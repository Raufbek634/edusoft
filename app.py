from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, has_request_context, send_file

import json
import os
import shutil
import io

import re
import uuid
import threading
import html
import time
from datetime import datetime, date, timedelta, timezone
import audit_log
from functools import wraps
import hashlib
import requests as http_requests
import calendar

import qrcode
from io import BytesIO

from platform_core import PlatformCtx, PLATFORM_FILES
from validation import validate_phone, validate_name, validate_amount, validate_date, validate_month, sanitize_html

# Lazy flags — heavy packages imported inside functions to speed up cold start
HAS_OPENPYXL = None
HAS_REPORTLAB = None

ROOT = os.path.dirname(os.path.abspath(__file__))

# Load .env file if present (local dev)
dotenv_path = os.path.join(ROOT, '.env')
if os.path.exists(dotenv_path):
    with open(dotenv_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#'):
                k, _, v = line.partition('=')
                os.environ.setdefault(k.strip(), v.strip())

app = Flask(__name__, static_folder=os.path.join(ROOT, 'static'))
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    stable = os.environ.get('SITE_URL') or os.environ.get('VERCEL_URL') or ''
    if stable:
        app.secret_key = hashlib.sha256(stable.encode()).hexdigest()
        print(f"[✓] SECRET_KEY derived from SITE_URL/VERCEL_URL")
    else:
        import secrets
        app.secret_key = secrets.token_hex(32)
        print("WARNING: SECRET_KEY env var not set, using auto-generated key")
app.config['SESSION_COOKIE_SAMESITE'] = 'Strict'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('VERCEL'))
app.config['SESSION_COOKIE_NAME'] = '__Secure-session' if os.environ.get('VERCEL') else 'session'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=60)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# PWA routes
@app.route('/manifest.json')
def pwa_manifest():
    return send_from_directory(os.path.join(ROOT, 'static'), 'manifest.json')

@app.route('/sw.js')
def pwa_sw():
    resp = send_from_directory(os.path.join(ROOT, 'static'), 'sw.js')
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-store'
    return resp

@app.route('/robots.txt')
def robots_txt():
    return send_from_directory(ROOT, 'robots.txt')

@app.route('/sitemap.xml')
def sitemap_xml():
    return send_from_directory(ROOT, 'sitemap.xml')

if os.environ.get('VERCEL'):
    app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
else:
    app.config['UPLOAD_FOLDER'] = os.path.join(ROOT, 'static/uploads')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB

DATA_DIR = os.environ.get('DATA_DIR', os.path.join(ROOT, 'data'))
if os.environ.get('VERCEL'):
    src = os.path.join(ROOT, 'data')
    dst = '/tmp/data'
    if os.path.exists(src) and not os.path.exists(dst):
        try:
            shutil.copytree(src, dst)
            print(f"[Vercel] Data copied to {dst}")
        except Exception as ex:
            print(f"[Vercel] Copy error: {ex}")
    DATA_DIR = dst
    if not os.environ.get('DATABASE_URL'):
        print(f"[Vercel] ⚠️  DATABASE_URL not set - data will NOT persist across instances!")
        print("[Vercel]    Set DATABASE_URL env var (Neon: https://neon.tech, free tier)")
    print(f"[Vercel] DATA_DIR={DATA_DIR} exists={os.path.isdir(DATA_DIR)}")
pc = PlatformCtx(DATA_DIR)

# ─── Visit tracking for daily stats ──────────────────────────────────────────
VISIT_LOG_FILE = 'visit_logs.json'

def esc_html(text):
    return html.escape(str(text), quote=False)

def _track_visit(ip, path):
    """Lightweight page visit counter per hour per day."""
    if not ip:
        return
    today = datetime.now().strftime('%Y-%m-%d')
    hour = datetime.now().strftime('%H')
    visits = pc.load_json(VISIT_LOG_FILE)
    if not isinstance(visits, dict):
        visits = {}
    day = visits.setdefault(today, {'views': 0, 'ips': {}, 'hours': {}})
    day['views'] = day.get('views', 0) + 1
    ips = day['ips']
    ips[ip] = ips.get(ip, 0) + 1
    hr = day['hours'].setdefault(hour, {'views': 0, 'ips': {}})
    hr['views'] = hr.get('views', 0) + 1
    h_ips = hr['ips']
    h_ips[ip] = h_ips.get(ip, 0) + 1
    # Trim old entries (keep 7 days)
    keys = sorted(visits.keys(), reverse=True)
    for k in keys[7:]:
        del visits[k]
    pc.save_json(VISIT_LOG_FILE, visits)

# ─── No-cache headers for all HTML responses ──────────────────────────────────
@app.after_request
def add_no_cache(resp):
    if resp.content_type and resp.content_type.startswith('text/html'):
        resp.headers['Cache-Control'] = 'no-store, must-revalidate'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
    # ─── Security headers (Z12) ───────────────────────────────────────────────
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'DENY')
    resp.headers.setdefault('X-XSS-Protection', '1; mode=block')
    resp.headers.setdefault('Strict-Transport-Security', 'max-age=63072000; includeSubDomains; preload')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    resp.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=(self)')
    if resp.content_type and resp.content_type.startswith('text/html'):
        csp = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://hcaptcha.com https://*.hcaptcha.com https://cdnjs.cloudflare.com; "
            "style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
            "img-src 'self' data: blob: https:; "
            "font-src 'self' https://cdnjs.cloudflare.com https://fonts.gstatic.com; "
            "frame-src https://hcaptcha.com https://*.hcaptcha.com; "
            "connect-src 'self' https://api.telegram.org; "
            "frame-ancestors 'none'"
        )
        resp.headers.setdefault('Content-Security-Policy', csp)
    # ─── Track page views (HTML only, skip API/static) ─────────────────────
    if resp.content_type and resp.content_type.startswith('text/html'):
        path = request.path
        if not path.startswith(('/static/', '/api/', '/receipt/')):
            try:
                ip = request.remote_addr or ''
                _track_visit(ip, path)
            except Exception:
                pass
    return resp

# ─── Session inactivity timeout (Z3) ──────────────────────────────────────────
SESSION_INACTIVITY_TIMEOUT = 55 * 60  # 55 minutes (warn at 55, expire at 60)
SESSION_MAX_AGE = 60 * 60  # 60 minutes absolute

@app.before_request
def check_session_inactivity():
    if request.path.startswith('/static/') or request.path == '/api/keepalive':
        return None
    # Maintenance mode check — faqat HTML sahifalarni bloklaydi, API ga ruxsat
    if not request.path.startswith('/api/') and not request.path.startswith('/static/') and request.path != '/login':
        try:
            plat = pc.load_platform()
            ms = plat.get('platform_settings', {})
            if ms.get('maintenance_mode') and not session.get('role'):
                return render_template('maintenance.html', message=ms.get('maintenance_message', 'Platformada texnik ishlar olib borilmoqda.'))
        except Exception:
            pass
    if session.get('admin_id') or session.get('role'):
        now = time.time()
        last_activity = session.get('_last_activity')
        if last_activity is not None and isinstance(last_activity, (int, float)):
            elapsed = now - last_activity
            if elapsed > SESSION_MAX_AGE:
                session.clear()
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'success': False, 'message': 'Sessiya vaqti tugadi. Qayta kiring.'}), 401
                return redirect(url_for('login'))
            session['_inactivity_warning'] = elapsed > SESSION_INACTIVITY_TIMEOUT
        else:
            if last_activity is not None:
                session.pop('_last_activity', None)
            session['_inactivity_warning'] = False
        session['_last_activity'] = now

@app.route('/api/keepalive')
def keepalive():
    if session.get('role'):
        session['_last_activity'] = time.time()
        session['_inactivity_warning'] = False
        return jsonify({'ok': True})
    return jsonify({'ok': False}), 401

# ─── Translations ─────────────────────────────────────────────────────────────
TRANS_PATH = os.path.join(DATA_DIR, 'translations.json')
_translations = {}
if os.path.exists(TRANS_PATH):
    with open(TRANS_PATH, 'r', encoding='utf-8') as _f:
        _translations = json.load(_f)

def _(key, lang='uz'):
    return _translations.get(lang, {}).get(key, _translations.get('uz', {}).get(key, key))

def _tr(key, lang=None):
    if lang is None:
        lang = session.get('lang', 'uz')
    return _(key, lang)

# ─── Theme ────────────────────────────────────────────────────────────────────
DEFAULT_THEME = 'dark'

# ─── Jinja2 globals ───────────────────────────────────────────────────────────
app.jinja_env.globals['_'] = _tr

def safe_int(val, default=0):
    try:
        return int(float(val)) if val is not None else default
    except (ValueError, TypeError):
        return default
def _generate_qr_image(receipt_id, base_url):
    """Generate QR PIL Image for a receipt URL."""
    receipt_url = f"{base_url}/receipt/{receipt_id}"
    return qrcode.make(receipt_url, box_size=6, border=2)

def generate_receipt_qr_bytes(receipt_id, base_url):
    """Generate QR code PNG bytes for a receipt."""
    img = _generate_qr_image(receipt_id, base_url)
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf

# ─── Brute Force Protection ──────────────────────────────────────────────────

ATTEMPT_LIMITS = [(5, 15), (10, 60), (20, -1)]  # (attempts, block_minutes) -1=permanent

def _get_ip():
    xff = request.headers.get('X-Forwarded-For', '')
    return xff.split(',')[0].strip() if xff else request.remote_addr or 'unknown'

def _rate_limit_key(ip):
    return f'rate_limit:{ip}'

def get_rate_limit(ip):
    data = pc.db.get(_rate_limit_key(ip)) if pc.db else {}
    return data or {}

def save_rate_limit(ip, data):
    if pc.db:
        pc.db.set(_rate_limit_key(ip), data)

def record_failed_login(ip):
    now = datetime.now(timezone.utc).timestamp()
    data = get_rate_limit(ip)
    if not data:
        data = {'count': 0, 'attempts': []}
    data['count'] = data.get('count', 0) + 1
    data['attempts'] = data.get('attempts', [])
    data['attempts'].append(now)
    # Keep last 50 attempts max
    data['attempts'] = data['attempts'][-50:]
    data['last_fail'] = now

    cnt = data['count']
    # Check thresholds
    for threshold, block_min in ATTEMPT_LIMITS:
        if cnt >= threshold:
            if block_min == -1:
                data['permanent_block'] = True
                data['blocked_until'] = None
                _notify_brute_force(ip, cnt)
            else:
                data['blocked_until'] = now + block_min * 60
            break
    save_rate_limit(ip, data)
    return data

def reset_login_attempts(ip):
    if pc.db:
        pc.db.delete_key(_rate_limit_key(ip))

def check_rate_limit(ip):
    """Returns (blocked: bool, remaining_seconds: int, attempts: int, needs_captcha: bool)"""
    data = get_rate_limit(ip)
    now = datetime.now(timezone.utc).timestamp()
    cnt = data.get('count', 0)
    blocked_until = data.get('blocked_until')
    permanent = data.get('permanent_block', False)

    if permanent:
        return True, -1, cnt, False

    if blocked_until and now < blocked_until:
        remaining = int(blocked_until - now)
        return True, remaining, cnt, False

    # If block expired, clear it
    if blocked_until and now >= blocked_until:
        data.pop('blocked_until', None)
        save_rate_limit(ip, data)

    needs_captcha = cnt >= 3
    return False, 0, cnt, needs_captcha

def _notify_brute_force(ip, count):
    """Send admin notification about permanent block."""
    try:
        plat = pc.load_platform()
        super_token = plat.get('super_bot_token', '').strip()
        super_chat = plat.get('super_telegram_chat_id', '').strip()
        if super_token and super_chat:
            from telegram_bot import send_message as tg_send
            tg_send(super_token, super_chat,
                f'🚨 <b>Brute Force hujumi aniqlandi!</b>\n\n'
                f'IP: <code>{ip}</code>\n'
                f'Urinishlar: {count}\n'
                f'Vaqt: {datetime.now().strftime("%d.%m.%Y %H:%M")}\n\n'
                f'IP bloklandi.')
    except Exception:
        pass

app.jinja_env.globals['statusBadge'] = lambda s: (
    '<span class="badge badge-success">Faol</span>' if s == 'active' else
    '<span class="badge badge-gray">Nofaol</span>' if s == 'inactive' else
    '<span class="badge badge-success">Keldi</span>' if s == 'present' else
    '<span class="badge badge-danger">Kelmadi</span>' if s == 'absent' else
    '<span class="badge badge-warning">Sababli</span>' if s == 'excused' else
    '<span class="badge badge-success">Faol</span>'
)

# ─── Data helpers (ko'p bog'cha) ───────────────────────────────────────────────

def _current_kg_id():
    if has_request_context():
        if session.get('role') == 'super':
            return None
        return session.get('kindergarten_id', 'default')
    return 'default'

def load_json(filename, kg_id=None):
    if filename in PLATFORM_FILES:
        return pc.load_json(filename)
    kid = kg_id if kg_id is not None else _current_kg_id()
    return pc.load_json(filename, kid or 'default')

def save_json(filename, data, kg_id=None):
    if filename in PLATFORM_FILES:
        pc.save_json(filename, data)
        return
    kid = kg_id if kg_id is not None else _current_kg_id()
    pc.save_json(filename, data, kid or 'default')

def load_settings(kg_id=None):
    kid = kg_id if kg_id is not None else _current_kg_id()
    data = pc.load_settings(kid or 'default')
    if isinstance(data, dict):
        if os.environ.get('BOT_TOKEN'):
            data['bot_token'] = os.environ['BOT_TOKEN']
        if os.environ.get('ADMIN_TELEGRAM_CHAT_ID'):
            data['admin_telegram_chat_id'] = os.environ['ADMIN_TELEGRAM_CHAT_ID']
    return data


def _audit(action, details=''):
    admin_id = session.get('admin_id', session.get('user_id', ''))
    admin_name = session.get('admin_name', session.get('login', ''))
    kg_id = session.get('kindergarten_id', '')
    ip = request.remote_addr or ''
    entry = audit_log.log(pc, action, details, str(admin_id), admin_name, kg_id or '', ip)
    try:
        if audit_log.check_suspicious(pc, action, details):
            plat = pc.load_platform()
            st = plat.get('super_bot_token', '').strip()
            sc = plat.get('super_telegram_chat_id', '').strip()
            if st and sc:
                from telegram_bot import send_message as tg_send
                tg_send(st, sc,
                    f"🚨 <b>Shubhali harakat!</b>\n\n"
                    f"Amal: {action}\n"
                    f"Admin: {admin_name} ({admin_id})\n"
                    f"Tafsilot: {details}\n"
                    f"IP: {ip}\n"
                    f"Vaqt: {entry.get('timestamp', '')[:19]}")
    except Exception:
        pass
    return entry

def save_settings_data(settings, kg_id=None):
    kid = kg_id if kg_id is not None else _current_kg_id()
    pc.save_settings(kid or 'default', settings)

def normalize_phone(phone):
    """Telefon raqamni solishtirish uchun standartlashtirish."""
    if not phone:
        return ''
    p = re.sub(r'\D', '', str(phone).strip())
    if p.startswith('998') and len(p) >= 12:
        return p[:12]
    if len(p) == 9 and p[0] in '9':
        return '998' + p
    if len(p) == 12:
        return p
    return p

def find_student_by_phone(phone, kg_id=None):
    norm = normalize_phone(phone)
    if not norm:
        return None, None
    tail = norm[-9:] if len(norm) >= 9 else norm
    kindergartens = [kg for kg in pc.load_kindergartens() if kg.get('status') == 'active']
    if kg_id:
        kindergartens = [kg for kg in kindergartens if kg['id'] == kg_id]
    for kg in kindergartens:
        students = pc.load_json('students.json', kg['id'])
        for s in students:
            sp = normalize_phone(s.get('parent_phone', ''))
            if not sp:
                continue
            if sp == norm or (tail and sp.endswith(tail)) or (tail and norm.endswith(sp[-9:])):
                return s, kg['id']
    return None, None

def is_valid_admin_chat_id(chat_id):
    cid = str(chat_id or '').strip()
    if not cid:
        return False
    return cid.lstrip('-').isdigit() and len(cid) >= 8

def send_telegram_message(chat_id, text, parse_mode='HTML', kg_id=None):
    settings = load_settings(kg_id)
    token = settings.get('bot_token', '')
    if not token or not chat_id:
        return False
    try:
        from telegram_bot import send_message
        return send_message(token, chat_id, text, parse_mode=parse_mode)
    except Exception:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        try:
            resp = http_requests.post(
                url,
                json={'chat_id': chat_id, 'text': text, 'parse_mode': parse_mode},
                timeout=8
            )
            return resp.status_code == 200
        except Exception:
            return False

def notify_super_admin(message, notif_type='info'):
    """Platforma egasi Raufbek — yangi bog'cha arizalari va premium."""
    supers = pc.load_json('super_admins.json')
    for sa in supers:
        add_notification(f"[PLATFORM] {message}", notif_type, target='platform')
    plat = pc.load_platform()
    super_chat = plat.get('super_telegram_chat_id', '')
    super_token = plat.get('super_bot_token', '').strip()
    if is_valid_admin_chat_id(super_chat) and super_token:
        try:
            from telegram_bot import send_message
            send_message(super_token, super_chat, f"🏢 <b>Platforma (Raufbek)</b>\n\n{message}")
        except Exception:
            pass

BASE_URL = os.environ.get('SITE_URL', 'https://sofgardercrm.vercel.app').rstrip('/')

def notify_login(admin_name, role, ip, lat='', lng='', extra=''):
    plat = pc.load_platform()
    chat = plat.get('super_telegram_chat_id', '')
    token = plat.get('super_bot_token', '').strip()
    if not is_valid_admin_chat_id(chat) or not token:
        return
    now_s = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    lines = [
        f"🔔 <b>Yangi kirish</b>",
        f"",
        f"👤 <b>Admin:</b> {admin_name}",
        f"🎭 <b>Rol:</b> {role}",
        f"🌐 <b>IP:</b> <code>{ip}</code>",
    ]
    if lat and lng:
        lines.append(f"📍 <b>Lokatsiya:</b> {lat}, {lng}")
        lines.append(f"🗺 <a href='https://www.google.com/maps?q={lat},{lng}'>Xaritada ko'rish</a>")
    if extra:
        lines.append(f"📎 {extra}")
    lines.append(f"⏰ {now_s} UTC")
    try:
        from telegram_bot import send_message as tg_send
        tg_send(token, chat, '\n'.join(lines))
    except Exception:
        pass

def notify_admin(message, notif_type='info', kg_id=None):
    kid = kg_id or _current_kg_id() or 'default'
    add_notification(message, notif_type, kg_id=kid)
    settings = load_settings(kid)
    admin_chat = settings.get('admin_telegram_chat_id', '')
    kg = pc.get_kindergarten(kid)
    if kg:
        owner_chat = kg.get('owner', {}).get('telegram_chat_id', '')
        if is_valid_admin_chat_id(owner_chat):
            send_telegram_message(owner_chat, f"🔔 <b>Bog'cha xabari</b>\n\n{message}", kg_id=kid)
    if is_valid_admin_chat_id(admin_chat):
        send_telegram_message(admin_chat, f"🔔 <b>Admin xabari</b>\n\n{message}", kg_id=kid)

def hash_password(pw):
    from werkzeug.security import generate_password_hash
    return generate_password_hash(pw)

def check_password(plain, hashed):
    if not hashed or not plain:
        return False
    from werkzeug.security import check_password_hash
    try:
        if check_password_hash(hashed, plain):
            return True
    except (ValueError, TypeError):
        pass
    return hashlib.sha256(plain.encode()).hexdigest() == hashed

def init_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    pc.init_platform_data()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ('kg_admin', 'super', 'teacher'):
            return redirect(url_for('login', login=1))
        if session.get('role') == 'super' and not session.get('kindergarten_id'):
            return f(*args, **kwargs)
        if not session.get('kindergarten_id'):
            return redirect(url_for('login', login=1))
        return f(*args, **kwargs)
    return decorated

def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'teacher':
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': 'Faqat o\'qituvchilar uchun'}), 403
            return redirect(url_for('login', login=1))
        return f(*args, **kwargs)
    return decorated

def super_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'super':
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': 'Avtorizatsiyadan o\'tmagan'}), 401
            return redirect(url_for('login', login=1))
        # Z4 — IP whitelist check
        allowed_ips = os.environ.get('SUPER_ADMIN_IPS', '').strip()
        if allowed_ips:
            ip = _get_ip()
            if ip and ip not in [x.strip() for x in allowed_ips.split(',')]:
                _audit('super_admin_blocked_ip', f"Ruxsatsiz IP: {ip}")
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'success': False, 'message': 'Ruxsat etilmagan IP'}), 403
                return redirect(url_for('login', login=1))
        # Z5 — Blocked IPs from platform config
        try:
            ip = _get_ip()
            if ip:
                plat = pc.load_platform()
                blocked = plat.get('blocked_ips', [])
                if any(b.get('ip') == ip for b in blocked):
                    _audit('super_admin_blocked_ip', f"Bloklangan IP: {ip}")
                    return jsonify({'success': False, 'message': 'IP manzilingiz bloklangan'}), 403
        except Exception:
            pass
        return f(*args, **kwargs)
    return decorated

def get_admin():
    if session.get('role') == 'super':
        return {'id': session.get('admin_id'), 'name': session.get('admin_name'), 'role': 'super'}
    kg = pc.get_kindergarten(session.get('kindergarten_id', 'default'))
    if kg:
        return {'id': kg['id'], 'name': session.get('admin_name'), 'role': 'kg_admin', 'kindergarten': kg}
    return None

def send_parent_payment_reminders(kg_id=None):
    """Ota-onalarga to'lov yaqinlashgani haqida Telegram."""
    kids = [kg_id] if kg_id else [k['id'] for k in pc.load_kindergartens() if k.get('status') == 'active']
    today = date.today()
    today_str = today.isoformat()
    for kid in kids:
        settings = load_settings(kid)
        token = settings.get('bot_token', '')
        if not token:
            continue
        students = load_json('students.json', kid)
        log = load_json('reminder_log.json', kid)
        sent_today = {x.get('key') for x in log if x.get('date') == today_str}
        y, m = today.year, today.month
        for s in students:
            if s.get('status') != 'active' or not s.get('telegram_chat_id'):
                continue
            payable = calc_payable_amount(s, y, m, kid)
            paid = get_paid_amount(s['id'], y, m, kid)
            if paid >= payable:
                continue
            due_day = int(s.get('payment_due_day', 1) or 1)
            try:
                due_date = date(y, m, min(due_day, 28))
                if due_date < today:
                    due_date = date(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, min(due_day, 28))
            except Exception:
                continue
            days_left = (due_date - today).days
            if days_left not in (1, 2):
                continue
            key = f"{s['id']}-{due_date.isoformat()}-{days_left}"
            if key in sent_today:
                continue
            fname = s['first_name']
            lname = s['last_name']
            msg = (
                f"💳 <b>To'lov eslatmasi</b>\n\n"
                f"Bolangiz <b>{lname} {fname}</b>ning to'lov muddati "
                f"<b>{days_left} kun</b>dan keyin tugaydi ({due_date.strftime('%d.%m.%Y')}).\n\n"
                f"Iltimos, admin bilan bog'laning.\n"
                f"🏫 {settings.get('name', 'Bog\'cha')}"
            )
            if send_telegram_message(s['telegram_chat_id'], msg, kg_id=kid):
                log.append({'key': key, 'date': today_str, 'student_id': s['id'], 'days_left': days_left})
                sent_today.add(key)
        if len(log) > 500:
            log = log[-500:]
        save_json('reminder_log.json', log, kid)

# ─── Attendance helpers ────────────────────────────────────────────────────────

def get_working_days(year, month):
    """Return count of working (Mon–Fri) days in a month."""
    cal = calendar.monthcalendar(year, month)
    days = 0
    for week in cal:
        for i, d in enumerate(week):
            if d != 0 and i < 5:  # Mon=0 … Fri=4
                days += 1
    return days

def get_attended_days(student_id, year, month, kg_id=None):
    attendance = load_json('attendance.json', kg_id)
    count = 0
    for rec in attendance:
        if rec['student_id'] != student_id:
            continue
        try:
            d = datetime.strptime(rec['date'], '%Y-%m-%d')
            if d.year == year and d.month == month and rec['status'] == 'present':
                count += 1
        except Exception:
            pass
    return count

def get_absent_days(student_id, year, month, kg_id=None):
    attendance = load_json('attendance.json', kg_id)
    count = 0
    for rec in attendance:
        if rec['student_id'] != student_id:
            continue
        try:
            d = datetime.strptime(rec['date'], '%Y-%m-%d')
            if d.year == year and d.month == month and rec['status'] == 'absent':
                count += 1
        except Exception:
            pass
    return count

def calc_payable_amount(student, year, month, kg_id=None):
    working = get_working_days(year, month)
    absent = get_absent_days(student['id'], year, month, kg_id)
    payable_days = max(working - absent, 0)
    daily = student['monthly_fee'] / working if working else 0
    payable = round(daily * payable_days)
    discount = safe_int(student.get('discount_percent'), 0)
    if discount > 0:
        payable = round(payable * (100 - discount) / 100)
    subsidy = safe_int(student.get('subsidy_amount'), 0)
    if subsidy > 0:
        payable = max(payable - subsidy, 0)
    return payable

def get_paid_amount(student_id, year, month, kg_id=None):
    payments = load_json('payments.json', kg_id)
    total = 0
    prefix = f"{year}-{month:02d}"
    for p in payments:
        if p.get('student_id') != student_id:
            continue
        if p.get('month', '') == prefix and p.get('status', 'paid') != 'cancelled':
            total += p['amount']
    return total

# ─── Auth routes ──────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if session.get('role') == 'super':
        return redirect(url_for('super_dashboard'))
    if session.get('role') == 'kg_admin':
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# ─── I18n & Theme ──────────────────────────────────────────────────────────────
@app.route('/api/lang', methods=['POST'])
def api_set_lang():
    data = request.get_json() or {}
    lang = data.get('lang', 'uz')
    if lang not in ('uz', 'ru', 'en'):
        lang = 'uz'
    session['lang'] = lang
    return jsonify({'success': True, 'lang': lang})

@app.route('/api/theme', methods=['POST'])
def api_set_theme():
    data = request.get_json() or {}
    theme = data.get('theme', 'dark')
    if theme not in ('dark', 'light'):
        theme = 'dark'
    session['theme'] = theme
    return jsonify({'success': True, 'theme': theme})

@app.context_processor
def inject_theme_lang():
    kg_id = session.get('kindergarten_id')
    kg_balance = 0
    if kg_id:
        try:
            s = pc.load_settings(kg_id)
            kg_balance = s.get('balance', 0)
        except Exception:
            pass
    return {
        'current_lang': session.get('lang', 'uz'),
        'current_theme': session.get('theme', DEFAULT_THEME),
        'now': datetime.now,
        'kg_balance': kg_balance
    }

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json() or {}
        login_val = data.get('login', '').strip()
        password = data.get('password', '')
        remember = data.get('remember', False)
        code = data.get('code', '').strip()
        is_maxsus = data.get('maxsus', False)
        captcha_token = data.get('captcha_token', '')
        lat = data.get('lat', '')
        lng = data.get('lng', '')

        ip = _get_ip()

        # ─── Rate limit check ──────────────────────────────────────────────
        blocked, remaining, attempts, needs_captcha = check_rate_limit(ip)
        if blocked:
            msg = '🚫 IP manzilingiz vaqtincha bloklangan.'
            if remaining > 0:
                msg += f' {remaining // 60} daqiqa kuting.'
            elif remaining == -1:
                msg += ' Bu doimiy blok. Admin bilan bog\'laning: @mr_turaqulov'
            return jsonify({
                'success': False, 'message': msg,
                'blocked': True, 'remaining': remaining,
                'attempts': attempts
            })

        # ─── hCaptcha verification after 3rd failed attempt ─────────────────
        if needs_captcha and not is_maxsus:
            settings = load_settings()
            hcaptcha_secret = settings.get('hcaptcha_secret', '')
            if hcaptcha_secret and captcha_token:
                verify = http_requests.post(
                    'https://api.hcaptcha.com/siteverify',
                    data={'secret': hcaptcha_secret, 'response': captcha_token}
                ).json()
                if not verify.get('success'):
                    return jsonify({
                        'success': False, 'message': '❌ CAPTCHA ni tasdiqlang.',
                        'captcha_required': True, 'attempts': attempts
                    })
            elif hcaptcha_secret and not captcha_token:
                return jsonify({
                    'success': False, 'message': 'CAPTCHA talab qilinadi.',
                    'captcha_required': True, 'attempts': attempts
                })

        # Step 2: verify Telegram code (super admin 2FA)
        if code:
            expected = session.pop('super_telegram_code', None)
            if expected and code == expected:
                auth = session.get('_pending_auth')
                if not auth:
                    return jsonify({'success': False, 'message': 'Sessiya tugagan. Qaytadan kirish'})
                session.pop('_pending_auth', None)
                session.clear()
                session['admin_id'] = auth['id']
                session['admin_name'] = auth['name']
                session['role'] = auth['role']
                session['kindergarten_id'] = auth.get('kindergarten_id')
                if remember:
                    session.permanent = True
                    app.permanent_session_lifetime = timedelta(days=30)
                reset_login_attempts(ip)
                loc = f" | loc: {lat},{lng}" if lat and lng else ""
                _audit('login', f"Super admin code-login: {auth['name']}{loc}")
                notify_login(auth['name'], 'super', ip, lat, lng)
                return jsonify({
                    'success': True,
                    'role': auth['role'],
                    'redirect': '/super'
                })
            record_failed_login(ip)
            return jsonify({'success': False, 'message': '❌ Kod noto\'g\'ri'})

        # Step 1a: Maxsus — super admin Telegram code login
        if is_maxsus:
            # Z4 — Super Admin PIN check
            expected_pin = os.environ.get('SUPER_ADMIN_PIN', '').strip()
            login_pin = data.get('pin', '').strip()
            if expected_pin and login_pin != expected_pin:
                record_failed_login(ip)
                loc = f" | loc: {lat},{lng}" if lat and lng else ""
                _audit('failed_login', f"Super admin PIN xato, IP: {ip}{loc}")
                return jsonify({'success': False, 'message': 'PIN noto\'g\'ri'})
            plat = pc.load_platform()
            super_token = plat.get('super_bot_token', '').strip()
            super_chat = plat.get('super_telegram_chat_id', '').strip()
            if not super_token or not super_chat:
                return jsonify({'success': False, 'message': 'Super admin bot sozlanmagan'})
            import random
            telegram_code = ''.join(random.choices('0123456789', k=6))
            session['super_telegram_code'] = telegram_code
            session['_pending_auth'] = {
                'role': 'super',
                'id': 'super',
                'name': 'Raufbek Turaqulov',
                'kindergarten_id': None
            }
            from telegram_bot import send_message as tg_send
            try:
                tg_send(super_token, super_chat,
                    f'🔐 <b>EduSoft — Maxsus kirish</b>\n\n'
                    f'Platformaga kirish kodi:\n\n'
                    f'<b>{telegram_code}</b>\n\n'
                    f'Agar bu siz bo\'lmasangiz, xabarni e\'tiborsiz qoldiring.')
                sent = True
            except Exception:
                sent = False
            reset_login_attempts(ip)
            return jsonify({
                'success': True,
                'code_required': True,
                'message': 'Telegram kodi yuborildi ✅' if sent else '⚠️ Kod yuborilmadi',
                'chat_hint': super_chat[-4:] if super_chat else ''
            })

        # Step 1b: normal login/password auth
        auth = pc.authenticate(login_val, password)
        if not auth:
            data = record_failed_login(ip)
            cnt = data.get('count', 0)
            _, _, _, captcha_needed = check_rate_limit(ip)
            msg = 'Noto\'g\'ri login yoki parol'
            if cnt >= 3:
                msg += f' ({cnt} marta noto\'g\'ri urinish)'
            loc = f" | loc: {lat},{lng}" if lat and lng else ""
            _audit('failed_login', f"Noto'g'ri login: {login_val}, IP: {ip}{loc}")
            return jsonify({
                'success': False, 'message': msg,
                'attempts': cnt,
                'captcha_required': captcha_needed
            })

        role = auth.get('role', '')
        if role == 'blocked':
            return jsonify({'success': False, 'message': '⛔ Bog\'cha bloklangan. Admin bilan bog\'laning: @mr_turaqulov'})

        # Super admin login (without 2FA — maxsus is the secure path)
        if role == 'super':
            reset_login_attempts(ip)
            session.clear()
            session['admin_id'] = auth.get('id', '')
            session['admin_name'] = auth.get('name', 'Admin')
            session['role'] = role
            session['kindergarten_id'] = auth.get('kindergarten_id')
            if remember:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=30)
            loc = f" | loc: {lat},{lng}" if lat and lng else ""
            _audit('login', f"Super admin: {auth.get('name', '')}{loc}")
            notify_login(auth.get('name', ''), 'super', ip, lat, lng)
            return jsonify({
                'success': True,
                'role': role,
                'redirect': '/super'
            })

        # Normal kg_admin login
        reset_login_attempts(ip)
        session.clear()
        session['admin_id'] = auth.get('id', '')
        session['admin_name'] = auth.get('name', 'Admin')
        session['role'] = role
        session['kindergarten_id'] = auth.get('kindergarten_id')
        if remember:
            session.permanent = True
            app.permanent_session_lifetime = timedelta(days=30)
        loc = f" | loc: {lat},{lng}" if lat and lng else ""
        _audit('login', f"Admin: {auth['name']} (bog'cha: {auth.get('kindergarten_id', 'default')}){loc}")
        notify_login(auth['name'], 'bog\'cha admini', ip, lat, lng, f"🏫 Bog'cha: {auth.get('kindergarten_id', 'default')}")
        return jsonify({
            'success': True,
            'role': auth['role'],
            'redirect': '/dashboard'
        })
    plans = pc.load_platform().get('plans', [])
    settings = load_settings()
    hcaptcha_sitekey = settings.get('hcaptcha_sitekey', '')
    return render_template('login.html', plans=plans, hcaptcha_sitekey=hcaptcha_sitekey)

# ─── Forgot Password / Reset ─────────────────────────────────────────────────

def _find_admin_by_phone(phone):
    """Find admin by phone/login. Returns (type, data) or None."""
    for sa in pc.load_json('super_admins.json'):
        if sa.get('login') == phone:
            return ('super', sa, None)
    for kg in pc.load_kindergartens():
        owner = kg.get('owner', {})
        if owner.get('login') == phone:
            return ('kg_admin', owner, kg.get('id'))
    for a in pc.load_json('admins.json', 'default'):
        if a.get('login') == phone:
            return ('legacy', a, 'default')
    return None

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json() or {}
    step = data.get('step', '')

    # Step 1: Send OTP
    if step == 'send_otp':
        phone = data.get('phone', '').strip()
        if not phone:
            return jsonify({'success': False, 'message': 'Telefon raqamni kiriting'})

        result = _find_admin_by_phone(phone)
        if not result:
            return jsonify({'success': False, 'message': 'Bu raqam tizimda topilmadi'})

        admin_type, admin_data, kg_id = result
        telegram_chat = admin_data.get('telegram_chat_id', '')

        import random
        otp = ''.join(random.choices('0123456789', k=6))
        session['reset_otp'] = otp
        session['reset_phone'] = phone
        session['reset_otp_expiry'] = (datetime.now(timezone.utc) + timedelta(minutes=5)).timestamp()
        session['reset_admin_type'] = admin_type
        session['reset_admin_data'] = admin_data
        session['reset_kg_id'] = kg_id

        # Send OTP via Telegram
        if telegram_chat:
            plat = pc.load_platform()
            super_token = plat.get('super_bot_token', '').strip()
            if super_token:
                from telegram_bot import send_message as tg_send
                try:
                    tg_send(super_token, telegram_chat,
                        f'🔐 <b>Parolni tiklash kodi</b>\n\n'
                        f'Parolingizni tiklash uchun kod:\n\n'
                        f'<b>{otp}</b>\n\n'
                        f'Kod 5 daqiqa amal qiladi.\n'
                        f'Agar siz so\'ramagan bo\'lsangiz, xabarni e\'tiborsiz qoldiring.')
                except Exception:
                    pass

        # Also send to super admin's own chat as fallback
        plat = pc.load_platform()
        super_token = plat.get('super_bot_token', '').strip()
        super_chat = plat.get('super_telegram_chat_id', '').strip()
        if super_token and super_chat and super_chat != telegram_chat:
            from telegram_bot import send_message as tg_send
            try:
                tg_send(super_token, super_chat,
                    f'🔐 <b>Parolni tiklash</b>\n\n'
                    f'Telefon: {phone}\n'
                    f'Admin: {admin_data.get("name", "Noma\'lum")}\n'
                    f'Kod: {otp}\n'
                    f'Muddat: 5 daqiqa')
            except Exception:
                pass

        return jsonify({
            'success': True,
            'message': 'Kod Telegram orqali yuborildi ✅' if telegram_chat else '⚠️ Admin Telegrami ulanmagan. Super adminga xabar yuborildi.',
            'otp_sent': True
        })

    # Step 2: Verify OTP
    if step == 'verify_otp':
        otp = data.get('otp', '').strip()
        expected = session.get('reset_otp')
        expiry = session.get('reset_otp_expiry', 0)

        if not expected or not otp:
            return jsonify({'success': False, 'message': 'Kodni kiriting'})

        if datetime.now(timezone.utc).timestamp() > expiry:
            session.pop('reset_otp', None)
            session.pop('reset_otp_expiry', None)
            return jsonify({'success': False, 'message': 'Kod muddati tugagan. Qaytadan yuboring.'})

        if otp != expected:
            return jsonify({'success': False, 'message': '❌ Noto\'g\'ri kod'})

        session['reset_otp_verified'] = True
        return jsonify({'success': True, 'message': 'Kod tasdiqlandi ✅', 'verified': True})

    # Step 3: Reset password
    if step == 'reset_password':
        if not session.get('reset_otp_verified'):
            return jsonify({'success': False, 'message': 'Avval kodni tasdiqlang'})

        new_password = data.get('new_password', '')
        if not new_password:
            return jsonify({'success': False, 'message': 'Yangi parolni kiriting'})

        # Password validation
        errors = []
        if len(new_password) < 8:
            errors.append('Kamida 8 ta belgi')
        if not re.search(r'[A-Z]', new_password):
            errors.append('Kamida 1 ta katta harf')
        if not re.search(r'[0-9]', new_password):
            errors.append('Kamida 1 ta raqam')
        if errors:
            return jsonify({'success': False, 'message': 'Talablar: ' + ', '.join(errors)})

        admin_type = session.get('reset_admin_type')
        admin_data = session.get('reset_admin_data')
        kg_id = session.get('reset_kg_id')
        phone = session.get('reset_phone')

        hashed = hash_password(new_password)

        if admin_type == 'super':
            supers = pc.load_json('super_admins.json')
            for s in supers:
                if s.get('id') == admin_data.get('id'):
                    s['password'] = hashed
                    break
            pc.save_json('super_admins.json', supers)

        elif admin_type == 'kg_admin':
            kgs = pc.load_kindergartens()
            for kg in kgs:
                owner = kg.get('owner', {})
                if owner.get('login') == phone:
                    kg['owner']['password'] = hashed
                    break
            pc.save_kindergartens(kgs)

        elif admin_type == 'legacy':
            legacy = pc.load_json('admins.json', 'default')
            for a in legacy:
                if a.get('login') == phone:
                    a['password'] = hashed
                    break
            pc.save_json('admins.json', legacy, 'default')

        # Clear session data
        for key in ['reset_otp', 'reset_otp_expiry', 'reset_otp_verified',
                     'reset_phone', 'reset_admin_type', 'reset_admin_data', 'reset_kg_id']:
            session.pop(key, None)

        # Invalidate all sessions by changing secret key (partial - clear session)
        session.clear()

        # Notify via Telegram
        admin_name = admin_data.get('name', 'Admin')
        plat = pc.load_platform()
        super_token = plat.get('super_bot_token', '').strip()
        super_chat = plat.get('super_telegram_chat_id', '').strip()
        if super_token and super_chat:
            from telegram_bot import send_message as tg_send
            try:
                tg_send(super_token, super_chat,
                    f'🔑 <b>Parol o\'zgartirildi</b>\n\n'
                    f'Admin: {admin_name}\n'
                    f'Telefon: {phone}\n'
                    f'Vaqt: {datetime.now().strftime("%d.%m.%Y %H:%M")}\n\n'
                    f'Agar admin tomonidan amalga oshirilmagan bo\'lsa, tekshiring!')
            except Exception:
                pass

        return jsonify({
            'success': True,
            'message': '✅ Parolingiz o\'zgartirildi. Agar siz qilmagan bo\'lsangiz, darhol @mr_turaqulov ga bog\'laning.'
        })

    return jsonify({'success': False, 'message': 'Noto\'g\'ri so\'rov'})

@app.route('/logout')
def logout():
    _audit('logout', f"Chiqish: {session.get('admin_name', '')}")
    session.clear()
    return redirect(url_for('login'))

# ─── Z4 — Honeypot routes ──────────────────────────────────────────────────────
HONEYPOT_IPS = set()

@app.route('/admin')
@app.route('/administrator')
@app.route('/wp-admin')
@app.route('/dashboard-old')
def honeypot_trap():
    ip = _get_ip()
    HONEYPOT_IPS.add(ip)
    _audit('honeypot_triggered', f"Honeypotga kirish: {request.path}, IP: {ip}")
    # If it's a scanner/bot, silently return 404
    return render_template('404.html'), 404

# ─── Contact form (server-side, no hardcoded tokens) ────────────────────────────
@app.route('/api/contact', methods=['POST'])
def contact_form():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()
    msg = data.get('msg', '').strip()
    if not name or not msg:
        return jsonify({'success': False, 'message': 'Ism va xabar matnini kiriting'})
    plat = pc.load_platform()
    super_token = plat.get('super_bot_token', '').strip()
    super_chat = plat.get('super_telegram_chat_id', '').strip()
    if not super_token or not super_chat:
        return jsonify({'success': False, 'message': 'Xabar yuborish imkoniyati hozircha mavjud emas'})
    text = f"📬 Bog'lanmoqchi: {name}"
    if email: text += f"\nEmail: {email}"
    if phone: text += f"\nTel: {phone}"
    text += f"\n\n{msg}"
    try:
        from telegram_bot import send_message as tg_send
        tg_send(super_token, super_chat, text)
        return jsonify({'success': True})
    except Exception:
        return jsonify({'success': False, 'message': 'Xatolik yuz berdi'})

# ─── Dashboard ────────────────────────────────────────────────────────────────

@app.route('/api/platform/alerts')
@login_required
def platform_alerts():
    kg_id = session.get('kindergarten_id', 'default')
    kg = pc.get_kindergarten(kg_id)
    was_active = kg.get('status') == 'active' if kg else False
    alerts = pc.get_kg_alerts(kg_id, calc_payable_amount, get_paid_amount, get_absent_days) if kg else []
    sub = pc.subscription_status(kg) if kg else {}
    # Check if just auto-blocked
    if was_active and sub.get('phase') == 'blocked':
        settings = pc.load_settings(kg_id)
        token = settings.get('bot_token', '')
        admin_chat = settings.get('admin_telegram_chat_id', '')
        if token and admin_chat:
            try:
                from telegram_bot import send_message as tg_send
                tg_send(token, admin_chat,
                    '⛔ <b>Bog\'cha bloklandi</b>\n\n'
                    f'To\'lov muddati o\'tganligi sababli bog\'changiz bloklandi.\n'
                    'Platforma egasi bilan bog\'lanib, to\'lovni amalga oshiring.')
            except Exception:
                pass
    return jsonify({'alerts': alerts, 'subscription': sub})

@app.route('/api/platform/run-reminders', methods=['POST'])
@login_required
def run_reminders_api():
    send_parent_payment_reminders(session.get('kindergarten_id'))
    return jsonify({'success': True, 'message': 'Eslatmalar yuborildi'})

@app.route('/api/dashboard/stats')
@login_required
def dashboard_stats_api():
    kg_id = _current_kg_id() or 'default'
    students = load_json('students.json')
    payments = load_json('payments.json')
    today = date.today()
    this_month = today.strftime('%Y-%m')
    # Active students
    active = sum(1 for s in students if s.get('status') == 'active')
    # Month payments total
    month_payments = [p for p in payments if p.get('month') == this_month and p.get('status') == 'paid']
    month_total = sum(p.get('amount', 0) for p in month_payments)
    # Month count
    month_count = len(month_payments)
    # Debtors (students without payment this month)
    paid_ids = {p['student_id'] for p in month_payments}
    debtors = [s for s in students if s.get('status', 'active') == 'active' and s['id'] not in paid_ids]
    # This week attendance
    attendance = load_json('attendance.json')
    week_start = (today - timedelta(days=today.weekday())).isoformat()
    week_att = [a for a in attendance if a.get('date', '') >= week_start]
    present = sum(1 for a in week_att if a.get('status') == 'present')
    absent = sum(1 for a in week_att if a.get('status') == 'absent')
    excused = sum(1 for a in week_att if a.get('status') == 'excused')
    # Monthly trend (last 6 months)
    trend = []
    for i in range(5, -1, -1):
        m = (today.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
        today = today.replace(day=1)
        m_pays = [p for p in payments if p.get('month') == m and p.get('status') == 'paid']
        trend.append({'month': m, 'total': sum(p.get('amount', 0) for p in m_pays), 'count': len(m_pays)})
        today = today - timedelta(days=1)
    today = date.today()
    return jsonify({
        'active_students': active,
        'month_total': month_total,
        'month_count': month_count,
        'debtors': len(debtors),
        'attendance': {'present': present, 'absent': absent, 'excused': excused, 'total': present + absent + excused},
        'trend': trend
    })

@app.route('/dashboard')
@login_required
def dashboard():
    try:
        settings = load_settings()
        send_parent_payment_reminders(session.get('kindergarten_id'))
        kg = pc.get_kindergarten(session.get('kindergarten_id', 'default'))
        subscription = pc.subscription_status(kg) if kg else {}
        return render_template('dashboard.html', settings=settings, admin_name=session.get('admin_name'),
                               subscription=subscription)
    except Exception as ex:
        import traceback
        return f'500 Dashboard error: {ex}\n{traceback.format_exc()}', 500

# ─── Super admin (Raufbek — platforma egasi) ───────────────────────────────────

@app.route('/super')
@super_required
def super_dashboard():
    plat = pc.load_platform()
    kgs = pc.load_kindergartens()
    apps = [a for a in load_json('kindergarten_applications.json') if a.get('status') == 'pending']
    kg_student_counts = {}
    for kg in kgs:
        kg_student_counts[kg['id']] = len(pc.load_json('students.json', kg['id']))
    for kg in kgs:
        kg['student_count'] = kg_student_counts.get(kg['id'], 0)
    return render_template('super_dashboard.html', platform=plat, kindergartens=kgs,
                           pending_apps=len(apps), admin_name=session.get('admin_name'))

@app.route('/super/kindergartens')
@super_required
def super_kindergartens():
    kgs = pc.load_kindergartens()
    for kg in kgs:
        s = pc.load_settings(kg['id'])
        kg['_balance'] = s.get('balance', 0)
    return render_template('super_kindergartens.html', kindergartens=kgs,
                           admin_name=session.get('admin_name'))

@app.route('/super/notifications')
@super_required
def super_notifications():
    kgs = pc.load_kindergartens()
    all_notifs = []
    for kg in kgs:
        if kg.get('status') != 'active':
            continue
        try:
            notifs = pc.load_json('notifications.json', kg['id'])
            for n in notifs:
                n['kg_name'] = kg['name']
                n['kg_id'] = kg['id']
            all_notifs.extend(notifs)
        except Exception:
            pass
    # Also load platform notifications
    try:
        plat_notifs = pc.load_json('notifications.json')
        if isinstance(plat_notifs, list):
            for n in plat_notifs:
                n['kg_name'] = 'Platforma'
                n['kg_id'] = 'platform'
            all_notifs.extend(plat_notifs)
    except Exception:
        pass
    all_notifs.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    all_notifs = all_notifs[:200]
    return render_template('super_notifications.html', notifications=all_notifs, kindergartens=kgs, admin_name=session.get('admin_name'))

@app.route('/super/applications')
@super_required
def super_applications():
    apps = load_json('kindergarten_applications.json')
    pending = sorted([a for a in apps if a.get('status') == 'pending'],
                     key=lambda x: x.get('created_at', ''), reverse=True)
    return render_template('super_applications.html', applications=pending,
                           admin_name=session.get('admin_name'))

@app.route('/super/audit-logs')
@super_required
def super_audit_logs():
    logs = audit_log.get_logs(pc, limit=200)
    return render_template('super_audit_logs.html', logs=logs,
                           admin_name=session.get('admin_name'))

@app.route('/api/super/audit-logs')
@super_required
def super_audit_logs_api():
    kg_id = request.args.get('kg_id', '')
    action = request.args.get('action', '')
    limit = safe_int(request.args.get('limit'), 100)
    logs = audit_log.get_logs(pc, limit=limit)
    if kg_id:
        logs = [l for l in logs if kg_id in l.get('details', '')]
    if action:
        logs = [l for l in logs if l.get('action') == action]
    return jsonify({'logs': logs, 'count': len(logs)})

@app.route('/api/super/audit-verify')
@super_required
def super_audit_verify():
    valid, idx = audit_log.verify_chain(pc)
    return jsonify({'verified': valid})

# ─── #9 — Super Admin login history ────────────────────────────────────────────
@app.route('/super/login-history')
@super_required
def super_login_history():
    return render_template('super_login_history.html', admin_name=session.get('admin_name'))

@app.route('/api/super/login-history')
@super_required
def super_login_history_api():
    days = safe_int(request.args.get('days'), 30)
    logs = audit_log.get_logs(pc, days=days, limit=200)
    login_logs = [l for l in logs if l.get('action') in ('login', 'logout', 'failed_login', 'super_admin_blocked_ip')]
    # Parse location data from details
    for l in login_logs:
        details = l.get('details', '')
        if 'loc:' in details:
            try:
                loc_part = details.split('loc:')[1].strip().split()[0]
                parts = loc_part.split(',')
                if len(parts) == 2:
                    l['lat'] = parts[0]; l['lng'] = parts[1]
            except Exception:
                pass
    return jsonify({'logs': login_logs, 'count': len(login_logs)})

# ─── #10 — Platform settings ────────────────────────────────────────────────────
PLATFORM_SETTINGS_DEFAULTS = {
    'default_trial_days': 30,
    'default_plan': 'standard',
    'currency': 'UZS',
    'platform_name': 'EduSoft',
    'platform_logo': '',
    'maintenance_mode': False,
    'maintenance_message': 'Platformada texnik ishlar olib borilmoqda. Birozdan so\'ng qayta urinib ko\'ring.',
}

@app.route('/super/settings')
@super_required
def super_platform_settings():
    plat = pc.load_platform()
    settings = plat.get('platform_settings', {})
    for k, v in PLATFORM_SETTINGS_DEFAULTS.items():
        settings.setdefault(k, v)
    plans = plat.get('plans', [])
    return render_template('super_settings.html', settings=settings, plans=plans, admin_name=session.get('admin_name'))

@app.route('/super/suspicious')
@super_required
def super_suspicious_page():
    return render_template('super_suspicious.html', admin_name=session.get('admin_name'))

@app.route('/super/stats')
@super_required
def super_daily_stats():
    return render_template('super_stats.html', admin_name=session.get('admin_name'))

@app.route('/api/super/daily-stats')
@super_required
def super_daily_stats_api():
    today = datetime.now().strftime('%Y-%m-%d')
    visits_data = pc.load_json(VISIT_LOG_FILE)
    if not isinstance(visits_data, dict):
        visits_data = {}
    today_data = visits_data.get(today, {'views': 0, 'ips': {}, 'hours': {}})
    # Audit log stats for today
    all_logs = audit_log.get_logs(pc, days=1)
    today_logs = [l for l in all_logs if l.get('timestamp', '').startswith(today)]
    action_counts = {}
    login_count = 0
    failed_login_count = 0
    for l in today_logs:
        act = l.get('action', 'unknown')
        action_counts[act] = action_counts.get(act, 0) + 1
        if act == 'login':
            login_count += 1
        elif act == 'failed_login':
            failed_login_count += 1
    return jsonify({
        'date': today,
        'page_views': today_data.get('views', 0),
        'unique_ips': len(today_data.get('ips', {})),
        'total_actions': len(today_logs),
        'login_count': login_count,
        'failed_login_count': failed_login_count,
        'action_counts': action_counts,
        'ip_breakdown': [{'ip': k, 'count': v} for k, v in
                         sorted(today_data.get('ips', {}).items(), key=lambda x: -x[1])[:50]],
        'hours': [{'hour': h, 'views': d['views'], 'unique_ips': len(d.get('ips', {}))}
                   for h, d in sorted(today_data.get('hours', {}).items())]
    })

@app.route('/super/admin-activity')
@super_required
def super_admin_activity():
    return render_template('super_admin_activity.html', admin_name=session.get('admin_name'))

@app.route('/api/super/admin-activity')
@super_required
def super_admin_activity_api():
    days = request.args.get('days', '7')
    try: days = int(days)
    except: days = 7
    admin_id = request.args.get('admin_id', '').strip()
    action_filter = request.args.get('action', '').strip().lower()
    admin_filter = request.args.get('admin', '').strip().lower()
    logs = audit_log.get_logs(pc, days=days, limit=2000)
    if admin_id:
        logs = [l for l in logs if l.get('admin_id') == admin_id]
    if admin_filter:
        logs = [l for l in logs if admin_filter in (l.get('admin_name') or '').lower()]
    if action_filter:
        logs = [l for l in logs if l.get('action', '').lower().startswith(action_filter)]
    # Group by admin
    by_admin = {}
    for l in logs:
        aid = l.get('admin_id', 'unknown')
        aname = l.get('admin_name', 'Noma\'lum')
        if aid not in by_admin:
            by_admin[aid] = {'admin_id': aid, 'admin_name': aname, 'total': 0, 'actions': {}}
        by_admin[aid]['total'] += 1
        act = l.get('action', 'unknown')
        by_admin[aid]['actions'][act] = by_admin[aid]['actions'].get(act, 0) + 1
    # Build summary (same as by_admin but with recent_actions)
    summary = []
    for v in sorted(by_admin.values(), key=lambda x: -x['total']):
        recent = sorted(v['actions'].items(), key=lambda x: -x[1])[:5]
        summary.append({'admin_name': v['admin_name'], 'total': v['total'], 'recent_actions': [a[0] for a in recent]})
    return jsonify({
        'logs': logs,
        'count': len(logs),
        'summary': summary,
        'total': len(logs),
        'by_admin': [v for k, v in sorted(by_admin.items(), key=lambda x: -x[1]['total'])],
        'days': days
    })

# ─── Super admin management (CRUD) ───────────────────────────────────────────

@app.route('/super/admins')
@super_required
def super_manage_admins():
    admins = pc.load_json('super_admins.json')
    for a in admins:
        a.pop('password', None)
    return render_template('super_admins.html', admins=admins, admin_name=session.get('admin_name'))

@app.route('/api/super/admins', methods=['GET', 'POST'])
@super_required
def super_admins_api():
    if request.method == 'GET':
        admins = pc.load_json('super_admins.json')
        for a in admins:
            a.pop('password', None)
        return jsonify({'admins': admins})
    data = request.get_json() or {}
    login = data.get('login', '').strip()
    password = data.get('password', '').strip()
    name = data.get('name', '').strip()
    if not login or not password or not name:
        return jsonify({'success': False, 'message': 'Login, parol va ism majburiy'})
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
    permissions = data.get('permissions', [])
    admins = pc.load_json('super_admins.json')
    for a in admins:
        if a['login'] == login:
            return jsonify({'success': False, 'message': 'Bu login band'})
    new_admin = {
        'id': 'super-' + str(uuid.uuid4())[:8],
        'login': login,
        'password': hash_password(password),
        'name': name,
        'role': 'super_admin',
        'permissions': permissions,
        'created_by': session.get('admin_name', ''),
        'created_at': datetime.now(timezone.utc).isoformat() + 'Z'
    }
    admins.append(new_admin)
    pc.save_json('super_admins.json', admins)
    _audit('super_admin_created', f"Yangi super admin: {name} ({login})")
    new_admin.pop('password', None)
    return jsonify({'success': True, 'admin': new_admin})

@app.route('/api/super/admins/<admin_id>', methods=['PUT', 'DELETE'])
@super_required
def super_admin_detail(admin_id):
    admins = pc.load_json('super_admins.json')
    idx = next((i for i, a in enumerate(admins) if a['id'] == admin_id), None)
    if idx is None:
        return jsonify({'success': False, 'message': 'Admin topilmadi'}), 404
    if request.method == 'DELETE':
        admin = admins.pop(idx)
        pc.save_json('super_admins.json', admins)
        _audit('super_admin_deleted', f"Super admin o'chirildi: {admin.get('name')} ({admin.get('login')})")
        return jsonify({'success': True})
    data = request.get_json() or {}
    admin = admins[idx]
    if 'name' in data:
        admin['name'] = data['name'].strip()
    if 'login' in data:
        admin['login'] = data['login'].strip()
    if data.get('password'):
        if len(data['password']) < 6:
            return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
        admin['password'] = hash_password(data['password'])
    if 'permissions' in data:
        admin['permissions'] = data['permissions']
    pc.save_json('super_admins.json', admins)
    _audit('super_admin_updated', f"Super admin yangilandi: {admin['name']}")
    resp = {k: v for k, v in admin.items() if k != 'password'}
    return jsonify({'success': True, 'admin': resp})

# ─── Kindergarten-level multiple admin management ────────────────────────────

@app.route('/api/kg-admins', methods=['GET', 'POST'])
@login_required
def kg_admins_api():
    kg_id = _current_kg_id() or 'default'
    settings = load_settings(kg_id)
    kg_admins = settings.get('additional_admins', [])
    if request.method == 'GET':
        safe = []
        for a in kg_admins:
            safe.append({k: v for k, v in a.items() if k != 'password'})
        return jsonify({'admins': safe})
    data = request.get_json() or {}
    login = data.get('login', '').strip()
    password = data.get('password', '').strip()
    name = data.get('name', '').strip()
    permissions = data.get('permissions', ['all'])
    if not login or not password or not name:
        return jsonify({'success': False, 'message': 'Login, parol va ism majburiy'})
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
    for a in kg_admins:
        if a['login'] == login:
            return jsonify({'success': False, 'message': 'Bu login band'})
    new_admin = {
        'id': 'adm-' + str(uuid.uuid4())[:8],
        'login': login,
        'password': hash_password(password),
        'name': name,
        'permissions': permissions,
        'created_at': datetime.now(timezone.utc).isoformat() + 'Z'
    }
    kg_admins.append(new_admin)
    settings['additional_admins'] = kg_admins
    save_settings_data(settings, kg_id)
    _audit('kg_admin_created', f"Yangi admin: {name} ({login})")
    new_admin.pop('password', None)
    return jsonify({'success': True, 'admin': new_admin})

@app.route('/api/kg-admins/<admin_id>', methods=['PUT', 'DELETE'])
@login_required
def kg_admin_detail(admin_id):
    kg_id = _current_kg_id() or 'default'
    settings = load_settings(kg_id)
    kg_admins = settings.get('additional_admins', [])
    idx = next((i for i, a in enumerate(kg_admins) if a['id'] == admin_id), None)
    if idx is None:
        return jsonify({'success': False, 'message': 'Admin topilmadi'}), 404
    if request.method == 'DELETE':
        admin = kg_admins.pop(idx)
        settings['additional_admins'] = kg_admins
        save_settings_data(settings, kg_id)
        _audit('kg_admin_deleted', f"Admin o'chirildi: {admin.get('name')} ({admin.get('login')})")
        return jsonify({'success': True})
    data = request.get_json() or {}
    admin = kg_admins[idx]
    if 'name' in data:
        admin['name'] = data['name'].strip()
    if 'login' in data:
        admin['login'] = data['login'].strip()
    if data.get('password'):
        if len(data['password']) < 6:
            return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
        admin['password'] = hash_password(data['password'])
    if 'permissions' in data:
        admin['permissions'] = data['permissions']
    settings['additional_admins'] = kg_admins
    save_settings_data(settings, kg_id)
    _audit('kg_admin_updated', f"Admin yangilandi: {admin['name']}")
    resp = {k: v for k, v in admin.items() if k != 'password'}
    return jsonify({'success': True, 'admin': resp})

@app.route('/api/check-login', methods=['POST'])
def api_check_login():
    """Check if a login is available (for admin creation forms)."""
    data = request.get_json() or {}
    login = data.get('login', '').strip()
    if not login:
        return jsonify({'available': False})
    # Check super_admins
    for a in pc.load_json('super_admins.json'):
        if a['login'] == login:
            return jsonify({'available': False})
    # Check kindergartens owners
    for kg in pc.load_kindergartens():
        owner = kg.get('owner', {})
        if owner.get('login') == login:
            return jsonify({'available': False})
        # Check additional admins
        settings = pc.load_settings(kg['id'])
        for aa in settings.get('additional_admins', []):
            if aa['login'] == login:
                return jsonify({'available': False})
    # Check legacy admins
    for a in pc.load_json('admins.json', 'default'):
        if a.get('login') == login:
            return jsonify({'available': False})
    return jsonify({'available': True})

@app.route('/api/super/admins/<admin_id>/password', methods=['POST'])
@super_required
def super_admin_reset_password(admin_id):
    admins = pc.load_json('super_admins.json')
    admin = next((a for a in admins if a['id'] == admin_id), None)
    if not admin:
        return jsonify({'success': False, 'message': 'Admin topilmadi'}), 404
    data = request.get_json() or {}
    new_pw = data.get('password', '').strip()
    if not new_pw or len(new_pw) < 6:
        return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
    admin['password'] = hash_password(new_pw)
    pc.save_json('super_admins.json', admins)
    _audit('super_admin_password_reset', f"Parol o'zgartirildi: {admin['name']}")
    return jsonify({'success': True})

@app.route('/api/kg-admins/<admin_id>/password', methods=['POST'])
@login_required
def kg_admin_reset_password(admin_id):
    kg_id = _current_kg_id() or 'default'
    settings = load_settings(kg_id)
    kg_admins = settings.get('additional_admins', [])
    admin = next((a for a in kg_admins if a['id'] == admin_id), None)
    if not admin:
        return jsonify({'success': False, 'message': 'Admin topilmadi'}), 404
    data = request.get_json() or {}
    new_pw = data.get('password', '').strip()
    if not new_pw or len(new_pw) < 6:
        return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
    admin['password'] = hash_password(new_pw)
    settings['additional_admins'] = kg_admins
    save_settings_data(settings, kg_id)
    _audit('kg_admin_password_reset', f"Parol o'zgartirildi: {admin['name']}")
    return jsonify({'success': True})

@app.route('/api/super/permissions')
@super_required
def super_permissions_list():
    """Return available permissions for super admins."""
    return jsonify({'permissions': [
        {'id': 'manage_admins', 'label': 'Adminlarni boshqarish'},
        {'id': 'manage_kindergartens', 'label': 'Bog\'chalarni boshqarish'},
        {'id': 'manage_applications', 'label': 'Arizalarni ko\'rib chiqish'},
        {'id': 'view_audit', 'label': 'Audit qaydlarini ko\'rish'},
        {'id': 'view_stats', 'label': 'Statistikani ko\'rish'},
        {'id': 'manage_platform', 'label': 'Platforma sozlamalari'},
        {'id': 'manage_broadcast', 'label': 'Xabarlarni yuborish'},
        {'id': 'manage_plans', 'label': 'Tariflarni boshqarish'},
    ]})

@app.route('/api/kg-admins/permissions')
@login_required
def kg_permissions_list():
    """Return available permissions for kindergarten admins."""
    return jsonify({'permissions': [
        {'id': 'students', 'label': 'O\'quvchilar'},
        {'id': 'attendance', 'label': 'Davomat'},
        {'id': 'payments', 'label': 'To\'lovlar'},
        {'id': 'registrations', 'label': 'Ro\'yxatga olish'},
        {'id': 'reports', 'label': 'Hisobotlar'},
        {'id': 'settings', 'label': 'Sozlamalar'},
        {'id': 'notifications', 'label': 'Xabarlar'},
        {'id': 'all', 'label': 'Barcha huquqlar'},
    ]})

@app.route('/api/super/admins/self-password', methods=['POST'])
@super_required
def super_admin_change_own_password():
    admin_id = session.get('admin_id', '')
    admins = pc.load_json('super_admins.json')
    admin = next((a for a in admins if a['id'] == admin_id), None)
    if not admin:
        return jsonify({'success': False, 'message': 'Admin topilmadi'}), 404
    data = request.get_json() or {}
    old = data.get('old_password', '')
    new = data.get('new_password', '').strip()
    if not check_password(old, admin.get('password', '')):
        return jsonify({'success': False, 'message': 'Eski parol noto\'g\'ri'})
    if not new or len(new) < 6:
        return jsonify({'success': False, 'message': 'Yangi parol kamida 6 belgi'})
    admin['password'] = hash_password(new)
    pc.save_json('super_admins.json', admins)
    _audit('super_admin_password_changed', f"O'z parolini o'zgartirdi: {admin['name']}")
    return jsonify({'success': True})

@app.route('/api/super/admins/self', methods=['GET'])
@super_required
def super_admin_self():
    admin_id = session.get('admin_id', '')
    admins = pc.load_json('super_admins.json')
    admin = next((a for a in admins if a['id'] == admin_id), None)
    if not admin:
        return jsonify({'success': False, 'message': 'Topilmadi'}), 404
    admin.pop('password', None)
    return jsonify({'admin': admin})

@app.route('/api/super/admins/self', methods=['PUT'])
@super_required
def super_admin_update_self():
    admin_id = session.get('admin_id', '')
    admins = pc.load_json('super_admins.json')
    idx = next((i for i, a in enumerate(admins) if a['id'] == admin_id), None)
    if idx is None:
        return jsonify({'success': False, 'message': 'Topilmadi'}), 404
    data = request.get_json() or {}
    if 'name' in data:
        admins[idx]['name'] = data['name'].strip()
    pc.save_json('super_admins.json', admins)
    _audit('super_admin_updated_self', f"O'z ma'lumotlarini yangiladi")
    resp = {k: v for k, v in admins[idx].items() if k != 'password'}
    return jsonify({'success': True, 'admin': resp})

@app.route('/api/super/admins/self/login', methods=['POST'])
@super_required
def super_admin_change_own_login():
    admin_id = session.get('admin_id', '')
    admins = pc.load_json('super_admins.json')
    admin = next((a for a in admins if a['id'] == admin_id), None)
    if not admin:
        return jsonify({'success': False, 'message': 'Topilmadi'}), 404
    data = request.get_json() or {}
    new_login = data.get('login', '').strip()
    if not new_login:
        return jsonify({'success': False, 'message': 'Login kiriting'})
    for a in admins:
        if a['id'] != admin_id and a['login'] == new_login:
            return jsonify({'success': False, 'message': 'Bu login band'})
    admin['login'] = new_login
    pc.save_json('super_admins.json', admins)
    _audit('super_admin_login_changed', f"Login o'zgartirildi: {new_login}")
    return jsonify({'success': True, 'login': new_login})

@app.route('/api/super/admins/self/logout-all', methods=['POST'])
@super_required
def super_admin_logout_all():
    """Force logout all sessions by changing secret key (signs all sessions out)."""
    import secrets
    app.secret_key = secrets.token_hex(32)
    _audit('super_admin_force_logout', "Barcha sessiyalar tugatildi")
    return jsonify({'success': True, 'message': 'Barcha sessiyalar tugatildi. Qaytadan kiring.'})

@app.route('/api/super/platform-settings', methods=['GET', 'POST'])
@super_required
def super_platform_settings_api():
    if request.method == 'GET':
        plat = pc.load_platform()
        settings = plat.get('platform_settings', {})
        for k, v in PLATFORM_SETTINGS_DEFAULTS.items():
            settings.setdefault(k, v)
        return jsonify({'settings': settings})
    data = request.get_json() or {}
    plat = pc.load_platform()
    existing = plat.get('platform_settings', {})
    for k in PLATFORM_SETTINGS_DEFAULTS:
        if k in data:
            existing[k] = data[k]
    plat['platform_settings'] = existing
    pc.save_json('platform.json', plat)
    _audit('platform_settings_updated', f"Sozlamalar o'zgartirildi: {', '.join(k for k in data if k in PLATFORM_SETTINGS_DEFAULTS)}")
    return jsonify({'success': True, 'settings': existing})

@app.route('/super/settings-audit')
@super_required
def super_settings_audit():
    logs = audit_log.get_logs(pc, limit=500)
    settings_logs = [l for l in logs if l.get('action', '') in ('platform_settings_updated', 'plan_created', 'plan_deleted', 'platform_channels_updated', 'announcement_created')]
    return render_template('super_settings_audit.html', logs=settings_logs, admin_name=session.get('admin_name'))

@app.route('/api/super/stats')
@super_required
def super_stats():
    kgs = pc.load_kindergartens()
    apps = load_json('kindergarten_applications.json')
    total_students = 0
    for kg in kgs:
        total_students += len(pc.load_json('students.json', kg['id']))
    return jsonify({
        'kindergartens': len(kgs),
        'pending_apps': sum(1 for a in apps if a.get('status') == 'pending'),
        'total_students': total_students,
        'plans': pc.load_platform().get('plans', [])
    })

@app.route('/api/super/audit-summary')
@super_required
def super_audit_summary():
    kgs = pc.load_kindergartens()
    apps = pc.load_json('kindergarten_applications.json')
    issues = []
    totals = {
        'kindergartens': len(kgs),
        'active': 0,
        'blocked': 0,
        'trial': 0,
        'payment_required': 0,
        'paid': 0,
        'students': 0,
        'active_students': 0,
        'pending_apps': sum(1 for a in apps if a.get('status') == 'pending'),
        'issues': 0
    }

    for kg in kgs:
        kg_id = kg.get('id', '')
        settings = pc.load_settings(kg_id)
        students = pc.load_json('students.json', kg_id)
        payments = pc.load_json('payments.json', kg_id)
        active_students = sum(1 for s in students if s.get('status') == 'active')
        totals['students'] += len(students)
        totals['active_students'] += active_students

        status = kg.get('status', 'active')
        if status == 'blocked':
            totals['blocked'] += 1
        else:
            totals['active'] += 1

        sub = pc.subscription_status(kg)
        phase = sub.get('phase', 'unknown')
        if phase in totals:
            totals[phase] += 1

        kg_issues = []
        severity = 'info'
        if status == 'blocked' or phase == 'blocked':
            kg_issues.append('Bog\'cha bloklangan')
            severity = 'danger'
        elif phase == 'payment_required':
            kg_issues.append('Obuna to\'lovi kerak')
            severity = 'danger'
        elif phase in ('trial', 'paid') and safe_int(sub.get('days_left'), 99) <= 5:
            kg_issues.append(f"Obuna/sinov {sub.get('days_left', 0)} kunda tugaydi")
            severity = 'warning'

        if not settings.get('bot_token'):
            kg_issues.append('Telegram bot token yo\'q')
            if severity == 'info':
                severity = 'warning'
        if not settings.get('admin_telegram_chat_id'):
            kg_issues.append('Admin Telegram chat ID yo\'q')
            if severity == 'info':
                severity = 'warning'
        if active_students == 0:
            kg_issues.append('Faol o\'quvchi yo\'q')
            if severity == 'info':
                severity = 'warning'
        if not kg.get('last_login'):
            kg_issues.append('Hali tizimga kirilmagan')
        if not payments and active_students > 0:
            kg_issues.append('To\'lov yozuvlari yo\'q')
            if severity == 'info':
                severity = 'warning'

        if kg_issues:
            issues.append({
                'id': kg_id,
                'name': kg.get('name', kg_id),
                'status': status,
                'plan': kg.get('plan', 'standard'),
                'phase': phase,
                'days_left': sub.get('days_left'),
                'students': len(students),
                'active_students': active_students,
                'last_login': kg.get('last_login'),
                'severity': severity,
                'issues': kg_issues[:6]
            })

    order = {'danger': 0, 'warning': 1, 'info': 2}
    issues.sort(key=lambda x: (order.get(x['severity'], 9), x['name'].lower()))
    totals['issues'] = len(issues)
    return jsonify({'totals': totals, 'issues': issues[:30]})

# ─── Super: Analytics page ──────────────────────────────────────────────────
@app.route('/super/analytics')
@super_required
def super_analytics():
    return render_template('super_analytics.html', admin_name=session.get('admin_name'))

@app.route('/api/super/chart-stats')
@super_required
def super_chart_stats():
    days = safe_int(request.args.get('days', 30))
    visits_data = pc.load_json(VISIT_LOG_FILE)
    if not isinstance(visits_data, dict):
        visits_data = {}
    dates = sorted(visits_data.keys())[-days:] if visits_data else []
    labels = []
    page_views = []
    unique_ips = []
    for d in dates:
        info = visits_data.get(d, {})
        labels.append(d[-5:])
        page_views.append(info.get('views', 0))
        unique_ips.append(len(info.get('ips', {})))
    # Payment stats across all KGs
    kgs = pc.load_kindergartens()
    total_paid = 0
    total_pending = 0
    monthly_revenue = {}
    for kg in kgs:
        payments = pc.load_json('payments.json', kg['id'])
        for p in payments:
            m = p.get('month', '')
            if p.get('status') == 'paid':
                total_paid += p.get('amount', 0)
                monthly_revenue[m] = monthly_revenue.get(m, 0) + p.get('amount', 0)
            elif p.get('status') == 'pending':
                total_pending += p.get('amount', 0)
    revenue_labels = sorted(monthly_revenue.keys())[-12:]
    revenue_data = [monthly_revenue.get(m, 0) for m in revenue_labels]
    # Top kindergartens by student count
    kg_stats = []
    for kg in kgs:
        students = pc.load_json('students.json', kg['id'])
        active = sum(1 for s in students if s.get('status') == 'active')
        payments = pc.load_json('payments.json', kg['id'])
        this_month = datetime.now().strftime('%Y-%m')
        monthly_paid = sum(p.get('amount', 0) for p in payments if p.get('month') == this_month and p.get('status') == 'paid')
        kg_stats.append({'id': kg['id'], 'name': kg['name'], 'students': len(students), 'active': active, 'monthly_paid': monthly_paid})
    kg_stats.sort(key=lambda x: -x['students'])
    return jsonify({
        'labels': labels,
        'page_views': page_views,
        'unique_ips': unique_ips,
        'revenue_labels': revenue_labels,
        'revenue_data': revenue_data,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'top_kgs': kg_stats[:10],
        'total_kgs': len(kgs),
        'total_students': sum(x['students'] for x in kg_stats)
    })

# ─── Super: Payments page ───────────────────────────────────────────────────
@app.route('/super/payments')
@super_required
def super_payments():
    return render_template('super_payments.html', admin_name=session.get('admin_name'))

@app.route('/api/super/payment-overview')
@super_required
def super_payment_overview():
    kg_id = request.args.get('kg_id', '')
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    kgs = pc.load_kindergartens()
    if kg_id:
        kgs = [kg for kg in kgs if kg['id'] == kg_id]
    result = []
    total_collected = 0
    total_debt = 0
    for kg in kgs:
        payments = pc.load_json('payments.json', kg['id'])
        students = pc.load_json('students.json', kg['id'])
        month_payments = [p for p in payments if p.get('month') == month]
        paid_sum = sum(p.get('amount', 0) for p in month_payments if p.get('status') == 'paid')
        pending_sum = sum(p.get('amount', 0) for p in month_payments if p.get('status') == 'pending')
        expected = sum(int(s.get('monthly_fee', 0)) for s in students if s.get('status') == 'active')
        debt = max(0, expected - paid_sum)
        total_collected += paid_sum
        total_debt += debt
        result.append({
            'id': kg['id'],
            'name': kg['name'],
            'plan': kg.get('plan', 'standard'),
            'status': kg.get('status', 'active'),
            'student_count': len(students),
            'active_count': sum(1 for s in students if s.get('status') == 'active'),
            'paid_count': len(set(p['student_id'] for p in month_payments if p.get('status') == 'paid')),
            'expected': expected,
            'paid_sum': paid_sum,
            'pending_sum': pending_sum,
            'debt': debt,
            'collection_rate': round((paid_sum / expected * 100) if expected > 0 else 0, 1)
        })
    return jsonify({
        'kindergartens': result,
        'total_collected': total_collected,
        'total_debt': total_debt,
        'total_kgs': len(result),
        'month': month
    })

# ─── Super: System Health page ──────────────────────────────────────────────
@app.route('/super/health')
@super_required
def super_health():
    return render_template('super_health.html', admin_name=session.get('admin_name'))

@app.route('/api/super/system-health')
@super_required
def super_system_health():
    checks = {}
    # Database / file storage
    checks['data_dir'] = {'ok': os.path.isdir(DATA_DIR), 'path': DATA_DIR}
    try:
        plat = pc.load_platform()
        checks['platform_config'] = {'ok': bool(plat), 'plans': len(plat.get('plans', []))}
    except Exception as e:
        checks['platform_config'] = {'ok': False, 'error': str(e)}
    try:
        kgs = pc.load_kindergartens()
        checks['kindergartens'] = {'ok': True, 'count': len(kgs)}
    except Exception as e:
        checks['kindergartens'] = {'ok': False, 'error': str(e)}
    # Super bot status
    plat = pc.load_platform()
    st = plat.get('super_bot_token', '')
    sc = plat.get('super_telegram_chat_id', '')
    checks['super_bot'] = {'ok': bool(st) and bool(sc), 'has_token': bool(st), 'has_chat_id': bool(sc)}
    if st:
        try:
            r = http_requests.get(f'https://api.telegram.org/bot{st}/getMe', timeout=5)
            data = r.json()
            checks['super_bot']['bot_online'] = data.get('ok', False)
            if data.get('ok'):
                checks['super_bot']['bot_name'] = data['result'].get('first_name', '')
                checks['super_bot']['bot_username'] = data['result'].get('username', '')
        except Exception as e:
            checks['super_bot']['bot_online'] = False
            checks['super_bot']['error'] = str(e)
    # KG bots status
    online_bots = 0
    total_bots = 0
    for kg in kgs:
        settings = pc.load_settings(kg['id'])
        token = settings.get('bot_token', '')
        if token:
            total_bots += 1
            try:
                r = http_requests.get(f'https://api.telegram.org/bot{token}/getMe', timeout=3)
                if r.json().get('ok'):
                    online_bots += 1
            except Exception:
                pass
    checks['kg_bots'] = {'ok': True, 'total': total_bots, 'online': online_bots}
    # Audit log integrity
    try:
        valid, idx = audit_log.verify_chain(pc)
        checks['audit_chain'] = {'ok': valid, 'entries': idx}
    except Exception:
        checks['audit_chain'] = {'ok': False}
    # Env vars
    checks['env'] = {
        'vercel': bool(os.environ.get('VERCEL')),
        'database_url': bool(os.environ.get('DATABASE_URL')),
        'cron_secret': bool(os.environ.get('CRON_SECRET')),
        'site_url': os.environ.get('SITE_URL', 'not set'),
        'secret_key': bool(app.secret_key)
    }
    # Uptime
    checks['health_check_at'] = datetime.now(timezone.utc).isoformat()
    checks['platform_name'] = plat.get('platform_settings', {}).get('platform_name', 'EduSoft')
    return jsonify(checks)

# ─── Super: IP Block page ───────────────────────────────────────────────────
@app.route('/super/ip-block')
@super_required
def super_ip_block():
    return render_template('super_ip_block.html', admin_name=session.get('admin_name'))

@app.route('/api/super/blocked-ips', methods=['GET', 'POST', 'DELETE'])
@super_required
def super_blocked_ips_api():
    plat = pc.load_platform()
    blocked = plat.get('blocked_ips', [])
    if request.method == 'GET':
        return jsonify({'ips': blocked, 'count': len(blocked)})
    data = request.get_json() or {}
    ip = data.get('ip', '').strip()
    if not ip:
        return jsonify({'success': False, 'message': 'IP manzil kiriting'})
    if request.method == 'POST':
        if any(b.get('ip') == ip for b in blocked):
            return jsonify({'success': False, 'message': 'Bu IP allaqachon bloklangan'})
        reason = data.get('reason', '')
        blocked.append({'ip': ip, 'reason': reason, 'blocked_at': datetime.now().isoformat(), 'blocked_by': session.get('admin_name', 'unknown')})
        plat['blocked_ips'] = blocked
        pc.save_json('platform.json', plat)
        _audit('ip_blocked', f"IP bloklandi: {ip} — {reason}")
        return jsonify({'success': True, 'ip': ip})
    if request.method == 'DELETE':
        plat['blocked_ips'] = [b for b in blocked if b.get('ip') != ip]
        pc.save_json('platform.json', plat)
        _audit('ip_unblocked', f"IP blok olib tashlandi: {ip}")
        return jsonify({'success': True})

# ─── Super: KG Admin Activity ────────────────────────────────────────────────
@app.route('/api/super/kg-admin-activity')
@super_required
def super_kg_admin_activity():
    days = safe_int(request.args.get('days'), 7)
    logs = audit_log.get_logs(pc, days=days, limit=5000)
    kgs = pc.load_kindergartens()
    kg_map = {kg['id']: kg['name'] for kg in kgs}
    by_kg = {}
    for l in logs:
        details = l.get('details', '')
        kg_id = None
        for kid in kg_map:
            if kid in details:
                kg_id = kid
                break
        if not kg_id:
            continue
        if kg_id not in by_kg:
            by_kg[kg_id] = {'kg_id': kg_id, 'kg_name': kg_map[kg_id], 'total': 0, 'actions': {}}
        by_kg[kg_id]['total'] += 1
        act = l.get('action', 'unknown')
        by_kg[kg_id]['actions'][act] = by_kg[kg_id]['actions'].get(act, 0) + 1
    summary = []
    for v in sorted(by_kg.values(), key=lambda x: -x['total']):
        recent = sorted(v['actions'].items(), key=lambda x: -x[1])[:5]
        summary.append({'kg_name': v['kg_name'], 'kg_id': v['kg_id'], 'total': v['total'], 'recent_actions': [a[0] for a in recent]})
    return jsonify({'summary': summary, 'days': days, 'total': len(summary)})

# ─── Enhanced application endpoint with rejection reason ─────────────────────
@app.route('/api/kindergarten-applications/<app_id>/status', methods=['POST'])
@super_required
def update_kindergarten_application(app_id):
    data = request.get_json() or {}
    status = data.get('status', 'reviewed')
    reason = data.get('reason', '')
    if status == 'approved':
        new_kg, app_rec = pc.approve_application(app_id, data.get('apply_logo'))
        if new_kg:
            notify_super_admin(f"✅ Tasdiqlandi: {new_kg['name']} (login: {new_kg['owner']['login']})")
            temp_pw = (app_rec or {}).get('temp_password', '12345678') if isinstance(app_rec, dict) else '12345678'
            new_kg['owner']['_temp'] = temp_pw
            return jsonify({'success': True, 'kindergarten': new_kg})
        return jsonify({'success': False, 'message': 'Topilmadi'})
    apps = load_json('kindergarten_applications.json')
    for i, a in enumerate(apps):
        if a['id'] == app_id:
            apps[i]['status'] = status
            if reason:
                apps[i]['rejection_reason'] = reason
            if status == 'rejected' and a.get('owner_telegram') and reason:
                _send_rejection_notification(a.get('owner_telegram'), a.get('kindergarten_name', ''), reason)
            save_json('kindergarten_applications.json', apps)
            return jsonify({'success': True, 'status': status})
    return jsonify({'success': False, 'message': 'Topilmadi'})

def _send_rejection_notification(chat_id, kg_name, reason):
    plat = pc.load_platform()
    token = plat.get('super_bot_token', '')
    if not token:
        return
    try:
        from telegram_bot import send_message
        msg = (
            f"❌ <b>Arizangiz rad etildi</b>\n\n"
            f"🏫 Bog'cha: <b>{esc_html(kg_name)}</b>\n"
            f"📝 Sabab: {esc_html(reason)}\n\n"
            f"Qo'shimcha ma'lumot uchun platforma admini bilan bog'lanishingiz mumkin."
        )
        send_message(token, chat_id, msg)
    except Exception:
        pass

# ─── Enhanced broadcast with targeting ──────────────────────────────────────
@app.route('/api/super/broadcast-enhanced', methods=['POST'])
@super_required
def super_broadcast_enhanced():
    data = request.get_json() or {}
    message = data.get('message', '').strip()
    target_plan = data.get('plan', '')  # empty = all
    target_status = data.get('status', '')  # empty = all
    if not message:
        return jsonify({'success': False, 'message': 'Xabar matni kiriting'})
    kgs = pc.load_kindergartens()
    sent = 0
    failed = 0
    skipped = 0
    for kg in kgs:
        if target_plan and kg.get('plan') != target_plan:
            skipped += 1
            continue
        if target_status and kg.get('status') != target_status:
            skipped += 1
            continue
        settings = pc.load_settings(kg['id'])
        token = settings.get('bot_token', '')
        if not token:
            skipped += 1
            continue
        # Send to kg admin chat
        admin_chat = settings.get('admin_telegram_chat_id', '')
        if admin_chat:
            try:
                from telegram_bot import send_message
                send_message(token, admin_chat, message)
                sent += 1
            except Exception:
                failed += 1
    return jsonify({'success': True, 'sent': sent, 'skipped': skipped, 'failed': failed, 'total': len(kgs)})

# ─── Super: All applications (with history) ──────────────────────────────────
@app.route('/api/super/all-applications')
@super_required
def super_all_applications():
    apps = load_json('kindergarten_applications.json')
    return jsonify({'applications': sorted(apps, key=lambda x: x.get('created_at', ''), reverse=True)})


# ─── Public payment link (per child per month) ─────────────────────────────
@app.route('/pay/<student_id>/<month>')
def public_pay(student_id, month):
    # Find student across all kindergartens
    kg_id = None
    student = None
    for kg in pc.load_kindergartens():
        students = pc.load_json('students.json', kg['id'])
        s = next((s for s in students if s['id'] == student_id), None)
        if s:
            student = s
            kg_id = kg['id']
            break
    if not student:
        return render_template('pay.html', found=False, error="O'quvchi topilmadi")
    settings = load_settings(kg_id)
    payments = load_json('payments.json', kg_id)
    # Check if paid for this month
    paid = any(p for p in payments if p['student_id'] == student_id
               and p.get('month') == month and p.get('status') == 'paid')
    month_payments = [p for p in payments if p['student_id'] == student_id
                      and p.get('month') == month]
    fee = int(student.get('monthly_fee', 0))
    paid_amount = sum(p.get('amount', 0) for p in month_payments if p.get('status') == 'paid')
    pending_amount = sum(p.get('amount', 0) for p in month_payments if p.get('status') == 'pending')
    return render_template('pay.html', found=True, paid=paid, student=student,
                          month=month, fee=fee, paid_amount=paid_amount,
                          pending_amount=pending_amount, kg_id=kg_id,
                          settings=settings)


@app.route('/api/payments/send-payment-request', methods=['POST'])
@login_required
def send_payment_request():
    """Admin sends payment request to unpaid parents via Telegram."""
    data = request.get_json() or {}
    kg_id = data.get('kg_id') or _current_kg_id()
    month = data.get('month', datetime.now().strftime('%Y-%m'))
    student_ids = data.get('student_ids', [])
    custom_msg = data.get('message', '')
    if not kg_id:
        return jsonify({'success': False, 'message': 'KG ID kerak'})
    settings = load_settings(kg_id)
    token = settings.get('bot_token', '')
    if not token:
        return jsonify({'success': False, 'message': 'Bot token sozlanmagan'})
    students = pc.load_json('students.json', kg_id)
    payments = load_json('payments.json', kg_id)
    paid_ids = {p['student_id'] for p in payments if p.get('month') == month and p.get('status') == 'paid'}
    sent = 0
    skipped_paid = 0
    no_chat = 0
    base_url = os.environ.get('SITE_URL', 'https://sofgardercrm.vercel.app').rstrip('/')
    target_students = [s for s in students if not student_ids or s['id'] in student_ids]
    for s in target_students:
        if s.get('status') != 'active':
            continue
        chat_id = s.get('telegram_chat_id', '')
        if not chat_id:
            no_chat += 1
            continue
        if s['id'] in paid_ids:
            skipped_paid += 1
            continue
        name = f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()
        fee = int(s.get('monthly_fee', 0))
        pay_url = f"{base_url}/pay/{s['id']}/{month}"
        msg = (
            f"💳 <b>To'lov bildirishnomasi</b>\n\n"
            f"👶 <b>{esc_html(name)}</b>\n"
            f"📅 Oy: <b>{month}</b>\n"
            f"💰 Summa: <b>{fee:,} UZS</b>\n\n"
        )
        if custom_msg:
            msg += f"📝 {esc_html(custom_msg)}\n\n"
        msg += f"🔗 <a href='{pay_url}'>To'lov qilish</a>"
        try:
            from telegram_bot import send_message
            send_message(token, chat_id, msg)
            sent += 1
        except Exception:
            pass
    return jsonify({
        'success': True,
        'sent': sent,
        'skipped_paid': skipped_paid,
        'no_chat': no_chat,
        'total': len(target_students)
    })


PUBLIC_STATS_TOKEN = None

def _get_public_stats_token():
    global PUBLIC_STATS_TOKEN
    if PUBLIC_STATS_TOKEN:
        return PUBLIC_STATS_TOKEN
    base = os.environ.get('PUBLIC_STATS_TOKEN', '').strip()
    if not base:
        sk = os.environ.get('SECRET_KEY', '')
        if sk:
            base = hashlib.sha256(sk.encode()).hexdigest()[:16]
        else:
            base = 'sofgarder2024'
    PUBLIC_STATS_TOKEN = base
    return base


@app.route('/public/stats')
def public_stats():
    token = request.args.get('token', '')
    expected = _get_public_stats_token()
    if token != expected:
        return render_template('public_stats.html', authorized=False, error='Noto\'g\'ri token'), 403
    # Gather stats
    logs = audit_log.get_logs(pc, days=30, limit=2000)
    login_logs = [l for l in logs if l.get('action') in ('login', 'logout', 'failed_login', 'super_admin_blocked_ip')]
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    today_logs = [l for l in login_logs if (l.get('timestamp') or '').startswith(today)]
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).strftime('%Y-%m-%d')
    week_logs = [l for l in login_logs if (l.get('timestamp') or '') >= week_ago]
    # Summary
    stats = {
        'today_logins': sum(1 for l in today_logs if l['action'] == 'login'),
        'today_failed': sum(1 for l in today_logs if l['action'] == 'failed_login'),
        'week_logins': sum(1 for l in week_logs if l['action'] == 'login'),
        'week_failed': sum(1 for l in week_logs if l['action'] == 'failed_login'),
        'month_logins': sum(1 for l in login_logs if l['action'] == 'login'),
        'month_failed': sum(1 for l in login_logs if l['action'] == 'failed_login'),
        'unique_ips': len(set(l.get('ip', '') for l in login_logs if l.get('ip'))),
    }
    # Recent logins (last 20)
    recent = [l for l in login_logs if l['action'] == 'login'][:20]
    return render_template('public_stats.html', authorized=True, stats=stats, recent=recent, token=token)

@app.route('/api/super/announcement', methods=['POST'])
@super_required
def super_announcement():
    data = request.get_json() or {}
    anns = pc.load_json('platform_announcements.json')
    anns.append({
        'id': str(uuid.uuid4())[:8],
        'title': data.get('title', 'Yangilanish'),
        'message': data.get('message', ''),
        'active': True,
        'created_at': datetime.now().isoformat()
    })
    pc.save_json('platform_announcements.json', anns[-50:])
    _audit('announcement_created', f"{data.get('title', '')}")
    return jsonify({'success': True})

@app.route('/api/super/platform-channels', methods=['POST'])
@super_required
def super_platform_channels():
    data = request.get_json() or {}
    plat = pc.load_platform()
    plat['platform_channels'] = [c.strip() for c in data.get('channels', []) if c.strip()]
    pc.save_json('platform.json', plat)
    _audit('platform_channels_updated', f"Kanallar yangilandi: {len(plat['platform_channels'])} ta")
    return jsonify({'success': True})

@app.route('/api/super/broadcast', methods=['POST'])
@super_required
def super_broadcast():
    """Send message to all kindergartens' Telegram bots (broadcast to parents)"""
    data = request.get_json() or {}
    msg = (data.get('message') or '').strip()
    if not msg:
        return jsonify({'success': False, 'message': 'Matn kiriting'})
    kgs = pc.load_kindergartens()
    sent = 0
    total = 0
    for kg in kgs:
        if kg.get('status') != 'active':
            continue
        kg_id = kg['id']
        settings = pc.load_settings(kg_id)
        token = settings.get('bot_token', '')
        if not token:
            continue
        students = pc.load_json('students.json', kg_id)
        parent_chats = set()
        for s in students:
            chat = s.get('telegram_chat_id', '')
            if chat:
                parent_chats.add(chat)
        # Also send to admin chat
        admin_chat = settings.get('admin_telegram_chat_id', '')
        if admin_chat:
            parent_chats.add(admin_chat)
        if not parent_chats:
            continue
        total += len(parent_chats)
        for chat_id in parent_chats:
            try:
                from telegram_bot import send_message as tg_send
                tg_send(token, chat_id,
                    f"📢 <b>Platforma xabari</b>\n\n{msg}\n\n— Raufbek Turaqulov (platforma egasi)")
                sent += 1
            except Exception:
                pass
    _audit('broadcast', f"Xabar yuborildi: {sent}/{total}")
    return jsonify({'success': True, 'sent': sent, 'total': total})

@app.route('/api/super/subscription/<kg_id>', methods=['POST'])
@super_required
def super_update_subscription(kg_id):
    data = request.get_json() or {}
    kgs = pc.load_kindergartens()
    for i, kg in enumerate(kgs):
        if kg['id'] == kg_id:
            if data.get('paid_until'):
                kgs[i].setdefault('subscription', {})['paid_until'] = data['paid_until']
            if data.get('extend_months'):
                months = int(data['extend_months'])
                base = date.today()
                pu = kgs[i].get('subscription', {}).get('paid_until')
                if pu:
                    try:
                        base = datetime.strptime(pu[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass
                kgs[i]['subscription']['paid_until'] = (base + timedelta(days=30 * months)).isoformat()
            if data.get('trial_days') is not None:
                kgs[i].setdefault('subscription', {})['trial_days'] = int(data['trial_days'])
            pc.save_kindergartens(kgs)
            _audit('subscription_updated', f"Bog'cha: {kg_id}")
            return jsonify({'success': True})
    return jsonify({'success': False}), 404

@app.route('/api/super/plan/save', methods=['POST'])
@super_required
def super_plan_save():
    data = request.get_json() or {}
    pid = (data.get('id') or '').strip().lower().replace(' ', '_')
    if not pid:
        return jsonify({'success': False, 'message': 'Plan ID kerak'}), 400
    plat = pc.load_platform()
    plans = plat.get('plans', [])
    idx = next((i for i, p in enumerate(plans) if p['id'] == pid), None)
    entry = {
        'id': pid,
        'name': (data.get('name') or 'Plansiz').strip(),
        'price_usd': safe_int(data.get('price_usd'), 10),
        'trial_days': safe_int(data.get('trial_days'), 30),
        'desc': (data.get('desc') or '').strip()
    }
    if idx is not None:
        plans[idx] = entry
    else:
        plans.append(entry)
    plat['plans'] = plans
    pc.save_json('platform.json', plat)
    _audit('plan_saved', f"Plan saqlandi: {pid} — {entry['name']}")
    return jsonify({'success': True, 'plans': plans})

@app.route('/api/super/plan/delete', methods=['POST'])
@super_required
def super_plan_delete():
    data = request.get_json() or {}
    pid = (data.get('id') or '').strip()
    plat = pc.load_platform()
    plat['plans'] = [p for p in plat.get('plans', []) if p['id'] != pid]
    pc.save_json('platform.json', plat)
    _audit('plan_deleted', f"Plan o'chirildi: {pid}")
    return jsonify({'success': True, 'plans': plat['plans']})

@app.route('/api/super/block', methods=['POST'])
@super_required
def super_block_kg():
    data = request.get_json() or {}
    kg_id = data.get('kg_id', '')
    status = data.get('status', 'blocked')
    kg = pc.get_kindergarten(kg_id)
    if not kg:
        return jsonify({'success': False, 'message': 'Bog\'cha topilmadi'}), 404
    pc.set_kg_status(kg_id, status)
    # When unblocking, extend subscription to prevent re-block
    if status == 'active':
        kgs = pc.load_kindergartens()
        for i, k in enumerate(kgs):
            if k['id'] == kg_id:
                old_until = k.get('subscription', {}).get('paid_until', '')
                base = date.today()
                if old_until:
                    try:
                        base = datetime.strptime(old_until[:10], '%Y-%m-%d').date()
                        if base < date.today():
                            base = date.today()
                    except Exception:
                        base = date.today()
                kgs[i].setdefault('subscription', {})['paid_until'] = (base + timedelta(days=30)).isoformat()
                pc.save_kindergartens(kgs)
                break
    # Notify admin via Telegram when blocked
    if status == 'blocked':
        settings = pc.load_settings(kg_id)
        token = settings.get('bot_token', '')
        admin_chat = settings.get('admin_telegram_chat_id', '')
        if token and admin_chat:
            try:
                from telegram_bot import send_message as tg_send
                tg_send(token, admin_chat,
                    '⛔ <b>Bog\'cha bloklandi</b>\n\n'
                    f'Bog\'cha: <b>{kg.get("name", kg_id)}</b>\n'
                    'Sabab: obuna to\'lovi amalga oshirilmagan.\n\n'
                    'Platforma egasi bilan bog\'lanib, to\'lovni amalga oshiring.')
            except Exception:
                pass
    _audit('kg_block_toggle', f"Bog'cha: {kg_id} → {status}")
    return jsonify({'success': True, 'message': 'Bloklandi' if status == 'blocked' else 'Faollashtirildi'})

@app.route('/api/super/kg-password', methods=['POST'])
@super_required
def super_kg_password():
    data = request.get_json() or {}
    kg_id = data.get('kg_id', '')
    new_pw = data.get('new_password', '')
    if len(new_pw) < 6:
        return jsonify({'success': False, 'message': 'Parol kamida 6 belgi'})
    kgs = pc.load_kindergartens()
    for i, kg in enumerate(kgs):
        if kg['id'] == kg_id:
            kgs[i].setdefault('owner', {})['password'] = hash_password(new_pw)
            pc.save_kindergartens(kgs)
            _audit('kg_password_reset', f"Bog'cha: {kg_id}")
            return jsonify({'success': True, 'message': 'Parol o\'zgartirildi'})
    return jsonify({'success': False, 'message': 'Bog\'cha topilmadi'}), 404

@app.route('/api/super/delete-kg', methods=['POST'])
@super_required
def super_delete_kg():
    data = request.get_json() or {}
    kg_id = data.get('kg_id', '')
    pc._delete_kindergarten(kg_id)
    _audit('kg_deleted', f"Bog'cha o'chirildi: {kg_id}")
    return jsonify({'success': True, 'message': 'Bog\'cha va barcha ma\'lumotlari o\'chirildi'})

# ─── Weekly audit report ──────────────────────────────────────────────────
@app.route('/api/super/weekly-report', methods=['POST'])
@super_required
def super_weekly_report():
    plat = pc.load_platform()
    super_token = plat.get('super_bot_token', '').strip()
    super_chat = plat.get('super_telegram_chat_id', '').strip()
    if not super_token or not super_chat:
        return jsonify({'success': False, 'message': 'Super bot token yoki chat ID sozlanmagan'})
    report = audit_log.weekly_report(pc)
    try:
        from telegram_bot import send_message as tg_send
        tg_send(super_token, super_chat, report)
        _audit('weekly_report_sent', f"Haftalik hisobot yuborildi ({len(report)} belgi)")
        return jsonify({'success': True, 'message': 'Hisobot yuborildi'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Yuborishda xatolik: {str(e)}'})

@app.route('/api/super/kg-balance', methods=['POST'])
@super_required
def super_kg_balance():
    data = request.get_json() or {}
    kg_id = data.get('kg_id', '')
    amount = safe_int(data.get('amount'), 0)
    note = data.get('note', '')
    if not kg_id:
        return jsonify({'success': False, 'message': 'kg_id majburiy'})
    if amount == 0:
        return jsonify({'success': False, 'message': 'Summa 0 bo\'lmasligi kerak'})
    kg = pc.get_kindergarten(kg_id)
    if not kg:
        return jsonify({'success': False, 'message': 'Bog\'cha topilmadi'})
    settings = pc.load_settings(kg_id)
    old_balance = settings.get('balance', 0)
    settings['balance'] = old_balance + amount
    pc.save_settings(kg_id, settings)
    action = 'qo\'shildi' if amount > 0 else 'yechildi'
    _audit('super_balance_changed',
           f"[Super Admin] {kg['name']} balansi: {old_balance:,} → {settings['balance']:,} so'm "
           f"({'+' if amount >= 0 else ''}{amount:,} {action})"
           f"{' — ' + note if note else ''}")
    return jsonify({'success': True, 'balance': settings['balance']})

@app.route('/api/super/suspicious')
@super_required
def super_suspicious():
    logs = audit_log.get_logs(pc, limit=500)
    suspicious = [e for e in logs if 'hours_ago' in e.get('details', '') or e.get('action') in (
        'payment_deleted', 'attendance_bulk_edit', 'upload_magic_mismatch', 'honeypot_triggered',
        'payment_duplicate_attempt',
    )]
    return jsonify({'logs': suspicious[:50], 'count': min(len(suspicious), 50)})

@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    students = load_json('students.json')
    payments = load_json('payments.json')
    attendance = load_json('attendance.json')

    now = datetime.now()
    y, m = now.year, now.month
    today_str = now.strftime('%Y-%m-%d')

    total_students = len(students)
    active_students = sum(1 for s in students if s.get('status') == 'active')

    # Monthly income
    monthly_income = sum(p['amount'] for p in payments
                         if p.get('month', '') == f'{y}-{m:02d}' and p.get('status', 'paid') != 'cancelled')

    # Debtors + unpaid this month
    debtors = 0
    unpaid = 0
    discount_count = 0
    subsidy_count = 0
    for s in students:
        if s.get('status') != 'active':
            continue
        if safe_int(s.get('discount_percent'), 0) > 0:
            discount_count += 1
        if safe_int(s.get('subsidy_amount'), 0) > 0:
            subsidy_count += 1
        payable = calc_payable_amount(s, y, m)
        paid = get_paid_amount(s['id'], y, m)
        if paid < payable:
            debtors += 1
            unpaid += payable - paid

    # Today attendance
    today_att = sum(1 for a in attendance
                    if a['date'] == today_str and a['status'] == 'present')

    # Recent payments (last 5)
    sorted_payments = sorted(payments, key=lambda x: x.get('date',''), reverse=True)[:5]
    recent_payments = []
    for p in sorted_payments:
        s = next((st for st in students if st['id'] == p['student_id']), None)
        recent_payments.append({
            **p,
            'student_name': f"{s['first_name']} {s['last_name']}" if s else 'Unknown'
        })

    # Recent students
    sorted_students = sorted(students, key=lambda x: x.get('join_date',''), reverse=True)[:5]

    # Recent attendance
    sorted_att = sorted(attendance, key=lambda x: x.get('date',''), reverse=True)[:5]
    recent_att = []
    for a in sorted_att:
        s = next((st for st in students if st['id'] == a['student_id']), None)
        recent_att.append({
            **a,
            'student_name': f"{s['first_name']} {s['last_name']}" if s else 'Unknown'
        })

    return jsonify({
        'total_students': total_students,
        'active_students': active_students,
        'debtors': debtors,
        'monthly_income': monthly_income,
        'unpaid': unpaid,
        'today_attendance': today_att,
        'recent_payments': recent_payments,
        'recent_students': sorted_students[:5],
        'recent_attendance': recent_att,
        'discount_count': discount_count,
        'subsidy_count': subsidy_count
    })

# ─── Students ─────────────────────────────────────────────────────────────────

@app.route('/students')
@login_required
def students_page():
    settings = load_settings()
    return render_template('students.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/students', methods=['GET'])
@login_required
def get_students():
    students = load_json('students.json')
    group = request.args.get('group', '')
    search = request.args.get('search', '').lower()
    status = request.args.get('status', '')
    course_id = request.args.get('course_id', '')

    result = students
    if group:
        result = [s for s in result if s.get('group', '') == group]
    if status:
        result = [s for s in result if s.get('status', '') == status]
    if course_id:
        result = [s for s in result if s.get('course_id', '') == course_id]
    if search:
        result = [s for s in result if
                  search in s.get('first_name', '').lower() or
                  search in s.get('last_name', '').lower() or
                  search in s.get('parent_name', '').lower() or
                  search in s.get('parent_phone', '').lower() or
                  search in s.get('id', '').lower()]

    return jsonify(result)

@app.route('/api/students', methods=['POST'])
@login_required
def add_student():
    data = request.get_json() or {}
    students = load_json('students.json')

    # Z7 — Input validation
    fname = sanitize_html(data.get('first_name', '')).strip()
    lname = sanitize_html(data.get('last_name', '')).strip()
    if not validate_name(fname):
        return jsonify({'success': False, 'message': 'Ism 2-50 harf va faqat harflar'})
    if not validate_name(lname):
        return jsonify({'success': False, 'message': 'Familiya 2-50 harf va faqat harflar'})

    # Check duplicate (use sanitized values for comparison)
    phone = data.get('parent_phone', '').strip()
    for s in students:
        if s['first_name'].lower() == fname.lower() and s['last_name'].lower() == lname.lower():
            return jsonify({'success': False, 'message': 'Бу исмли ўқувчи аллақачон мавжуд'})

    student = {
        'id': 'STU-' + str(uuid.uuid4())[:8].upper(),
        'first_name': fname,
        'last_name': lname,
        'birth_date': data.get('birth_date', ''),
        'parent_name': data.get('parent_name', '').strip(),
        'parent_phone': normalize_phone(phone) or phone,
        'group': data.get('group', '').strip(),
        'course_id': data.get('course_id', ''),
        'monthly_fee': safe_int(data.get('monthly_fee'), 0),
        'payment_due_day': safe_int(data.get('payment_due_day'), 1),
        'join_date': data.get('join_date', date.today().isoformat()),
        'status': data.get('status', 'active'),
        'image': data.get('image', ''),
        'telegram_chat_id': data.get('telegram_chat_id', ''),
        'discount_percent': safe_int(data.get('discount_percent'), 0),
        'subsidy_amount': safe_int(data.get('subsidy_amount'), 0),
        'notes': data.get('notes', '')
    }
    students.append(student)
    save_json('students.json', students)

    add_notification(f"Yangi o'quvchi qo'shildi: {fname} {lname}", 'info')
    _audit('student_created', f"{student['id']} — {fname} {lname}")

    return jsonify({'success': True, 'student': student})

@app.route('/api/students/bulk-import', methods=['POST'])
@login_required
def bulk_import_students():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Fayl topilmadi'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Fayl tanlanmagan'})
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls'):
        return jsonify({'success': False, 'message': 'Faqat .xlsx yoki .xls fayllar'})
    # Parse Excel
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({'success': False, 'message': f'Excel o\'qishda xatolik: {str(e)}'})
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    students = load_json('students.json')
    added = 0
    errors = []
    for i, row in enumerate(rows):
        if not row or not row[0]:
            continue
        try:
            fname = str(row[0] or '').strip()
            lname = str(row[1] or '').strip()
            phone = str(row[2] or '').strip()
            parent = str(row[3] or '').strip()
            group = str(row[4] or '').strip()
            fee = 0
            try: fee = int(float(str(row[5] or 0)))
            except: pass
            if not fname or not lname:
                errors.append(f"Qator {i+2}: ism/familiya bo'sh")
                continue
            student = {
                'id': 'STU-' + str(uuid.uuid4())[:8].upper(),
                'first_name': fname,
                'last_name': lname,
                'birth_date': '',
                'parent_name': parent,
                'parent_phone': normalize_phone(phone),
                'group': group,
                'monthly_fee': fee,
                'payment_due_day': 1,
                'join_date': date.today().isoformat(),
                'status': 'active',
                'image': '',
                'telegram_chat_id': '',
                'discount_percent': 0,
                'subsidy_amount': 0,
                'notes': ''
            }
            students.append(student)
            added += 1
        except Exception as e:
            errors.append(f"Qator {i+2}: {str(e)}")
    save_json('students.json', students)
    _audit('students_bulk_import', f"{added} ta o'quvchi import qilindi, {len(errors)} ta xatolik")
    return jsonify({'success': True, 'added': added, 'errors': errors, 'total': len(rows)})

@app.route('/api/students/<student_id>', methods=['GET'])
@login_required
def get_student(student_id):
    students = load_json('students.json')
    s = next((s for s in students if s['id'] == student_id), None)
    if not s:
        return jsonify({'success': False, 'message': 'Ўқувчи топилмади'}), 404
    courses_map = {c['id']: c['name'] for c in load_json('courses.json')}
    s['course_name'] = courses_map.get(s.get('course_id', ''), '')
    return jsonify(s)

@app.route('/api/students/<student_id>', methods=['PUT'])
@login_required
def update_student(student_id):
    data = request.get_json() or {}
    students = load_json('students.json')

    fname = sanitize_html(data.get('first_name', '')).strip()
    lname = sanitize_html(data.get('last_name', '')).strip()
    if fname and not validate_name(fname):
        return jsonify({'success': False, 'message': 'Ism 2-50 harf va faqat harflar'})
    if lname and not validate_name(lname):
        return jsonify({'success': False, 'message': 'Familiya 2-50 harf va faqat harflar'})

    for i, s in enumerate(students):
        if s['id'] == student_id:
            students[i].update({
                'first_name': fname or s['first_name'],
                'last_name': lname or s['last_name'],
                'birth_date': data.get('birth_date', s['birth_date']),
                'parent_name': sanitize_html(data.get('parent_name', s['parent_name'])).strip(),
                'parent_phone': data.get('parent_phone', s['parent_phone']).strip(),
                'group': data.get('group', s['group']).strip(),
                'course_id': data.get('course_id', s.get('course_id', '')),
                'monthly_fee': safe_int(data.get('monthly_fee'), s.get('monthly_fee', 0)),
                'payment_due_day': safe_int(data.get('payment_due_day'), s.get('payment_due_day', 1)),
                'status': data.get('status', s['status']),
                'image': data.get('image', s.get('image', '')),
                'telegram_chat_id': data.get('telegram_chat_id', s.get('telegram_chat_id', '')),
                'discount_percent': safe_int(data.get('discount_percent'), s.get('discount_percent', 0)),
                'subsidy_amount': safe_int(data.get('subsidy_amount'), s.get('subsidy_amount', 0)),
                'notes': sanitize_html(data.get('notes', s.get('notes', ''))).strip()
            })
            save_json('students.json', students)
            _audit('student_updated', f"{student_id} — {students[i]['first_name']} {students[i]['last_name']}")
            return jsonify({'success': True, 'student': students[i]})
    return jsonify({'success': False, 'message': 'Ўқувчи топилмади'}), 404

@app.route('/api/students/<student_id>', methods=['DELETE'])
@login_required
def delete_student(student_id):
    students = load_json('students.json')
    student = next((s for s in students if s['id'] == student_id), None)
    if not student:
        return jsonify({'success': False, 'message': 'O\'quvchi topilmadi'}), 404
    parent_phone = student.get('parent_phone', '')
    students = [s for s in students if s['id'] != student_id]
    save_json('students.json', students)

    for fname, key in [('attendance.json', 'student_id'), ('payments.json', 'student_id')]:
        data = load_json(fname)
        data = [r for r in data if r.get(key) != student_id]
        save_json(fname, data)

    # Clean up related records
    for fname, key in [('parent_portfolios.json', 'student_id'), ('complaints.json', 'student_id'),
                        ('payment_checks.json', 'student_id')]:
        try:
            data = load_json(fname)
            data = [r for r in data if r.get(key) != student_id]
            save_json(fname, data)
        except Exception:
            pass

    # Mark registration as cancelled
    try:
        regs = load_json('registrations.json')
        changed = False
        for i, r in enumerate(regs):
            if r.get('student_id') == student_id:
                regs[i]['status'] = 'cancelled'
                changed = True
        if changed:
            save_json('registrations.json', regs)
    except Exception:
        pass

    # Clean bot sessions with this phone
    if parent_phone:
        try:
            sessions = load_json('bot_sessions.json')
            pn = normalize_phone(parent_phone)
            sessions = [s for s in sessions if s.get('phone', '') != pn]
            save_json('bot_sessions.json', sessions)
        except Exception:
            pass

    add_notification(f"O'quvchi o'chirildi: {student['first_name']} {student['last_name']}", 'warning')
    _audit('student_deleted', f"{student_id} — {student['first_name']} {student['last_name']}")
    return jsonify({'success': True})

@app.route('/api/groups')
@login_required
def get_groups():
    students = load_json('students.json')
    groups = list(set(s.get('group', '') for s in students if s.get('group')))
    return jsonify(sorted(groups))

# ─── Global Search ────────────────────────────────────────────────────────────
@app.route('/api/search')
@login_required
def global_search():
    q = request.args.get('q', '').strip().lower()
    if not q or len(q) < 1:
        return jsonify({'students': [], 'parents': [], 'groups': []})

    students = load_json('students.json')
    portfolios = load_json('parent_portfolios.json')

    phone_norm = ''
    if q.replace('+', '').isdigit():
        phone_norm = normalize_phone(q)

    # Search students
    student_results = []
    for s in students:
        score = 0
        fname = s.get('first_name', '').lower()
        lname = s.get('last_name', '').lower()
        pname = s.get('parent_name', '').lower()
        pphone = s.get('parent_phone', '')
        group = s.get('group', '').lower()
        sid = s.get('id', '').lower()

        full = f"{fname} {lname}"
        if q == fname or q == lname or q == full:
            score = 100
        elif q in sid:
            score = 90
        elif q in fname or q in lname:
            score = 80
        elif q in full:
            score = 75
        elif q in pname:
            score = 70
        elif q in group:
            score = 60
        elif phone_norm and normalize_phone(pphone) == phone_norm:
            score = 95
        elif phone_norm and (
            normalize_phone(pphone).endswith(phone_norm[-7:]) if normalize_phone(pphone) else False
        ):
            score = 85

        if score:
            student_results.append({
                'id': s['id'],
                'name': f"{s.get('first_name', '')} {s.get('last_name', '')}",
                'parent': pname.title() if pname else '',
                'phone': pphone,
                'group': s.get('group', ''),
                'status': s.get('status', 'active'),
                'score': score
            })

    student_results.sort(key=lambda x: -x['score'])

    # Search parents
    parent_results = []
    sid_map = {s['id']: s for s in students}
    for p in portfolios:
        score = 0
        pname = p.get('parent_name', '').lower()
        pphone = p.get('phone', '')
        notes = p.get('notes', '').lower()
        pid = p.get('id', '').lower()
        children_names = []
        for sid in p.get('student_ids', []):
            s = sid_map.get(sid)
            if s:
                children_names.append(f"{s.get('first_name', '')} {s.get('last_name', '')}")

        if q in pid:
            score = 90
        elif q in pname:
            score = 80
        elif q in notes:
            score = 60
        elif any(q in cn.lower() for cn in children_names):
            score = 70
        elif phone_norm and normalize_phone(pphone) == phone_norm:
            score = 95

        if score:
            parent_results.append({
                'id': p['id'],
                'name': p.get('parent_name', ''),
                'phone': pphone,
                'children': children_names,
                'score': score
            })

    parent_results.sort(key=lambda x: -x['score'])

    # Search groups
    all_groups = list(set(s.get('group', '') for s in students if s.get('group')))
    group_results = [g for g in all_groups if q in g.lower()]

    # Phone-only results (extra — match any phone field)
    phone_results = []
    if phone_norm and not student_results and not parent_results:
        for s in students:
            if phone_norm and normalize_phone(s.get('parent_phone', '')) == phone_norm:
                phone_results.append({
                    'type': 'student',
                    'id': s['id'],
                    'name': f"{s.get('first_name', '')} {s.get('last_name', '')}",
                    'phone': s.get('parent_phone', '')
                })

    return jsonify({
        'students': student_results[:15],
        'parents': parent_results[:15],
        'groups': group_results[:10],
        'phone_matches': phone_results
    })

# ─── Courses ──────────────────────────────────────────────────────────────────

@app.route('/courses')
@login_required
def courses_page():
    return render_template('courses.html', admin_name=session.get('admin_name'))

@app.route('/api/courses', methods=['GET'])
@login_required
def get_courses():
    courses = load_json('courses.json')
    students = load_json('students.json')
    for c in courses:
        c['student_count'] = len([s for s in students if s.get('course_id') == c['id'] and s.get('status') == 'active'])
        c['groups'] = c.get('groups', [])
    return jsonify(courses)

@app.route('/api/courses', methods=['POST'])
@login_required
def create_course():
    data = request.get_json() or {}
    name = sanitize_html(data.get('name', '')).strip()
    if not name:
        return jsonify({'success': False, 'message': 'Kurs nomini kiriting'})
    courses = load_json('courses.json')
    groups = data.get('groups', [])
    if isinstance(groups, str):
        groups = [g.strip() for g in groups.split(',') if g.strip()]
    course = {
        'id': f"CRS-{uuid.uuid4().hex[:8].upper()}",
        'name': name,
        'description': sanitize_html(data.get('description', '')).strip(),
        'price': safe_int(data.get('price', 0)),
        'status': data.get('status', 'active'),
        'start_date': data.get('start_date', ''),
        'end_date': data.get('end_date', ''),
        'groups': groups
    }
    courses.append(course)
    save_json('courses.json', courses)
    _audit('course_created', f"Kurs: {name}")
    return jsonify({'success': True, 'course': course})

@app.route('/api/courses/<course_id>', methods=['PUT'])
@login_required
def update_course(course_id):
    data = request.get_json() or {}
    courses = load_json('courses.json')
    for c in courses:
        if c['id'] == course_id:
            c['name'] = sanitize_html(data.get('name', c['name'])).strip()
            c['description'] = sanitize_html(data.get('description', c.get('description', ''))).strip()
            c['price'] = safe_int(data.get('price', c.get('price', 0)))
            c['status'] = data.get('status', c.get('status', 'active'))
            c['start_date'] = data.get('start_date', c.get('start_date', ''))
            c['end_date'] = data.get('end_date', c.get('end_date', ''))
            groups = data.get('groups', c.get('groups', []))
            if isinstance(groups, str):
                groups = [g.strip() for g in groups.split(',') if g.strip()]
            c['groups'] = groups
            save_json('courses.json', courses)
            _audit('course_updated', f"Kurs: {c['name']}")
            return jsonify({'success': True, 'course': c})
    return jsonify({'success': False, 'message': 'Kurs topilmadi'}), 404

@app.route('/api/courses/<course_id>', methods=['DELETE'])
@login_required
def delete_course(course_id):
    courses = load_json('courses.json')
    courses = [c for c in courses if c['id'] != course_id]
    save_json('courses.json', courses)
    _audit('course_deleted', f"Kurs ID: {course_id}")
    return jsonify({'success': True})

# ─── Teachers ─────────────────────────────────────────────────────────────────

@app.route('/teachers')
@login_required
def teachers_page():
    return render_template('teachers.html', admin_name=session.get('admin_name'))

@app.route('/api/teachers', methods=['GET'])
@login_required
def get_teachers():
    teachers = load_json('teachers.json')
    courses = {c['id']: c['name'] for c in load_json('courses.json')}
    for t in teachers:
        t['_courses'] = [courses.get(cid, cid) for cid in t.get('course_ids', [])]
    return jsonify(teachers)

@app.route('/api/teachers', methods=['POST'])
@login_required
def create_teacher():
    data = request.get_json() or {}
    name = sanitize_html(data.get('name', '')).strip()
    login = sanitize_html(data.get('login', '')).strip()
    password = data.get('password', '')
    if not name or not login or not password:
        return jsonify({'success': False, 'message': 'Ism, login va parolni kiriting'})
    if len(password) < 4:
        return jsonify({'success': False, 'message': 'Parol kamida 4 belgi'})
    teachers = load_json('teachers.json')
    for t in teachers:
        if t.get('login') == login:
            return jsonify({'success': False, 'message': 'Bu login band'})
    teacher = {
        'id': f"TCH-{uuid.uuid4().hex[:8].upper()}",
        'name': name,
        'login': login,
        'password': hash_password(password),
        'phone': sanitize_html(data.get('phone', '')).strip(),
        'course_ids': data.get('course_ids', []),
        'status': data.get('status', 'active')
    }
    teachers.append(teacher)
    save_json('teachers.json', teachers)
    _audit('teacher_created', f"O'qituvchi: {name}, login: {login}")
    return jsonify({'success': True, 'teacher': {k: v for k, v in teacher.items() if k != 'password'}})

@app.route('/api/teachers/<teacher_id>', methods=['PUT'])
@login_required
def update_teacher(teacher_id):
    data = request.get_json() or {}
    teachers = load_json('teachers.json')
    for t in teachers:
        if t['id'] == teacher_id:
            if data.get('name'):
                t['name'] = sanitize_html(data['name']).strip()
            if data.get('login'):
                new_login = sanitize_html(data['login']).strip()
                if new_login != t['login']:
                    for other in teachers:
                        if other.get('login') == new_login:
                            return jsonify({'success': False, 'message': 'Bu login band'})
                t['login'] = new_login
            if data.get('password'):
                t['password'] = hash_password(data['password'])
            if 'phone' in data:
                t['phone'] = sanitize_html(data.get('phone', '')).strip()
            if 'course_ids' in data:
                t['course_ids'] = data['course_ids']
            if 'status' in data:
                t['status'] = data['status']
            save_json('teachers.json', teachers)
            _audit('teacher_updated', f"O'qituvchi: {t['name']}")
            return jsonify({'success': True, 'teacher': {k: v for k, v in t.items() if k != 'password'}})
    return jsonify({'success': False, 'message': 'O\'qituvchi topilmadi'}), 404

@app.route('/api/teachers/<teacher_id>', methods=['DELETE'])
@login_required
def delete_teacher(teacher_id):
    teachers = load_json('teachers.json')
    teachers = [t for t in teachers if t['id'] != teacher_id]
    save_json('teachers.json', teachers)
    _audit('teacher_deleted', f"O'qituvchi ID: {teacher_id}")
    return jsonify({'success': True})

# ─── Attendance ───────────────────────────────────────────────────────────────

@app.route('/attendance')
@login_required
def attendance_page():
    settings = load_settings()
    return render_template('attendance.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/attendance', methods=['GET'])
@login_required
def get_attendance():
    att = load_json('attendance.json')
    date_filter = request.args.get('date', '')
    student_id = request.args.get('student_id', '')
    month = request.args.get('month', '')
    course_id = request.args.get('course_id', '')

    result = att
    if date_filter:
        result = [a for a in result if a['date'] == date_filter]
    if student_id:
        result = [a for a in result if a['student_id'] == student_id]
    if month:
        result = [a for a in result if a['date'].startswith(month)]
    if course_id:
        students_map = {s['id']: s.get('course_id', '') for s in load_json('students.json')}
        result = [a for a in result if students_map.get(a['student_id'], '') == course_id]

    return jsonify(result)

@app.route('/api/attendance', methods=['POST'])
@login_required
def mark_attendance():
    data = request.get_json() or {}
    att = load_json('attendance.json')
    student_id = data.get('student_id', '')
    att_date = data.get('date', '')
    status = data.get('status', 'present')
    if not student_id or not att_date:
        return jsonify({'success': False, 'message': 'student_id va date majburiy'}), 400

    # Update or insert
    found = False
    for i, a in enumerate(att):
        if a['student_id'] == student_id and a['date'] == att_date:
            att[i]['status'] = status
            att[i]['note'] = data.get('note', '')
            found = True
            break
    if not found:
        att.append({
            'id': str(uuid.uuid4())[:8],
            'student_id': student_id,
            'date': att_date,
            'status': status,
            'note': data.get('note', ''),
            'marked_by': session.get('admin_name', 'Admin'),
            'created_at': datetime.now().isoformat()
        })
    save_json('attendance.json', att)
    _audit('attendance_marked', f"{student_id} {att_date} → {status}")
    return jsonify({'success': True})

@app.route('/api/attendance/bulk', methods=['POST'])
@login_required
def bulk_attendance():
    data = request.get_json() or {}
    att_date = data.get('date', '')
    records = data.get('records', [])
    if not att_date or not records:
        return jsonify({'success': False, 'message': 'date va records majburiy'}), 400

    att = load_json('attendance.json')
    for rec in records:
        found = False
        for i, a in enumerate(att):
            if a['student_id'] == rec['student_id'] and a['date'] == att_date:
                att[i]['status'] = rec['status']
                att[i]['note'] = rec.get('note', '')
                found = True
                break
        if not found:
            att.append({
                'id': str(uuid.uuid4())[:8],
                'student_id': rec['student_id'],
                'date': att_date,
                'status': rec['status'],
                'note': rec.get('note', ''),
                'marked_by': session.get('admin_name', 'Admin'),
                'created_at': datetime.now().isoformat()
            })
    save_json('attendance.json', att)
    _audit('attendance_bulk_edit', f"{att_date} — {len(records)} records")
    return jsonify({'success': True})

@app.route('/api/attendance/month-summary')
@login_required
def month_summary():
    year = safe_int(request.args.get('year'), datetime.now().year)
    month = safe_int(request.args.get('month'), datetime.now().month)
    students = load_json('students.json')

    result = []
    working_days = get_working_days(year, month)
    for s in students:
        if s.get('status') != 'active':
            continue
        attended = get_attended_days(s['id'], year, month)
        absent = get_absent_days(s['id'], year, month)
        payable = calc_payable_amount(s, year, month)
        paid = get_paid_amount(s['id'], year, month)
        result.append({
            'student': s,
            'working_days': working_days,
            'attended': attended,
            'absent': absent,
            'payable_amount': payable,
            'paid_amount': paid,
            'debt': max(payable - paid, 0)
        })
    return jsonify(result)

# ─── Payments ─────────────────────────────────────────────────────────────────

@app.route('/payments')
@login_required
def payments_page():
    settings = load_settings()
    return render_template('payments.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/payments', methods=['GET'])
@login_required
def get_payments():
    try:
        payments = load_json('payments.json')
        students = load_json('students.json')
    except Exception as e:
        print(f"[ERROR] get_payments: yuklash xatosi: {e}")
        return jsonify([])
    try:
        student_map = {s['id']: s for s in students}
    except Exception as e:
        print(f"[ERROR] get_payments: student_map xatosi: {e}")
        student_map = {}
    student_id = request.args.get('student_id', '')
    month = request.args.get('month', '')
    category = request.args.get('category', '')
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '').lower()

    result = payments
    if student_id:
        result = [p for p in result if p['student_id'] == student_id]
    if month:
        result = [p for p in result if p.get('month', '').startswith(month)]
    if category:
        result = [p for p in result if p.get('category') == category]
    if status_filter:
        result = [p for p in result if p.get('status', 'paid') == status_filter]
    if search:
        result = [p for p in result if search in p.get('student_name', '').lower() or search in p.get('note', '').lower()]

    # Resolve current student name
    for p in result:
        s = student_map.get(p.get('student_id', ''))
        if s:
            p['student_name'] = f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()
        p.setdefault('category', 'tuition')
        p.setdefault('status', 'paid')

    result = sorted(result, key=lambda x: x.get('date', ''), reverse=True)
    return jsonify(result)

@app.route('/api/payments', methods=['POST'])
@login_required
def add_payment():
    data = request.get_json() or {}
    payments = load_json('payments.json')
    students = load_json('students.json')

    student_id = data.get('student_id', '')
    if not student_id:
        return jsonify({'success': False, 'message': 'Student ID kerak'})
    student = next((s for s in students if s['id'] == student_id), None)
    if not student:
        return jsonify({'success': False, 'message': 'Ўқувчи топилмади'})

    amount = safe_int(data.get('amount'), 0)
    if amount <= 0:
        return jsonify({'success': False, 'message': 'To\'lov summasi noto\'g\'ri'})

    month = data.get('month', datetime.now().strftime('%Y-%m'))

    # Smarter duplicate: allow partials, only block if total would exceed monthly_fee
    monthly_fee = safe_int(student.get('monthly_fee'), 0)
    existing_total = sum(
        p['amount'] for p in payments
        if p.get('student_id') == student_id and p.get('month') == month
        and p.get('status', 'paid') != 'cancelled'
    )
    if existing_total + amount > monthly_fee and monthly_fee > 0:
        _audit('payment_duplicate_attempt', f"{student_id} — {month} (mavjud: {existing_total}, yangi: {amount}, limit: {monthly_fee})")
        return jsonify({'success': False, 'message': f'Bu oy uchun to\'lov limiti ({monthly_fee:,} UZS) dan oshib ketdi. Mavjud to\'lov: {existing_total:,} UZS.'})

    receipt_id = 'RCP-' + str(uuid.uuid4())[:8].upper()
    payment = {
        'id': receipt_id,
        'student_id': student_id,
        'student_name': f"{student['first_name']} {student['last_name']}",
        'amount': safe_int(data.get('amount'), 0),
        'date': data.get('date', date.today().isoformat()),
        'month': month,
        'type': data.get('type', 'full'),  # full / partial
        'category': data.get('category', 'tuition'),  # tuition / meals / materials / event / uniform / other
        'status': data.get('status', 'paid'),  # paid / pending / cancelled / refunded
        'note': data.get('note', ''),
        'admin_name': session.get('admin_name', 'Admin'),
        'created_at': datetime.now().isoformat()
    }
    payments.append(payment)
    save_json('payments.json', payments)

    add_notification(f"To'lov qabul qilindi: {student['first_name']} {student['last_name']} - {payment['amount']:,} UZS", 'payment')
    _audit('payment_created', f"{receipt_id} — {student['first_name']} {student['last_name']} — {payment['amount']:,} UZS ({payment.get('category','tuition')})")

    return jsonify({'success': True, 'payment': payment, 'receipt_id': receipt_id})

@app.route('/api/payments/bulk', methods=['POST'])
@login_required
def bulk_create_payments():
    data = request.get_json() or {}
    student_ids = data.get('student_ids', [])
    amount = data.get('amount', 0)
    month = data.get('month', '')
    category = data.get('category', 'tuition')
    note = data.get('note', '')
    payment_date = data.get('date', date.today().isoformat())
    if not student_ids or not amount:
        return jsonify({'success': False, 'message': 'O\'quvchilar va summa kiriting'})
    students = load_json('students.json')
    payments = load_json('payments.json')
    s_map = {s['id']: s for s in students}
    created = []
    for sid in student_ids:
        s = s_map.get(sid)
        if not s:
            continue
        p = {
            'id': 'PAY-' + str(uuid.uuid4())[:8].upper(),
            'student_id': sid,
            'student_name': f"{s['first_name']} {s['last_name']}",
            'amount': amount,
            'month': month,
            'category': category,
            'status': 'paid',
            'date': payment_date,
            'note': note,
            'admin_id': session.get('admin_id', ''),
            'admin_name': session.get('admin_name', ''),
            'kg_id': _current_kg_id() or 'default'
        }
        payments.insert(0, p)
        created.append(p)
    save_json('payments.json', payments)
    _audit('payments_bulk_created', f"{len(created)} ta to'lov yaratildi")
    return jsonify({'success': True, 'count': len(created), 'payments': created})

@app.route('/api/payments/<payment_id>', methods=['PUT'])
@login_required
def update_payment(payment_id):
    data = request.get_json() or {}
    payments = load_json('payments.json')
    idx = next((i for i, p in enumerate(payments) if p['id'] == payment_id), None)
    if idx is None:
        return jsonify({'success': False, 'message': 'To\'lov topilmadi'}), 404
    p = payments[idx]
    changed = []
    for field in ('amount', 'date', 'month', 'type', 'category', 'status', 'note'):
        if field in data:
            val = data[field]
            if field == 'amount':
                val = safe_int(val, p['amount'])
                if val <= 0:
                    continue
            old_val = p.get(field)
            p[field] = val
            if str(old_val) != str(val):
                changed.append(f"{field}: {old_val} → {val}")
    p['updated_at'] = datetime.now().isoformat()
    payments[idx] = p
    save_json('payments.json', payments)
    _audit('payment_updated', f"{payment_id} — o'zgartirildi: {'; '.join(changed)}")
    return jsonify({'success': True, 'payment': p})

@app.route('/api/payments/<payment_id>', methods=['DELETE'])
@login_required
def delete_payment(payment_id):
    kg_id = _current_kg_id()
    payments = load_json('payments.json', kg_id)
    removed = [p for p in payments if p['id'] == payment_id]
    payments = [p for p in payments if p['id'] != payment_id]
    save_json('payments.json', payments, kg_id)
    print(f"[DEBUG] delete_payment: payment={payment_id}, kg_id={kg_id}, "
          f"removed={'yes' if removed else 'no'}, remaining={len(payments)}")
    if removed:
        hours_ago = 0
        try:
            created = removed[0].get('created_at', '')
            if created:
                dt = datetime.fromisoformat(created)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                hours_ago = (datetime.now(timezone.utc) - dt).total_seconds() / 3600
        except Exception:
            pass
        details = f"{payment_id} — {removed[0].get('student_name','')} — {removed[0].get('amount',0):,} UZS hours_ago:{int(hours_ago)}"
        _audit('payment_deleted', details)
        if hours_ago >= 24:
            plat = pc.load_platform()
            st = plat.get('super_bot_token', '').strip()
            sc = plat.get('super_telegram_chat_id', '').strip()
            if st and sc:
                from telegram_bot import send_message as tg_send
                try:
                    tg_send(st, sc,
                        f"⚠️ <b>Shubhali: to'lov o'chirildi</b>\n\n"
                        f"To'lov: {payment_id}\n"
                        f"O'quvchi: {removed[0].get('student_name','')}\n"
                        f"Summa: {removed[0].get('amount',0):,} UZS\n"
                        f"Admin: {session.get('admin_name','')}\n"
                        f"Bog'cha: {kg_id or 'default'}\n"
                        f"Yaratilgan: {removed[0].get('created_at','')[:16]}\n"
                        f"Hozir: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}\n"
                        f"⏱ {int(hours_ago)} soat avval yaratilgan")
                except Exception:
                    pass
    return jsonify({'success': True, 'removed': bool(removed), 'kg_id': kg_id})

@app.route('/api/student-balance/<student_id>')
@login_required
def student_balance(student_id):
    students = load_json('students.json')
    student = next((s for s in students if s['id'] == student_id), None)
    if not student:
        return jsonify({'success': False}), 404

    now = datetime.now()
    y, m = now.year, now.month

    payable = calc_payable_amount(student, y, m)
    paid = get_paid_amount(student_id, y, m)
    working = get_working_days(y, m)
    absent = get_absent_days(student_id, y, m)
    attended = get_attended_days(student_id, y, m)

    return jsonify({
        'student': student,
        'payable_amount': payable,
        'paid_amount': paid,
        'debt': max(payable - paid, 0),
        'working_days': working,
        'absent_days': absent,
        'attended_days': attended,
        'discount_percent': safe_int(student.get('discount_percent'), 0),
        'subsidy_amount': safe_int(student.get('subsidy_amount'), 0)
    })

# ─── Receipt ──────────────────────────────────────────────────────────────────

@app.route('/receipt/<payment_id>')
def receipt_page(payment_id):
    settings = load_settings()
    payment = None
    def _match(p):
        return p.get('id') == payment_id or p.get('order_id') == payment_id
    for kg in pc.load_kindergartens():
        kg_payments = pc.load_json('payments.json', kg['id'])
        payment = next((p for p in kg_payments if _match(p)), None)
        if payment:
            settings = pc.load_settings(kg['id'])
            break
    if not payment:
        default_payments = pc.load_json('payments.json', 'default')
        payment = next((p for p in default_payments if _match(p)), None)
    return render_template('receipt.html', payment=payment, settings=settings, payment_id=payment_id)

@app.route('/api/receipt/<payment_id>/qr')
def receipt_qr(payment_id):
    """Dynamic QR code generation — returns the QR image PNG directly."""
    base_url = request.host_url.rstrip('/')
    buf = generate_receipt_qr_bytes(payment_id, base_url)
    return send_file(buf, mimetype='image/png')

# ─── Parents portal ───────────────────────────────────────────────────────────

@app.route('/parent')
def parent_portal_select():
    kindergartens = [kg for kg in pc.load_kindergartens() if kg.get('status') == 'active']
    return render_template('parent_select.html', kindergartens=kindergartens)

@app.route('/parent/<kg_id>')
def parent_portal(kg_id):
    kg = pc.get_kindergarten(kg_id)
    if not kg:
        return redirect(url_for('parent_portal_select'))
    settings = load_settings(kg_id)
    return render_template('parent.html', settings=settings, kg_id=kg_id)

@app.route('/api/parent/lookup')
def parent_lookup():
    phone = request.args.get('phone', '').strip()
    student_id = request.args.get('student_id', '').strip()
    filter_kg = request.args.get('kg_id', '').strip()
    kg_id = None
    student = None

    if student_id:
        if filter_kg:
            students = pc.load_json('students.json', filter_kg)
            student = next((s for s in students if s['id'] == student_id), None)
            if student:
                kg_id = filter_kg
        else:
            for kg in pc.load_kindergartens():
                students = pc.load_json('students.json', kg['id'])
                student = next((s for s in students if s['id'] == student_id), None)
                if student:
                    kg_id = kg['id']
                    break
    elif phone:
        student, kg_id = find_student_by_phone(phone, filter_kg if filter_kg else None)

    if not student:
        return jsonify({'success': False, 'message': 'O\'quvchi topilmadi'})

    now = datetime.now()
    y, m = now.year, now.month

    payments = load_json('payments.json', kg_id)
    student_payments = [p for p in payments if p['student_id'] == student['id']]
    student_payments = sorted(student_payments, key=lambda x: x.get('date',''), reverse=True)

    att = load_json('attendance.json', kg_id)
    student_att = [a for a in att if a['student_id'] == student['id']]

    payable = calc_payable_amount(student, y, m, kg_id)
    paid = get_paid_amount(student['id'], y, m, kg_id)

    notifications = load_json('notifications.json', kg_id)
    student_notifs = [n for n in notifications if n.get('target') in ['all', student['id']]][-10:]

    discount = safe_int(student.get('discount_percent'), 0)
    subsidy = safe_int(student.get('subsidy_amount'), 0)

    return jsonify({
        'success': True,
        'student': {k: v for k, v in student.items() if k != 'telegram_chat_id'},
        'payments': student_payments[:20],
        'attendance': student_att[-30:],
        'current_month': {
            'payable': payable,
            'paid': paid,
            'debt': max(payable - paid, 0),
            'discount_percent': discount,
            'subsidy_amount': subsidy
        },
        'notifications': student_notifs
    })

# ─── Notifications ────────────────────────────────────────────────────────────

def add_notification(message, notif_type='info', target='all', kg_id=None):
    if target == 'platform':
        notifications = pc.load_json('platform_announcements.json')
        notifications.append({
            'id': str(uuid.uuid4())[:8],
            'title': 'Platforma',
            'message': message,
            'active': True,
            'created_at': datetime.now().isoformat()
        })
        pc.save_json('platform_announcements.json', notifications[-100:])
        return
    kid = kg_id or _current_kg_id() or 'default'
    notifications = load_json('notifications.json', kid)
    notifications.append({
        'id': str(uuid.uuid4())[:8],
        'message': message,
        'type': notif_type,
        'target': target,
        'read': False,
        'created_at': datetime.now().isoformat()
    })
    if len(notifications) > 200:
        notifications = notifications[-200:]
    save_json('notifications.json', notifications, kid)

@app.route('/notifications')
@login_required
def notifications_page():
    settings = load_settings()
    return render_template('notifications.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    notifications = load_json('notifications.json')
    return jsonify(sorted(notifications, key=lambda x: x.get('created_at',''), reverse=True))

@app.route('/api/notifications/unread-count')
@login_required
def unread_count():
    notifications = load_json('notifications.json')
    count = sum(1 for n in notifications if not n.get('read', False))
    return jsonify({'count': count})

@app.route('/api/notifications/mark-read', methods=['POST'])
@login_required
def mark_notifications_read():
    notifications = load_json('notifications.json')
    for n in notifications:
        n['read'] = True
    save_json('notifications.json', notifications)
    return jsonify({'success': True})

@app.route('/api/notifications/dismiss/<notif_id>', methods=['DELETE'])
@login_required
def dismiss_notification(notif_id):
    notifications = load_json('notifications.json')
    notifications = [n for n in notifications if n['id'] != notif_id]
    save_json('notifications.json', notifications)
    return jsonify({'success': True})

@app.route('/api/notifications/smart-check')
@login_required
def smart_check():
    students = load_json('students.json')
    today = date.today()
    now = datetime.now()
    y, m = now.year, now.month
    alerts = []

    for s in students:
        if s.get('status') != 'active':
            continue

        # Birthday check
        try:
            bd = datetime.strptime(s['birth_date'], '%Y-%m-%d')
            if bd.month == today.month and bd.day == today.day:
                alerts.append({'type': 'birthday', 'message': f"🎂 Bugun {s['first_name']} {s['last_name']} ning tug'ilgan kuni!", 'student': s})
        except Exception:
            pass

        # Debt check
        payable = calc_payable_amount(s, y, m)
        paid = get_paid_amount(s['id'], y, m)
        if paid < payable:
            debt = payable - paid
            alerts.append({'type': 'debt', 'message': f"💳 {s['first_name']} {s['last_name']}: qarzdorlik {debt:,} UZS", 'student': s})

        # Today absence
        att = load_json('attendance.json')
        today_str = today.isoformat()
        today_att = next((a for a in att if a['student_id'] == s['id'] and a['date'] == today_str), None)
        if today_att and today_att['status'] == 'absent':
            alerts.append({'type': 'absent', 'message': f"❌ {s['first_name']} {s['last_name']} bugun kelmadi", 'student': s})

    return jsonify(alerts)

# ─── Telegram ─────────────────────────────────────────────────────────────────

@app.route('/api/telegram/send', methods=['POST'])
@login_required
def send_telegram():
    data = request.get_json() or {}
    settings = load_settings()
    token = settings.get('bot_token', '')
    if not token:
        return jsonify({'success': False, 'message': 'Bot token sozlanmagan'})

    target = data.get('target', 'all')  # all / group / student_id
    message = data.get('message', '')
    students = load_json('students.json')

    if target == 'all':
        recipients = [s for s in students if s.get('status') == 'active' and s.get('telegram_chat_id')]
    elif target.startswith('group:'):
        group = target.split(':', 1)[1]
        recipients = [s for s in students if s.get('group') == group and s.get('telegram_chat_id')]
    else:
        recipients = [s for s in students if s['id'] == target and s.get('telegram_chat_id')]

    sent = 0
    failed = 0
    for s in recipients:
        chat_id = s['telegram_chat_id']
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        try:
            resp = http_requests.post(url, json={'chat_id': chat_id, 'text': message, 'parse_mode': 'HTML'}, timeout=5)
            if resp.status_code == 200:
                sent += 1
            else:
                failed += 1
        except Exception:
            failed += 1

    return jsonify({'success': True, 'sent': sent, 'failed': failed})

@app.route('/api/debt-reminders/send', methods=['POST'])
@login_required
def send_debt_reminders():
    data = request.get_json() or {}
    settings = load_settings()
    token = settings.get('bot_token', '')
    if not token:
        return jsonify({'success': False, 'message': 'Bot token sozlanmagan'})

    students = load_json('students.json')
    now = datetime.now()
    y, m = now.year, now.month
    template = data.get('message', '') or settings.get('reminder_message',
        'Hurmatli ona/ota, farzandingizning {name} uchun {month} oyidagi qarzdorligi {debt} {currency}. Iltimos to\'lovni amalga oshiring.\n\nSizning EduSoft')
    currency = settings.get('currency', 'UZS')
    month_names = ['', 'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun', 'Iyul', 'Avgust', 'Sentabr', 'Oktabr', 'Noyabr', 'Dekabr']

    sent = 0
    failed = 0
    for s in students:
        if s.get('status') != 'active':
            continue
        chat_id = s.get('telegram_chat_id')
        if not chat_id:
            continue
        payable = calc_payable_amount(s, y, m)
        paid = get_paid_amount(s['id'], y, m)
        debt = max(payable - paid, 0)
        if debt <= 0:
            continue
        name = f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()
        msg = template.replace('{name}', name).replace('{debt}', f"{debt:,}").replace('{currency}', currency).replace('{month}', month_names[m])
        try:
            resp = http_requests.post(f"https://api.telegram.org/bot{token}/sendMessage",
                json={'chat_id': chat_id, 'text': msg, 'parse_mode': 'HTML'}, timeout=5)
            if resp.status_code == 200:
                sent += 1
            else:
                failed += 1
        except Exception:
            failed += 1

    return jsonify({'success': True, 'sent': sent, 'failed': failed})

@app.route('/api/telegram/verify-channel', methods=['POST'])
@login_required
def verify_channel():
    data = request.get_json() or {}
    settings = load_settings()
    token = settings.get('bot_token', '')
    chat_id = data.get('chat_id', '')
    channel = data.get('channel', '')

    if not token:
        return jsonify({'success': False, 'message': 'Bot token sozlanmagan'})

    url = f"https://api.telegram.org/bot{token}/getChatMember"
    try:
        resp = http_requests.post(url, json={'chat_id': channel, 'user_id': chat_id}, timeout=5)
        result = resp.json()
        if result.get('ok'):
            status = result['result']['status']
            is_member = status in ['member', 'administrator', 'creator']
            return jsonify({'success': True, 'is_member': is_member})
    except Exception:
        pass
    return jsonify({'success': False, 'message': 'Tekshirib bo\'lmadi'})

def get_bot_session(chat_id):
    sessions = load_json('bot_sessions.json')
    sid = str(chat_id)
    for s in sessions:
        if s.get('chat_id') == sid:
            return s
    return {'chat_id': sid, 'step': 'start'}

def save_bot_session(chat_id, step, extra=None):
    sessions = load_json('bot_sessions.json')
    sid = str(chat_id)
    found = False
    for i, s in enumerate(sessions):
        if s.get('chat_id') == sid:
            sessions[i]['step'] = step
            if extra:
                sessions[i].update(extra)
            found = True
            break
    if not found:
        rec = {'chat_id': sid, 'step': step}
        if extra:
            rec.update(extra)
        sessions.append(rec)
    if len(sessions) > 500:
        sessions = sessions[-500:]
    save_json('bot_sessions.json', sessions)

def link_telegram_to_student(chat_id, phone, kg_id=None):
    student, student_kg_id = find_student_by_phone(phone, kg_id)
    if not student:
        return None, 'not_found', None
    actual_kg_id = student_kg_id or kg_id or 'default'
    students = load_json('students.json', actual_kg_id)
    for i, s in enumerate(students):
        if s['id'] == student['id']:
            students[i]['telegram_chat_id'] = str(chat_id)
            save_json('students.json', students, actual_kg_id)
            return students[i], 'linked', actual_kg_id
    return None, 'error', None

def telegram_app_context():
    site = os.environ.get('SITE_URL', '').strip()
    try:
        if not site and request:
            site = request.host_url.rstrip('/')
    except RuntimeError:
        pass
    return {
        'pc': pc,
        'load_settings': load_settings,
        'normalize_phone': normalize_phone,
        'get_bot_session': get_bot_session,
        'save_bot_session': save_bot_session,
        'link_telegram': link_telegram_to_student,
        'notify_admin': notify_admin,
        'site_url': site,
    }

@app.route('/api/telegram/webhook', defaults={'kg_id': None}, methods=['POST'])
@app.route('/api/telegram/webhook/<kg_id>', methods=['POST'])
def telegram_webhook(kg_id):
    update = request.get_json(force=True) or {}
    if kg_id:
        # Validate secret token matches kg_id
        header_secret = request.headers.get('X-Telegram-Bot-Api-Secret-Token', '')
        if header_secret and header_secret != kg_id:
            return jsonify({'error': 'unauthorized'}), 403
        s = load_settings(kg_id)
        if s.get('bot_token'):
            session['kindergarten_id'] = kg_id
            if 'role' not in session:
                session['role'] = 'webhook'
            try:
                from telegram_bot import handle_update
                handle_update(update, telegram_app_context(), bot_kg_id=kg_id)
            except Exception:
                pass
        return jsonify({'ok': True})
    # Legacy: no kg_id in URL — use global secret or iterate all
    secret = os.environ.get('WEBHOOK_SECRET', '')
    if secret and request.headers.get('X-Telegram-Bot-Api-Secret-Token') != secret:
        return jsonify({'error': 'unauthorized'}), 403
    kgs = pc.load_kindergartens()
    if not isinstance(kgs, list):
        kgs = []
    for kg in kgs:
        kg_id = kg.get('id')
        if not kg_id:
            continue
        s = load_settings(kg_id)
        if s.get('bot_token'):
            session['kindergarten_id'] = kg_id
            if 'role' not in session:
                session['role'] = 'webhook'
            try:
                from telegram_bot import handle_update
                handle_update(update, telegram_app_context(), bot_kg_id=kg_id)
            except Exception:
                pass
            return jsonify({'ok': True})
    return jsonify({'ok': True})

@app.route('/api/telegram/super-webhook', methods=['POST'])
def super_telegram_webhook():
    plat = pc.load_platform()
    super_token = plat.get('super_bot_token', '').strip()
    if not super_token:
        return jsonify({'ok': True})
    secret = os.environ.get('SUPER_WEBHOOK_SECRET', '')
    if secret and request.headers.get('X-Telegram-Bot-Api-Secret-Token') != secret:
        return jsonify({'error': 'unauthorized'}), 403
    try:
        update = request.get_json(force=True) or {}
        from telegram_bot import handle_super_update
        handle_super_update(update, plat, telegram_app_context())
    except Exception as ex:
        print(f"[SuperWebhook] xato: {ex}")
    return jsonify({'ok': True})

@app.route('/api/telegram/status')
@login_required
def telegram_status():
    settings = load_settings()
    token = settings.get('bot_token', '')
    if not token:
        return jsonify({'success': False, 'message': 'Bot token yo\'q', 'has_token': False})
    from telegram_bot import get_webhook_info, get_token as tb_token
    info = get_webhook_info(token)
    wh = info.get('result', {}) if info.get('ok') else {}
    polling_local = not os.environ.get('VERCEL') and os.environ.get('TELEGRAM_POLLING', '1') != '0'
    return jsonify({
        'success': True,
        'has_token': True,
        'webhook_url': wh.get('url') or '',
        'webhook_active': bool(wh.get('url')),
        'polling_local': polling_local,
        'admin_chat_ok': is_valid_admin_chat_id(settings.get('admin_telegram_chat_id', '')),
        'hint': 'Mahalliy: python app.py — bot avtomatik polling. Vercel: Webhook o\'rnating.'
    })

@app.route('/api/telegram/set-webhook', methods=['POST'])
@login_required
def set_telegram_webhook():
    settings = load_settings()
    token = settings.get('bot_token', '')
    if not token:
        return jsonify({'success': False, 'message': 'Bot token sozlanmagan'})
    site_url = (request.get_json() or {}).get('site_url', '').strip()
    if not site_url:
        site_url = request.host_url.rstrip('/')
    webhook_url = f"{site_url}/api/telegram/webhook"
    url = f"https://api.telegram.org/bot{token}/setWebhook"
    payload = {'url': webhook_url, 'drop_pending_updates': True}
    secret = os.environ.get('WEBHOOK_SECRET', '')
    if secret:
        payload['secret_token'] = secret
    try:
        resp = http_requests.post(url, json=payload, timeout=10)
        data = resp.json()
        if data.get('ok'):
            return jsonify({'success': True, 'url': webhook_url})
        return jsonify({'success': False, 'message': data.get('description', 'Xatolik')})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# ─── Registration (ota-ona) ───────────────────────────────────────────────────

@app.route('/register')
def register_page():
    settings = load_settings()
    return render_template('register.html', settings=settings)

@app.route('/api/register', methods=['POST'])
def submit_registration():
    data = request.get_json() or {}
    phone = normalize_phone(data.get('parent_phone', ''))
    if not phone or len(phone) < 9:
        return jsonify({'success': False, 'message': 'Telefon raqam noto\'g\'ri'})
    fname = (data.get('child_first_name') or '').strip()
    lname = (data.get('child_last_name') or '').strip()
    parent_name = (data.get('parent_name') or '').strip()
    if not fname or not lname or not parent_name:
        return jsonify({'success': False, 'message': 'Barcha majburiy maydonlarni to\'ldiring'})

    existing, _ = find_student_by_phone(phone)
    if existing:
        return jsonify({'success': False, 'message': 'Bu telefon allaqachon tizimda mavjud. Ota-ona portaliga kiring.'})

    regs = load_json('registrations.json')
    for r in regs:
        if r.get('status') == 'pending' and normalize_phone(r.get('parent_phone', '')) == phone:
            return jsonify({'success': False, 'message': 'Arizangiz allaqachon ko\'rib chiqilmoqda'})

    reg = {
        'id': 'REG-' + str(uuid.uuid4())[:8].upper(),
        'child_first_name': fname,
        'child_last_name': lname,
        'birth_date': data.get('birth_date', ''),
        'parent_name': parent_name,
        'parent_phone': normalize_phone(phone) or phone,
        'group_preference': (data.get('group_preference') or '').strip(),
        'notes': (data.get('notes') or '').strip(),
        'status': 'pending',
        'created_at': datetime.now().isoformat()
    }
    regs.append(reg)
    save_json('registrations.json', regs)
    notify_admin(
        f"📝 Yangi ro'yxat arizasi: {fname} {lname} — {parent_name} ({phone})",
        'info'
    )
    return jsonify({'success': True, 'message': 'Arizangiz qabul qilindi. Admin tez orada bog\'lanadi.', 'id': reg['id']})

@app.route('/api/registrations', methods=['GET'])
@login_required
def get_registrations():
    regs = load_json('registrations.json')
    pending = [r for r in regs if r.get('status') == 'pending']
    return jsonify(sorted(pending, key=lambda x: x.get('created_at', ''), reverse=True))

@app.route('/api/registrations/<reg_id>/approve', methods=['POST'])
@login_required
def approve_registration(reg_id):
    regs = load_json('registrations.json')
    reg = next((r for r in regs if r['id'] == reg_id), None)
    if not reg:
        return jsonify({'success': False, 'message': 'Ariza topilmadi'})

    data = request.get_json() or {}
    students = load_json('students.json')
    student = {
        'id': 'STU-' + str(uuid.uuid4())[:8].upper(),
        'first_name': reg['child_first_name'],
        'last_name': reg['child_last_name'],
        'birth_date': reg.get('birth_date', ''),
        'parent_name': reg['parent_name'],
        'parent_phone': reg['parent_phone'],
        'group': data.get('group', reg.get('group_preference', '')).strip(),
        'monthly_fee': safe_int(data.get('monthly_fee'), 0),
        'payment_due_day': safe_int(data.get('payment_due_day'), 1),
        'join_date': date.today().isoformat(),
        'status': 'active',
        'image': '',
        'telegram_chat_id': '',
        'notes': reg.get('notes', '')
    }
    students.append(student)
    save_json('students.json', students)

    for i, r in enumerate(regs):
        if r['id'] == reg_id:
            regs[i]['status'] = 'approved'
            regs[i]['student_id'] = student['id']
            break
    save_json('registrations.json', regs)
    add_notification(f"Ro'yxat tasdiqlandi: {student['first_name']} {student['last_name']}", 'info')
    return jsonify({'success': True, 'student': student})

@app.route('/api/registrations/<reg_id>/reject', methods=['POST'])
@login_required
def reject_registration(reg_id):
    regs = load_json('registrations.json')
    for i, r in enumerate(regs):
        if r['id'] == reg_id:
            regs[i]['status'] = 'rejected'
            save_json('registrations.json', regs)
            return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Ariza topilmadi'})

# ─── Complaints / feedback ────────────────────────────────────────────────────

@app.route('/api/parent/complaint', methods=['POST'])
def submit_complaint():
    data = request.get_json() or {}
    phone = normalize_phone(data.get('phone', ''))
    student_id = (data.get('student_id') or '').strip()
    message = (data.get('message') or '').strip()
    category = (data.get('category') or 'shikoyat').strip()

    if not message or len(message) < 5:
        return jsonify({'success': False, 'message': 'Xabar kamida 5 belgidan iborat bo\'lishi kerak'})

    student = None
    kg_id = None
    if student_id:
        for kg in pc.load_kindergartens():
            students = pc.load_json('students.json', kg['id'])
            student = next((s for s in students if s['id'] == student_id), None)
            if student:
                kg_id = kg['id']
                break
    elif phone:
        student, kg_id = find_student_by_phone(phone)

    complaints = load_json('complaints.json', kg_id or 'default')
    complaint = {
        'id': 'CMP-' + str(uuid.uuid4())[:8].upper(),
        'student_id': student['id'] if student else '',
        'student_name': f"{student['first_name']} {student['last_name']}" if student else 'Noma\'lum',
        'parent_name': student.get('parent_name', '') if student else data.get('parent_name', ''),
        'parent_phone': phone or (student.get('parent_phone', '') if student else ''),
        'category': category,
        'message': message,
        'status': 'new',
        'created_at': datetime.now().isoformat()
    }
    complaints.append(complaint)
    if len(complaints) > 300:
        complaints = complaints[-300:]
    save_json('complaints.json', complaints, kg_id or 'default')

    cat_label = {'shikoyat': 'Shikoyat', 'taklif': 'Taklif', 'norozilik': 'Norozilik'}.get(category, category)
    notify_admin(
        f"📩 {cat_label}: {complaint['student_name']}\n"
        f"📞 {complaint['parent_phone']}\n"
        f"💬 {message[:200]}",
        'warning'
    )
    return jsonify({'success': True, 'message': 'Xabaringiz adminga yetkazildi. Rahmat!'})

@app.route('/api/complaints', methods=['GET'])
@login_required
def get_complaints():
    complaints = load_json('complaints.json')
    return jsonify(sorted(complaints, key=lambda x: x.get('created_at', ''), reverse=True))

@app.route('/api/complaints/<complaint_id>/status', methods=['POST'])
@login_required
def update_complaint_status(complaint_id):
    data = request.get_json() or {}
    status = data.get('status', 'read')
    complaints = load_json('complaints.json')
    for i, c in enumerate(complaints):
        if c['id'] == complaint_id:
            complaints[i]['status'] = status
            save_json('complaints.json', complaints)
            return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Topilmadi'})

# ─── Settings ─────────────────────────────────────────────────────────────────

@app.route('/settings')
@login_required
def settings_page():
    settings = load_settings()
    kg_id = _current_kg_id() or ''
    return render_template('settings.html', settings=settings, kg_id=kg_id, admin_name=session.get('admin_name'),
                           is_super=session.get('role') == 'super')

@app.route('/api/settings/public')
def get_public_settings():
    s = load_settings()
    return jsonify({
        'name': s.get('name', ''),
        'currency': s.get('currency', 'UZS'),
        'logo': s.get('logo', ''),
        'tagline': s.get('tagline', ''),
        'bot_username': s.get('bot_username', ''),
        'payment_card': s.get('payment_card', ''),
        'payment_provider': s.get('payment_provider', 'manual_card'),
        'payment_merchant_id': s.get('payment_merchant_id', '')
    })

# ─── Public Pricing / Tariffs ──────────────────────────────────────────────────
@app.route('/api/public-fees')
def public_fees():
    s = load_settings()
    group_fees = s.get('group_fees', {})
    result = []
    if group_fees:
        for gname, fee in group_fees.items():
            result.append({
                'name': gname,
                'fee': fee,
                'fee_min': fee,
                'fee_max': fee,
                'fee_same': True,
                'count': 0,
                'configured': True
            })
        result.sort(key=lambda x: x['fee'])
    else:
        students = load_json('students.json')
        groups = {}
        for stu in students:
            g = stu.get('group', '') or "Guruhsiz"
            fee = stu.get('monthly_fee', 0)
            if g not in groups:
                groups[g] = {'fees': set(), 'count': 0}
            groups[g]['fees'].add(fee)
            groups[g]['count'] += 1
        for gname, gdata in groups.items():
            fees = sorted(gdata['fees'])
            result.append({
                'name': gname,
                'fee_min': min(fees),
                'fee_max': max(fees),
                'fee_same': len(fees) == 1,
                'fee': fees[0],
                'count': gdata['count'],
                'configured': False
            })
        result.sort(key=lambda x: x['fee'])
    return jsonify({
        'groups': result,
        'name': s.get('name', ''),
        'currency': s.get('currency', 'UZS'),
        'logo': s.get('logo', ''),
        'tagline': s.get('tagline', ''),
        'payment_card': s.get('payment_card', ''),
        'payment_provider': s.get('payment_provider', 'manual_card')
    })

@app.route('/investor')
def investor_pitch():
    return send_file('investor_pitch.html')

@app.route('/pricing')
def pricing_page():
    settings = load_settings()
    return render_template('pricing.html', settings=settings)

@app.route('/api/payment-checks/submit', methods=['POST'])
def submit_payment_check():
    """Parent submits a payment check from the portal."""
    data = request.get_json() or {}
    student_id = data.get('student_id', '')
    student_name = data.get('student_name', '')
    parent_phone = data.get('parent_phone', '')
    amount = safe_int(data.get('amount'), 0)
    photo_url = data.get('photo_url', '')
    month = data.get('month', datetime.now().strftime('%Y-%m'))

    if not student_id or not amount or not photo_url:
        return jsonify({'success': False, 'message': 'Barcha maydonlarni to\'ldiring'})

    # Find kg_id + parent_chat_id from student
    kg_id = None
    parent_chat_id = ''
    for kg in pc.load_kindergartens():
        students = pc.load_json('students.json', kg['id'])
        for s in students:
            if s['id'] == student_id:
                kg_id = kg['id']
                parent_chat_id = s.get('telegram_chat_id', '')
                break
        if kg_id:
            break

    if not kg_id:
        kg_id = _current_kg_id() or 'default'

    checks = load_json('payment_checks.json', kg_id)
    check_id = 'CHK-' + str(uuid.uuid4())[:8].upper()
    receipt_id = 'RCP-' + str(uuid.uuid4())[:8].upper()

    # Create pending payment record immediately
    payments = load_json('payments.json', kg_id)
    pending_payment = {
        'id': receipt_id,
        'student_id': student_id,
        'student_name': student_name,
        'amount': amount,
        'date': datetime.now().strftime('%Y-%m-%d'),
        'month': month,
        'type': 'partial',
        'category': 'tuition',
        'status': 'pending',
        'note': f"Chek orqali to'lov",
        'admin_name': 'Kutilmoqda',
        'created_at': datetime.now().isoformat()
    }
    payments.append(pending_payment)
    save_json('payments.json', payments, kg_id)

    check = {
        'id': check_id,
        'receipt_id': receipt_id,
        'student_id': student_id,
        'student_name': student_name,
        'parent_phone': parent_phone,
        'parent_chat_id': parent_chat_id,
        'amount': amount,
        'photo_url': photo_url,
        'status': 'pending',
        'created_at': datetime.now().isoformat()
    }
    checks.append(check)
    save_json('payment_checks.json', checks, kg_id)

    # Send photo to admin's Telegram with approve/reject buttons
    settings = load_settings(kg_id)
    token = settings.get('bot_token', '')
    admin_chat = settings.get('admin_telegram_chat_id', '')
    if token and admin_chat:
        try:
            from telegram_bot import send_photo
            esc = html.escape
            admin_msg = (
                f"📩 <b>Yangi to'lov cheki</b>\n\n"
                f"👤 Ota-ona: <b>{esc(student_name)}</b>\n"
                f"📞 Telefon: {esc(parent_phone)}\n"
                f"💵 Summa: <b>{amount:,} UZS</b>\n"
                f"🆔 ID: <code>{check_id}</code>\n\n"
                f"Chek rasmini tekshirib, tasdiqlang!"
            )
            approve_btn = {
                'inline_keyboard': [[
                    {'text': '✅ Tasdiqlash', 'callback_data': f'approve_pay_{kg_id}_{check_id}'},
                    {'text': '❌ Rad etish', 'callback_data': f'reject_pay_{kg_id}_{check_id}'}
                ]]
            }
            send_photo(token, admin_chat, photo_url, admin_msg, reply_markup=approve_btn)
        except Exception:
            pass

    # Also notify via in-app notification as fallback
    notify_admin(
        f"📩 Yangi to'lov cheki: {student_name} — {amount:,} UZS (ID: {check_id})",
        'payment', kg_id=kg_id
    )

    return jsonify({'success': True, 'check_id': check_id, 'receipt_id': receipt_id, 'check': check})

@app.route('/api/payment-checks', methods=['GET'])
@login_required
def get_payment_checks():
    checks = load_json('payment_checks.json')
    status = request.args.get('status', '')
    result = sorted(checks, key=lambda x: x.get('created_at', ''), reverse=True)
    if status:
        result = [c for c in result if c.get('status') == status]
    return jsonify(result)

@app.route('/api/payment-checks/<check_id>/status', methods=['POST'])
@login_required
def update_payment_check(check_id):
    data = request.get_json() or {}
    status = data.get('status', 'approved')
    checks = load_json('payment_checks.json')

    for i, c in enumerate(checks):
        if c['id'] == check_id:
            students = load_json('students.json')
            student = next((s for s in students if s['id'] == c['student_id']), None)

            now = datetime.now()
            kg_id_local = _current_kg_id() or 'default'
            receipt_id = c.get('receipt_id', 'RCP-' + str(uuid.uuid4())[:8].upper())

            # Update existing pending payment
            payments = load_json('payments.json')
            found = False
            for p in payments:
                if p['id'] == receipt_id:
                    if status == 'approved':
                        p['status'] = 'paid'
                        p['admin_name'] = session.get('admin_name', 'Admin')
                        p['date'] = now.strftime('%Y-%m-%d')
                    else:
                        p['status'] = 'cancelled'
                        p['admin_name'] = session.get('admin_name', 'Admin')
                    found = True
                    break
            if not found:
                # Fallback: create new payment (for legacy checks without receipt_id)
                payment = {
                    'id': receipt_id,
                    'student_id': c['student_id'],
                    'student_name': c['student_name'],
                    'amount': c['amount'],
                    'date': now.strftime('%Y-%m-%d'),
                    'month': now.strftime('%Y-%m'),
                    'type': 'check',
                    'category': 'tuition',
                    'status': 'paid' if status == 'approved' else 'cancelled',
                    'note': f"Chek orqali to'lov (ID: {check_id})",
                    'admin_name': session.get('admin_name', 'Admin'),
                    'created_at': now.isoformat()
                }
                payments.append(payment)
            save_json('payments.json', payments)

            if status == 'approved':
                checks[i]['status'] = 'approved'
                checks[i]['approved_by'] = session.get('admin_name', 'Admin')
                checks[i]['approved_at'] = now.isoformat()
                toast_text = f"✅ To'lov tasdiqlandi: {c['student_name']} — {c['amount']:,} UZS"
            else:
                checks[i]['status'] = 'rejected'
                checks[i]['rejected_by'] = session.get('admin_name', 'Admin')
                toast_text = f"❌ To'lov rad etildi: {c['student_name']}"

            # Notify parent via telegram (both approved & rejected)
            parent_chat = c.get('parent_chat_id', '')
            if not parent_chat and student:
                parent_chat = student.get('telegram_chat_id', '')
            if parent_chat:
                sname = html.escape(c['student_name'])
                if status == 'approved':
                    base = request.host_url.rstrip('/')
                    receipt_link = f"{base}/receipt/{receipt_id}"
                    qr_link = f"{base}/api/receipt/{receipt_id}/qr"
                    msg = (
                        f"✅ <b>To'lov tasdiqlandi!</b>\n\n"
                        f"👨‍🎓 Farzand: <b>{sname}</b>\n"
                        f"💵 Summa: <b>{c['amount']:,} UZS</b>\n"
                        f"🧾 Chek ID: <code>{receipt_id}</code>\n"
                        f"📅 Sana: {now.strftime('%d.%m.%Y')}\n\n"
                        f"📄 Chek: {receipt_link}\n"
                        f"📱 QR: {qr_link}\n\n"
                        f"Rahmat! 🙏"
                    )
                else:
                    msg = (
                        f"❌ <b>To'lov rad etildi</b>\n\n"
                        f"👨‍🎓 Farzand: <b>{sname}</b>\n"
                        f"💵 Summa: <b>{c['amount']:,} UZS</b>\n"
                        f"ℹ️ Iltimos, admin bilan bog'laning yoki qayta to'lov qiling."
                    )
                send_telegram_message(parent_chat, msg, kg_id=kg_id_local)

            save_json('payment_checks.json', checks)
            add_notification(toast_text, 'payment')
            return jsonify({'success': True, 'status': status})

    return jsonify({'success': False, 'message': 'Chek topilmadi'}), 404

@app.route('/api/payment-checks/pending-count')
@login_required
def pending_checks_count():
    checks = load_json('payment_checks.json')
    count = sum(1 for c in checks if c.get('status') == 'pending')
    return jsonify({'count': count})

# ─── Auto-payment (TezCheck) ─────────────────────────────────────────────────

@app.route('/api/auto-payment/create', methods=['POST'])
@login_required
def auto_payment_create():
    """Create a tezcheck.uz payment invoice and return pay_url."""
    data = request.get_json() or {}
    amount = safe_int(data.get('amount'), 0)
    if amount < 100:
        return jsonify({'success': False, 'message': 'Minimal summa 100 so\'m'})
    kg_id = _current_kg_id()
    settings = load_settings()
    if not settings.get('auto_payment'):
        return jsonify({'success': False, 'message': 'Avtomatik to\'lov o\'chirilgan'})
    result = pc.create_invoice(amount)
    if result:
        return jsonify({
            'success': True,
            'order_id': result.get('order_id'),
            'pay_url': result.get('pay_url'),
            'data': result
        })
    return jsonify({'success': False, 'message': 'TezCheck API xatolik'})

@app.route('/api/auto-payment/status/<int:order_id>')
@login_required
def auto_payment_status(order_id):
    """Check a tezcheck.uz invoice status."""
    result = pc.check_invoice(order_id)
    if result and result.get('payment'):
        return jsonify({'success': True, 'payment': result['payment']})
    return jsonify({'success': False, 'message': 'Buyurtma topilmadi'})

@app.route('/api/kg-balance', methods=['GET'])
@login_required
def get_kg_balance():
    """Get current kindergarten balance."""
    settings = load_settings()
    return jsonify({'balance': settings.get('balance', 0)})

@app.route('/api/kg-balance/add', methods=['POST'])
@login_required
def add_kg_balance():
    """Admin adds or subtracts funds from KG balance."""
    data = request.get_json() or {}
    amount = safe_int(data.get('amount'), 0)
    note = data.get('note', '')
    kg_id = _current_kg_id()
    settings = load_settings()
    old_balance = settings.get('balance', 0)
    settings['balance'] = old_balance + amount
    save_settings_data(settings, kg_id)
    _audit('balance_changed',
           f"Balans o\'zgartirildi: {old_balance:,} → {settings['balance']:,} so\'m ({'+' if amount >= 0 else ''}{amount:,})"
           f"{' — ' + note if note else ''}")
    return jsonify({'success': True, 'balance': settings['balance']})

@app.route('/api/kg-balance/history', methods=['GET'])
@login_required
def kg_balance_history():
    """Return all balance-related events (payments + manual adjustments)."""
    kg_id = _current_kg_id()
    payments = pc.load_payments(kg_id)
    history = []
    for p in payments:
        entry = {
            'type': 'payment',
            'amount': p.get('amount', 0),
            'status': p.get('status', 'pending'),
            'parent_name': p.get('parent_name', ''),
            'parent_phone': p.get('parent_phone', ''),
            'created_at': p.get('created_at', ''),
            'paid_at': p.get('paid_at', ''),
            'transaction_id': p.get('transaction_id', ''),
            'order_id': p.get('order_id', ''),
            'student_id': p.get('student_id', ''),
            'month': p.get('month', '')
        }
        history.append(entry)
    history.sort(key=lambda x: x.get('paid_at') or x.get('created_at') or '', reverse=True)
    return jsonify({'success': True, 'history': history})

@app.route('/api/kg-balance/withdrawals', methods=['GET'])
@login_required
def kg_withdrawals():
    """List all withdrawal requests for current KG."""
    kg_id = _current_kg_id()
    withdrawals = pc.load_withdrawals(kg_id)
    return jsonify({'success': True, 'withdrawals': withdrawals})

@app.route('/api/kg-balance/withdraw', methods=['POST'])
@login_required
def kg_withdraw():
    """Create a withdrawal request. Returns a unique code."""
    data = request.get_json() or {}
    amount = safe_int(data.get('amount'), 0)
    recipient_name = data.get('recipient_name', '').strip()
    recipient_phone = data.get('recipient_phone', '').strip()
    recipient_card = data.get('recipient_card', '').strip()
    if amount < 100:
        return jsonify({'success': False, 'message': 'Minimal summa 100 so\'m'})
    kg_id = _current_kg_id()
    settings = load_settings()
    if amount > settings.get('balance', 0):
        return jsonify({'success': False, 'message': 'Balansda yetarli mablag\' yo\'q'})
    w = pc.add_withdrawal(kg_id, amount, recipient_name, recipient_phone, recipient_card)
    _audit('withdrawal_created',
           f"Yechib olish so\'rovi yaratildi: {amount:,} so\'m (kod: {w['code']})"
           f"{' — ' + recipient_name if recipient_name else ''}")
    return jsonify({'success': True, 'withdrawal': w})

@app.route('/api/kg-balance/withdraw/confirm', methods=['POST'])
@login_required
def kg_withdraw_confirm():
    """Confirm a withdrawal by code and deduct from balance."""
    data = request.get_json() or {}
    code = data.get('code', '').strip().upper()
    if not code:
        return jsonify({'success': False, 'message': 'Kod majburiy'})
    kg_id = _current_kg_id()
    w = pc.confirm_withdrawal(kg_id, code)
    if not w:
        return jsonify({'success': False, 'message': 'Kod topilmadi yoki allaqachon tasdiqlangan'})
    # Deduct from balance
    settings = load_settings()
    old_balance = settings.get('balance', 0)
    settings['balance'] = old_balance - w['amount']
    save_settings_data(settings, kg_id)
    _audit('withdrawal_confirmed',
           f"Yechib olish tasdiqlandi: {w['amount']:,} so'm (kod: {code})"
           f" — qabul qiluvchi: {w.get('recipient_name', '')} {w.get('recipient_phone', '')}")
    return jsonify({'success': True, 'withdrawal': w, 'balance': settings['balance']})

@app.route('/api/auto-payment/toggle', methods=['POST'])
@login_required
def toggle_auto_payment():
    """Toggle auto-payment on/off for current KG."""
    data = request.get_json() or {}
    enabled = bool(data.get('enabled', False))
    kg_id = _current_kg_id()
    settings = load_settings()
    settings['auto_payment'] = enabled
    save_settings_data(settings, kg_id)
    _audit('auto_payment_toggled',
           f"Avtomatik to\'lov {'yoqildi' if enabled else 'o\'chirildi'}")
    return jsonify({'success': True, 'auto_payment': enabled})

# ─── TezCheck public payment page for parents ─────────────────────────────────

@app.route('/tezcheck-pay/<kg_id>')
def tezcheck_pay_page(kg_id):
    """Public payment page for parents. Shows payment form and generates invoice."""
    kg = pc.get_kindergarten(kg_id)
    if not kg:
        return render_template('tezcheck_pay.html', found=False, error="Bog'cha topilmadi", kg=None)
    settings = pc.load_settings(kg_id)
    return render_template('tezcheck_pay.html', found=True, kg=kg, settings=settings)

@app.route('/api/tezcheck/create-payment', methods=['POST'])
def tezcheck_create_payment():
    """Create a tezcheck.uz invoice and return pay_url."""
    import traceback
    try:
        data = request.get_json() or {}
        kg_id = data.get('kg_id', '')
        amount = safe_int(data.get('amount'), 0)
        description = data.get('description', '')
        parent_name = data.get('parent_name', '')
        parent_phone = data.get('parent_phone', '')
        if amount < 100:
            return jsonify({'success': False, 'message': 'Minimal summa 100 so\'m'})
        kg = pc.get_kindergarten(kg_id)
        if not kg:
            return jsonify({'success': False, 'message': 'Bog\'cha topilmadi'})
        result = pc.create_invoice(amount, description, kg_id, parent_name, parent_phone)
        if result:
            return jsonify({
                'success': True,
                'order_id': result.get('order_id'),
                'pay_url': result.get('pay_url')
            })
        return jsonify({'success': False, 'message': 'TezCheck API xatolik'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()})

@app.route('/api/tezcheck/check-status/<order_id>')
def tezcheck_check_status(order_id):
    """Check tezcheck.uz invoice status (polling for clients)."""
    result = pc.check_invoice(str(order_id))
    if result and result.get('payment'):
        p = result['payment']
        if p.get('status') == 'paid':
            amount = int(p.get('amount', 0))
            for kg in pc.load_kindergartens():
                kg_id = kg.get('id', '')
                payments = pc.load_payments(kg_id)
                for payment in payments:
                    if str(payment.get('order_id', '')) == str(order_id):
                        if payment.get('status') != 'paid':
                            payment['status'] = 'paid'
                            payment['paid_at'] = datetime.now(timezone.utc).isoformat() + 'Z'
                            payment['transaction_id'] = p.get('transaction_id')
                            pc.save_payments(kg_id, payments)
                            settings = pc.load_settings(kg_id)
                            settings['balance'] = settings.get('balance', 0) + amount
                            pc.save_settings(kg_id, settings)
                        return jsonify({'success': True, 'payment': p})
        else:
            return jsonify({'success': True, 'payment': p})
    return jsonify({'success': False, 'message': 'To\'lov topilmadi'})

@app.route('/api/tezcheck/webhook', methods=['POST'])
def tezcheck_webhook():
    """Webhook for tezcheck.uz payment callbacks."""
    data = request.get_json() or {}
    seen_id = str(data.get('order_id') or data.get('id') or '')
    if seen_id:
        result = pc.check_invoice(seen_id)
        if result and result.get('payment'):
            p = result['payment']
            if p.get('status') == 'paid':
                amount = int(p.get('amount', 0))
                # Find which kg this payment belongs to
                for kg in pc.load_kindergartens():
                    kg_id = kg.get('id', '')
                    payments = pc.load_payments(kg_id)
                    for payment in payments:
                        if str(payment.get('order_id', '')) == str(seen_id):
                            if payment.get('status') == 'paid':
                                return jsonify({'ok': True})
                            payment['status'] = 'paid'
                            payment['paid_at'] = datetime.now(timezone.utc).isoformat() + 'Z'
                            payment['transaction_id'] = p.get('transaction_id')
                            pc.save_payments(kg_id, payments)
                            settings = pc.load_settings(kg_id)
                            settings['balance'] = settings.get('balance', 0) + amount
                            pc.save_settings(kg_id, settings)
                            return jsonify({'ok': True})
    return jsonify({'ok': True})

@app.route('/api/debug/db-state')
def debug_db_state():
    import os
    keys = pc.db.list_keys('') if pc.db else []
    rows = []
    for k in sorted(keys):
        val = pc.db.get(k)
        if isinstance(val, list):
            rows.append({'key': k, 'type': 'list', 'count': len(val), 'ids': [v.get('id','') for v in val if isinstance(v, dict)][:5]})
        elif isinstance(val, dict):
            rows.append({'key': k, 'type': 'dict', 'keys': list(val.keys())[:10]})
        else:
            rows.append({'key': k, 'type': type(val).__name__, 'val': str(val)[:100]})
    env_vars = {k: '***' if 'PASS' in k.upper() or 'KEY' in k.upper() else os.environ.get(k,'') 
                for k in ['DATABASE_URL','POSTGRES_URL','POSTGRES_URL_NON_POOLING','POSTGRES_PRISMA_URL','STORAGE_URL','STORAGE_URL_NON_POOLING','VERCEL_URL'] 
                if os.environ.get(k)}
    return jsonify({'db_connected': pc.db is not None, 'key_count': len(keys), 'keys': rows, 'env_vars': env_vars})

@app.route('/api/debug/checkout-test')
def debug_checkout_test():
    """Test checkout.uz API call from Vercel and return full response."""
    import urllib.request, json, traceback
    result = {}
    try:
        try:
            with urllib.request.urlopen('https://api.ipify.org?format=json', timeout=10) as r:
                result['outbound_ip'] = json.loads(r.read())['ip']
        except:
            result['outbound_ip'] = 'unknown'
        
        plat = pc.load_platform()
        result['api_key_found'] = bool(plat.get('checkout_api_key'))
        result['provider'] = plat.get('payment_api_provider', 'not set')
        
        key = (plat.get('checkout_api_key') or '').strip()
        if not key:
            result['api_key_masked'] = '(empty)'
        elif len(key) > 6:
            result['api_key_masked'] = key[:4] + '*' * (len(key)-8) + key[-4:]
        else:
            result['api_key_masked'] = key[:2] + '***'
        
        if key:
            tests = {}
            ua = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
            # 1) Minimal request — no custom headers except Content-Type
            try:
                body_data = json.dumps({"amount": 10000, "description": "Test"}).encode()
                req = urllib.request.Request(
                    'https://checkout.uz/api/v1/create_payment',
                    data=body_data, method='POST')
                req.add_header('Content-Type', 'application/json')
                req.add_header('User-Agent', ua)
                with urllib.request.urlopen(req, timeout=20) as resp:
                    tests['minimal'] = {'status': f'HTTP {resp.getcode()}', 'response': json.loads(resp.read())}
            except urllib.error.HTTPError as e:
                body = e.read()
                tests['minimal'] = {'code': e.code, 'reason': e.reason, 'body': body.decode('utf-8', errors='replace')[:500]}
            except Exception as e:
                tests['minimal'] = {'error': str(e)[:200]}
            # 2) Bearer token
            methods = [
                ('Bearer', f'Bearer {key}'),
                ('X-API-Key', key),
                ('apikey in body', None),
                ('Basic', 'Basic ' + __import__('base64').b64encode(f'{key}:'.encode()).decode()),
            ]
            for label, auth_val in methods:
                body_data = json.dumps({"amount": 10000, "description": "Test"}).encode()
                if label == 'apikey in body':
                    body_data = json.dumps({"api_key": key, "amount": 10000, "description": "Test"}).encode()
                req = urllib.request.Request(
                    'https://checkout.uz/api/v1/create_payment',
                    data=body_data, method='POST')
                req.add_header('Content-Type', 'application/json')
                req.add_header('User-Agent', ua)
                if auth_val:
                    req.add_header('Authorization', auth_val)
                try:
                    with urllib.request.urlopen(req, timeout=20) as resp:
                        resp_data = json.loads(resp.read())
                        tests[label] = {'status': f'HTTP {resp.getcode()}', 'response': resp_data}
                except urllib.error.HTTPError as e:
                    body = e.read()
                    try:
                        resp_body = json.loads(body)
                    except:
                        resp_body = body.decode('utf-8', errors='replace')[:300]
                    tests[label] = {'code': e.code, 'reason': e.reason, 'body': resp_body}
                except Exception as e:
                    tests[label] = {'error': str(e)[:200]}
            result['auth_tests'] = tests
        else:
            result['status'] = 'no_api_key'
    except Exception as e:
        result['status'] = 'error'
        result['traceback'] = traceback.format_exc()
    return jsonify(result)



@app.route('/api/settings', methods=['GET'])
@login_required
def get_settings():
    return jsonify(load_settings())

@app.route('/api/settings', methods=['POST'])
@login_required
def update_settings():
    data = request.get_json() or {}
    settings = load_settings()
    for k in ['name', 'currency', 'bot_token', 'required_channels', 'logo',
              'admin_telegram_chat_id', 'admin_phone', 'bot_username', 'tagline',
              'payment_card', 'payment_provider', 'payment_merchant_id', 'payment_service_id',
              'group_fees', 'reminder_frequency', 'reminder_message', 'max_capacity', 'teachers',
              'auto_payment']:
        if k in data:
            val = data[k]
            if k == 'bot_token' and (not val or '…' in str(val) or val == '***'):
                continue
            settings[k] = val
    save_settings_data(settings)
    # Sync admin_telegram_chat_id to owner record in kindergartens.json
    if 'admin_telegram_chat_id' in data:
        kg_id = _current_kg_id() or 'default'
        kg = pc.get_kindergarten(kg_id)
        if kg:
            kgs = pc.load_kindergartens()
            for i, k in enumerate(kgs):
                if k['id'] == kg_id:
                    kgs[i]['owner']['telegram_chat_id'] = data['admin_telegram_chat_id']
                    pc.save_kindergartens(kgs)
                    break
    if os.environ.get('VERCEL') and settings.get('bot_token'):
        try:
            site_url = request.host_url.rstrip('/')
            kg_id = _current_kg_id()
            _register_webhook(settings['bot_token'], f"{site_url}/api/telegram/webhook/{kg_id}", secret_token=kg_id)
        except Exception:
            pass
    changed_keys = [k for k in data if k != 'bot_token']
    _audit('settings_updated', f"Sozlamalar o'zgartirildi: {', '.join(changed_keys[:10])}")
    return jsonify({'success': True, 'settings': settings})

@app.route('/api/settings/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json() or {}
    old_pw = data.get('old_password', '')
    new_pw = data.get('new_password', '')

    if len(new_pw) < 6:
        return jsonify({'success': False, 'message': 'Yangi parol kamida 6 ta belgidan iborat bo\'lishi kerak'})

    admin = get_admin()
    if not admin:
        return jsonify({'success': False, 'message': 'Admin topilmadi'})

    # Update in kindergartens.json (current auth system)
    kg_id = _current_kg_id()
    if kg_id:
        kgs = pc.load_kindergartens()
        for i, kg in enumerate(kgs):
            if kg['id'] == kg_id:
                owner = kg.get('owner', {})
                if not check_password(old_pw, owner.get('password', '')):
                    return jsonify({'success': False, 'message': 'Eski parol noto\'g\'ri'})
                kgs[i]['owner']['password'] = hash_password(new_pw)
                pc.save_kindergartens(kgs)
                _audit('password_changed', f"kg_admin: {kg_id}")
                return jsonify({'success': True})

    # Fallback to legacy admins.json
    admins = load_json('admins.json')
    for i, a in enumerate(admins):
        if a['id'] == admin['id']:
            if not check_password(old_pw, a.get('password', '')):
                return jsonify({'success': False, 'message': 'Eski parol noto\'g\'ri'})
            admins[i]['password'] = hash_password(new_pw)
            save_json('admins.json', admins)
            _audit('password_changed', f"admin: {admin['id']}")
            return jsonify({'success': True})

    return jsonify({'success': False, 'message': 'Xatolik'})

# ─── Image upload (base64 → DB, filesystem-free for Vercel) ──────────────────

@app.route('/api/upload-image', methods=['POST'])
@login_required
def upload_image():
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'message': 'Fayl topilmadi'})
        file = request.files['image']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Fayl tanlanmagan'})

        parts = file.filename.rsplit('.', 1)
        ext = parts[-1].lower() if len(parts) > 1 else ''
        if ext not in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            return jsonify({'success': False, 'message': 'Noto\'g\'ri fayl turi'})

        data = file.read()
        # Z11 — Magic bytes check
        if len(data) < 8:
            return jsonify({'success': False, 'message': 'Fayl hajmi juda kichik'})
        if len(data) > 2 * 1024 * 1024:
            return jsonify({'success': False, 'message': 'Rasm hajmi 2MB dan katta'})
        magic = data[:8]
        is_jpeg = magic[:2] == b'\xff\xd8'
        is_png = magic[:8] == b'\x89PNG\r\n\x1a\n'
        is_gif = magic[:6] in (b'GIF87a', b'GIF89a')
        is_webp = magic[:4] == b'RIFF' and data[8:12] == b'WEBP'
        if not (is_jpeg or is_png or is_gif or is_webp):
            _audit('upload_magic_mismatch', f"Magic bytes mos kelmadi: {data[:4].hex()}")
            return jsonify({'success': False, 'message': 'Haqiqiy rasm fayli emas'})

        # Re-encode to strip metadata & malware; resize if needed
        safe_data = data
        mime = f'image/{ext}'
        try:
            from PIL import Image
            import io
            img = Image.open(io.BytesIO(data))
            if img.width > 800 or img.height > 800:
                img.thumbnail((800, 800), Image.LANCZOS)
            img = img.convert('RGB') if img.mode in ('RGBA', 'P') else img
            out = io.BytesIO()
            fmt = 'JPEG' if is_jpeg else ('PNG' if is_png else ('GIF' if is_gif else 'WEBP'))
            img.save(out, format=fmt, quality=80)
            safe_data = out.getvalue()
            mime = f'image/{fmt.lower()}'
        except Exception:
            pass

        # Encode as base64 data URL — stored in DB, no filesystem needed
        import base64
        b64 = base64.b64encode(safe_data).decode('ascii')
        data_url = f'data:{mime};base64,{b64}'
        return jsonify({'success': True, 'url': data_url})
    except Exception as ex:
        return jsonify({'success': False, 'message': f'Fayl yuklashda xatolik: {str(ex)}'})

# ─── Parent portfolios (admin yaratadi) ─────────────────────────────────────────

def sync_students_from_portfolio(portfolio):
    """Portfelga bog'langan o'quvchilarga ota-ona ma'lumotini yozish."""
    students = load_json('students.json')
    phone = normalize_phone(portfolio.get('phone', ''))
    pname = portfolio.get('parent_name', '').strip()
    sids = set(portfolio.get('student_ids', []))
    changed = False
    for i, s in enumerate(students):
        if s['id'] in sids:
            students[i]['parent_name'] = pname
            students[i]['parent_phone'] = phone
            changed = True
    if changed:
        save_json('students.json', students)

@app.route('/api/parent-portfolios', methods=['GET'])
@login_required
def list_parent_portfolios():
    portfolios = load_json('parent_portfolios.json')
    students = load_json('students.json')
    sid_map = {s['id']: s for s in students}
    result = []
    for p in portfolios:
        children = []
        for sid in p.get('student_ids', []):
            s = sid_map.get(sid)
            if s:
                children.append({
                    'id': s['id'],
                    'name': f"{s['first_name']} {s['last_name']}",
                    'group': s.get('group', ''),
                    'status': s.get('status', 'active')
                })
        result.append({**p, 'children': children})
    return jsonify(sorted(result, key=lambda x: x.get('created_at', ''), reverse=True))

@app.route('/api/parent-portfolios', methods=['POST'])
@login_required
def create_parent_portfolio():
    data = request.get_json() or {}
    phone = normalize_phone(data.get('phone', ''))
    if not phone or len(phone) < 9:
        return jsonify({'success': False, 'message': 'Telefon raqam noto\'g\'ri'})
    pname = (data.get('parent_name') or '').strip()
    if not pname:
        return jsonify({'success': False, 'message': 'Ota-ona ismini kiriting'})

    portfolios = load_json('parent_portfolios.json')
    portfolio = {
        'id': 'PAR-' + str(uuid.uuid4())[:8].upper(),
        'parent_name': pname,
        'phone': phone,
        'email': (data.get('email') or '').strip(),
        'notes': (data.get('notes') or '').strip(),
        'student_ids': data.get('student_ids', []) or [],
        'active': True,
        'created_at': datetime.now().isoformat(),
        'created_by': session.get('admin_name', 'Admin')
    }
    portfolios.append(portfolio)
    save_json('parent_portfolios.json', portfolios)
    sync_students_from_portfolio(portfolio)
    add_notification(f"Ota-ona portfeli yaratildi: {pname} ({phone})", 'info')
    return jsonify({'success': True, 'portfolio': portfolio})

@app.route('/api/parent-portfolios/<pid>', methods=['PUT'])
@login_required
def update_parent_portfolio(pid):
    data = request.get_json() or {}
    portfolios = load_json('parent_portfolios.json')
    for i, p in enumerate(portfolios):
        if p['id'] == pid:
            phone = normalize_phone(data.get('phone', p['phone']))
            portfolios[i].update({
                'parent_name': (data.get('parent_name', p['parent_name']) or '').strip(),
                'phone': phone,
                'email': (data.get('email', p.get('email', '')) or '').strip(),
                'notes': (data.get('notes', p.get('notes', '')) or '').strip(),
                'student_ids': data.get('student_ids', p.get('student_ids', [])),
                'active': data.get('active', p.get('active', True))
            })
            save_json('parent_portfolios.json', portfolios)
            sync_students_from_portfolio(portfolios[i])
            return jsonify({'success': True, 'portfolio': portfolios[i]})
    return jsonify({'success': False, 'message': 'Portfel topilmadi'}), 404

@app.route('/api/parent-portfolios/<pid>', methods=['DELETE'])
@login_required
def delete_parent_portfolio(pid):
    portfolios = load_json('parent_portfolios.json')
    portfolios = [p for p in portfolios if p['id'] != pid]
    save_json('parent_portfolios.json', portfolios)
    return jsonify({'success': True})

# ─── Yangi bog'cha arizasi ─────────────────────────────────────────────────────

@app.route('/bogcha-register')
def bogcha_register_page():
    return render_template('bogcha_register.html')

@app.route('/api/bogcha-register', methods=['POST'])
def submit_bogcha_application():
    data = request.get_json() or {}
    name = (data.get('kindergarten_name') or '').strip()
    director = (data.get('director_name') or '').strip()
    phone = normalize_phone(data.get('phone', ''))
    plan = (data.get('plan') or 'standard').strip()
    if not name or not director or len(phone) < 9:
        return jsonify({'success': False, 'message': 'Bog\'cha nomi, rahbar va telefon majburiy'})

    plat = pc.load_platform()
    plans = {p['id']: p for p in plat.get('plans', [])}
    plan_info = plans.get(plan, plans.get('standard', {'name': 'Standart', 'price_usd': 10}))

    temp_pw = data.get('temp_password', '12345678')
    if len(temp_pw) < 6:
        temp_pw = '12345678'

    apps = load_json('kindergarten_applications.json')
    app_rec = {
        'id': 'KG-' + str(uuid.uuid4())[:8].upper(),
        'kindergarten_name': name,
        'director_name': director,
        'phone': phone,
        'plan': plan,
        'plan_name': plan_info.get('name', 'Standart'),
        'price_usd': plan_info.get('price_usd', 10),
        'city': (data.get('city') or '').strip(),
        'address': (data.get('address') or '').strip(),
        'email': (data.get('email') or '').strip(),
        'capacity': (data.get('capacity') or '').strip(),
        'owner_telegram': (data.get('owner_telegram') or '').strip(),
        'temp_password': temp_pw,
        'notes': (data.get('notes') or '').strip(),
        'status': 'pending',
        'created_at': datetime.now().isoformat()
    }
    apps.append(app_rec)
    save_json('kindergarten_applications.json', apps)
    notify_super_admin(
        f"🏫 <b>Yangi bog'cha arizasi</b>\n"
        f"Nom: {name}\nRahbar: {director}\nTel: {phone}\n"
        f"📦 Tarif: <b>{plan_info.get('name')}</b> — 1 oy bepul, keyin <b>${plan_info.get('price_usd')}/oy</b>\n"
        f"Premium alohida ko'rib chiqiladi."
    )
    return jsonify({'success': True, 'message': 'Arizangiz qabul qilindi! Raufbek Turaqulov tez orada bog\'lanadi.', 'id': app_rec['id']})

@app.route('/api/platform/plans')
def public_plans():
    return jsonify(pc.load_platform())

@app.route('/api/kindergarten-applications', methods=['GET'])
@super_required
def list_kindergarten_applications():
    apps = load_json('kindergarten_applications.json')
    pending = [a for a in apps if a.get('status') == 'pending']
    return jsonify(sorted(pending, key=lambda x: x.get('created_at', ''), reverse=True))

# ─── Parents section ──────────────────────────────────────────────────────────

@app.route('/parents')
@login_required
def parents_page():
    settings = load_settings()
    return render_template('parents.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/parents')
@login_required
def get_parents():
    """Eski API — portfellar bilan bir xil ma'lumot."""
    portfolios = load_json('parent_portfolios.json')
    if portfolios:
        students = load_json('students.json')
        sid_map = {s['id']: s for s in students}
        result = []
        for p in portfolios:
            children = []
            for sid in p.get('student_ids', []):
                s = sid_map.get(sid)
                if s:
                    children.append({
                        'id': s['id'],
                        'name': f"{s['first_name']} {s['last_name']}",
                        'group': s.get('group', ''),
                        'status': s.get('status', 'active')
                    })
            result.append({
                'id': p['id'],
                'name': p.get('parent_name', ''),
                'phone': p.get('phone', ''),
                'children': children
            })
    courses_map = {c['id']: c['name'] for c in load_json('courses.json')}
    for s in result:
        s['course_name'] = courses_map.get(s.get('course_id', ''), '')
    return jsonify(result)

    students = load_json('students.json')
    parents = {}
    for s in students:
        phone = s.get('parent_phone', '')
        if phone not in parents:
            parents[phone] = {
                'name': s.get('parent_name', ''),
                'phone': phone,
                'children': []
            }
        parents[phone]['children'].append({
            'id': s['id'],
            'name': f"{s['first_name']} {s['last_name']}",
            'group': s.get('group', ''),
            'status': s.get('status', 'active')
        })
    return jsonify(list(parents.values()))

# ─── Report Module ─────────────────────────────────────────────────────────────
def _report_data(year, month):
    students = load_json('students.json')
    payments = load_json('payments.json')
    attendance = load_json('attendance.json')
    portfolios = load_json('parent_portfolios.json')
    checks = load_json('payment_checks.json', None) or []

    ym = f'{year}-{month:02d}'
    ym_prefix = lambda d: d.startswith(ym) if d else False

    active = [s for s in students if s.get('status') == 'active']
    working_days = get_working_days(year, month)

    # Group breakdown
    groups = {}
    for s in students:
        g = s.get('group', '') or "Guruhsiz"
        groups.setdefault(g, {'total': 0, 'active': 0, 'paid': 0, 'debt': 0, 'fee_sum': 0, 'count': 0})
        groups[g]['total'] += 1
        groups[g]['count'] += 1
        if s.get('status') == 'active':
            groups[g]['active'] += 1
            groups[g]['fee_sum'] += s.get('monthly_fee', 0)
            payable = calc_payable_amount(s, year, month)
            paid = get_paid_amount(s['id'], year, month)
            groups[g]['paid'] += paid
            if paid < payable:
                groups[g]['debt'] += payable - paid

    # Per-student details
    details = []
    debtors_list = []
    for s in active:
        payable = calc_payable_amount(s, year, month)
        paid = get_paid_amount(s['id'], year, month)
        attended = get_attended_days(s['id'], year, month)
        absent = get_absent_days(s['id'], year, month)
        debt = max(payable - paid, 0)
        row = {
            'id': s['id'],
            'name': f"{s['first_name']} {s['last_name']}",
            'group': s.get('group', ''),
            'fee': s.get('monthly_fee', 0),
            'payable': payable,
            'paid': paid,
            'debt': debt,
            'attended': attended,
            'absent': absent,
            'working': working_days,
            'phone': s.get('parent_phone', ''),
            'parent': s.get('parent_name', '')
        }
        details.append(row)
        if debt > 0:
            debtors_list.append(row)

    # Payments this month
    month_payments = [p for p in payments if p.get('month', '') == ym and p.get('status', 'paid') != 'cancelled']
    total_collected = sum(p['amount'] for p in month_payments)

    # Pending checks
    pending_checks = [c for c in checks if c.get('status') == 'pending']
    pending_amount = sum(c.get('amount', 0) for c in pending_checks)

    # Attendance summary
    month_att = [a for a in attendance if ym_prefix(a.get('date', ''))]
    total_present = sum(1 for a in month_att if a['status'] == 'present')
    total_absent = sum(1 for a in month_att if a['status'] == 'absent')
    total_excused = sum(1 for a in month_att if a['status'] == 'excused')
    expected = working_days * len(active)
    att_pct = round(total_present / expected * 100, 1) if expected else 0

    # New students this month
    new_students = [s for s in students if ym_prefix(s.get('join_date', ''))]

    # Total debts
    total_debt = sum(d['debt'] for d in details)
    debtor_count = len(debtors_list)

    group_list = []
    for gname, gdata in sorted(groups.items(), key=lambda x: -x[1]['active']):
        group_list.append({
            'name': gname,
            'total': gdata['total'],
            'active': gdata['active'],
            'collected': gdata['paid'],
            'debt': gdata['debt'],
            'avg_fee': round(gdata['fee_sum'] / gdata['active'], 0) if gdata['active'] else 0
        })

    return {
        'year': year,
        'month': month,
        'ym': ym,
        'total_students': len(students),
        'active_students': len(active),
        'new_students': len(new_students),
        'working_days': working_days,
        'total_present': total_present,
        'total_absent': total_absent,
        'total_excused': total_excused,
        'attendance_pct': att_pct,
        'total_collected': total_collected,
        'pending_checks': len(pending_checks),
        'pending_amount': pending_amount,
        'total_debt': total_debt,
        'debtor_count': debtor_count,
        'groups': group_list,
        'details': details[:500],
        'debtors': debtors_list[:50],
        'currency': load_settings().get('currency', 'UZS')
    }

@app.route('/report')
@login_required
def report_page():
    settings = load_settings()
    return render_template('report.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/report/data')
@login_required
def report_data():
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    return jsonify(_report_data(year, month))

@app.route('/api/report/excel')
@login_required
def report_excel():
    # Rate-limit: max 1 export per 30 seconds
    last = session.get('last_export_time', 0)
    if time.time() - last < 30:
        return jsonify({'success': False, 'message': 'Iltimos, 30 soniya kuting'}), 429
    session['last_export_time'] = time.time()
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    data = _report_data(year, month)
    global HAS_OPENPYXL
    if HAS_OPENPYXL is None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            HAS_OPENPYXL = True
        except ImportError:
            HAS_OPENPYXL = False
    if not HAS_OPENPYXL:
        return jsonify({'success': False, 'message': 'openpyxl o\'rnatilmagan'}), 500

    wb = Workbook()
    hdr = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='center')

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = 'Xulosa'
    ws1.merge_cells('A1:B1')
    ws1.cell(row=1, column=1, value=f"Hisobot — {data['ym']}").font = Font(bold=True, size=14, color='4F46E5')
    summary = [
        ['Jami o\'quvchilar', data['total_students']],
        ['Faol o\'quvchilar', data['active_students']],
        ['Yangi o\'quvchilar', data['new_students']],
        ['Ish kunlari', data['working_days']],
        ['Kelganlar (jami)', data['total_present']],
        ['Kelmaganlar (jami)', data['total_absent']],
        ['Davomat foizi', f"{data['attendance_pct']}%"],
        ['Yig\'ilgan to\'lov', data['total_collected']],
        ['Kutilayotgan cheklar', data['pending_amount']],
        ['Jami qarzdorlik', data['total_debt']],
        ['Qarzdorlar soni', data['debtor_count']],
    ]
    for i, (k, v) in enumerate(summary, 3):
        ws1.cell(row=i, column=1, value=k).border = border
        ws1.cell(row=i, column=2, value=v).border = border
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 16

    # Sheet 2: Groups
    ws2 = wb.create_sheet('Guruhlar')
    ws2.append(['Guruh', 'Jami', 'Faol', 'Yig\'ilgan', 'Qarzdorlik', "O'rtacha to'lov"])
    style_header(ws2, 1, 6)
    for g in data['groups']:
        ws2.append([g['name'], g['total'], g['active'], g['collected'], g['debt'], g['avg_fee']])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=6):
        for cell in row:
            cell.border = border
    ws2.column_dimensions['A'].width = 18
    for c in 'BCDEF': ws2.column_dimensions[c].width = 14

    # Sheet 3: Students
    ws3 = wb.create_sheet('O\'quvchilar')
    ws3.append(['ID', 'Ism', 'Guruh', 'Oylik to\'lov', 'To\'lanadigan', 'To\'langan', 'Qarzdorlik', 'Kelgan', 'Kelmagan', 'Ota-ona', 'Telefon'])
    style_header(ws3, 1, 11)
    for d in data['details']:
        ws3.append([d['id'], d['name'], d['group'], d['fee'], d['payable'], d['paid'], d['debt'], d['attended'], d['absent'], d['parent'], d['phone']])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=11):
        for cell in row:
            cell.border = border
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 24
    ws3.column_dimensions['C'].width = 12
    for c in 'DEFGHI': ws3.column_dimensions[c].width = 14
    ws3.column_dimensions['J'].width = 20
    ws3.column_dimensions['K'].width = 16

    # Sheet 4: Debtors
    ws4 = wb.create_sheet('Qarzdorlar')
    ws4.append(['ID', 'Ism', 'Guruh', 'Qarzdorlik', 'Ota-ona', 'Telefon'])
    style_header(ws4, 1, 6)
    for d in data['debtors']:
        ws4.append([d['id'], d['name'], d['group'], d['debt'], d['parent'], d['phone']])
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, max_col=6):
        for cell in row:
            cell.border = border
    ws4.column_dimensions['A'].width = 18
    ws4.column_dimensions['B'].width = 24
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 14
    ws4.column_dimensions['E'].width = 20
    ws4.column_dimensions['F'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"hisobot_{data['ym']}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/report/pdf')
@login_required
def report_pdf():
    # Rate-limit: max 1 export per 30 seconds
    last = session.get('last_export_time', 0)
    if time.time() - last < 30:
        return jsonify({'success': False, 'message': 'Iltimos, 30 soniya kuting'}), 429
    session['last_export_time'] = time.time()
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    data = _report_data(year, month)
    global HAS_REPORTLAB
    if HAS_REPORTLAB is None:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm, cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            HAS_REPORTLAB = True
        except ImportError:
            HAS_REPORTLAB = False
    if not HAS_REPORTLAB:
        return jsonify({'success': False, 'message': 'reportlab o\'rnatilmagan'}), 500

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6*mm, alignment=TA_CENTER)
    style_h = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#1E293B'), spaceAfter=3*mm, spaceBefore=5*mm)
    style_n = styles['Normal']
    style_c = ParagraphStyle('Center', parent=style_n, alignment=TA_CENTER, fontSize=9)
    style_r = ParagraphStyle('Right', parent=style_n, alignment=TA_RIGHT, fontSize=9)

    elements = []
    elements.append(Paragraph(f"Hisobot — {data['ym']}", style_title))
    elements.append(Spacer(1, 3*mm))

    # Summary table
    summary_data = [
        ['Jami', str(data['total_students']), 'Faol', str(data['active_students']), 'Yangi', str(data['new_students'])],
        ['Ish kuni', str(data['working_days']), 'Kelgan', str(data['total_present']), 'Kelmagan', str(data['total_absent'])],
        ['Davomat', f"{data['attendance_pct']}%", 'Yig\'ilgan', f"{data['total_collected']:,}", 'Qarzdorlik', f"{data['total_debt']:,}"],
    ]
    t = Table(summary_data, colWidths=[45, 50, 40, 55, 40, 55])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 5*mm))

    # Groups table
    elements.append(Paragraph('Guruhlar bo\'yicha', style_h))
    gdata = [['Guruh', 'Jami', 'Faol', 'Yig\'ilgan', 'Qarzdorlik']]
    for g in data['groups']:
        gdata.append([g['name'], str(g['total']), str(g['active']), f"{g['collected']:,}", f"{g['debt']:,}"])
    t2 = Table(gdata, colWidths=[80, 50, 50, 70, 70])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 5*mm))

    # Debtors
    if data['debtors']:
        elements.append(Paragraph('Qarzdor o\'quvchilar', style_h))
        ddata = [['Ism', 'Guruh', 'Qarzdorlik', 'Telefon']]
        for d in data['debtors'][:20]:
            ddata.append([d['name'], d['group'], f"{d['debt']:,}", d['phone']])
        t3 = Table(ddata, colWidths=[100, 60, 70, 70])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(t3)

    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph(f"Hisobot avtomatik tarzda yaratildi — {datetime.now().strftime('%d.%m.%Y %H:%M')}", ParagraphStyle('Footer', parent=style_n, fontSize=7, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(elements)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"hisobot_{data['ym']}.pdf", mimetype='application/pdf')

init_data()

def _boot_telegram():
    try:
        from telegram_bot import start_polling_background, super_bot_polling
        start_polling_background(telegram_app_context())
        # Start super admin bot
        plat = pc.load_platform()
        super_token = plat.get('super_bot_token', '').strip()
        super_chat = plat.get('super_telegram_chat_id', '').strip()
        if super_token and super_chat:
            import threading
            def _start_super():
                try:
                    super_bot_polling(super_token, super_chat, telegram_app_context())
                except Exception as ex:
                    print(f"[SuperBot] ishga tushmadi: {ex}")
            t = threading.Thread(target=_start_super, daemon=True, name='super-bot-poll')
            t.start()
            print(f"[SuperBot] polling boshlandi (chat: {super_chat})")
        else:
            print(f"[SuperBot] token={bool(super_token)} chat={bool(super_chat)}")
    except Exception as ex:
        print(f"[Telegram] polling ishga tushmadi: {ex}")

# ─── Vercel webhook setup ──────────────────────────────────────────────────
def _register_webhook(bot_token, webhook_url, secret_token=None):
    try:
        url = f"https://api.telegram.org/bot{bot_token}/setWebhook"
        payload = {
            'url': webhook_url,
            'max_connections': 5,
        }
        if secret_token:
            payload['secret_token'] = secret_token
        resp = http_requests.post(url, json=payload, timeout=10)
        data = resp.json()
        ok = data.get('ok', False)
        print(f"[Webhook] {webhook_url} ok={ok} {data.get('description', '')}")
        return ok
    except Exception as ex:
        print(f"[Webhook] xato: {ex}")
        return False

# ─── Performance Indicators ────────────────────────────────────────────────────

@app.route('/performance')
@login_required
def performance_page():
    settings = load_settings()
    return render_template('performance.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/performance/stats')
@login_required
def performance_stats():
    students = load_json('students.json')
    attendance = load_json('attendance.json')
    settings = load_settings()

    active_students = [s for s in students if s.get('status') == 'active']
    total_active = len(active_students)
    max_cap = int(settings.get('max_capacity', 0)) or 100
    capacity_pct = round(total_active / max_cap * 100, 1)

    groups = {}
    for s in active_students:
        g = s.get('group', 'Guruhsiz')
        groups.setdefault(g, {'students': 0, 'present': 0, 'total': 0})
        groups[g]['students'] += 1

    teacher_count = int(settings.get('teachers', 0)) or max(1, total_active // 15)
    student_teacher_ratio = round(total_active / teacher_count, 1) if teacher_count else 0

    now = datetime.now()
    monthly = {}
    for a in attendance:
        sid = a.get('student_id')
        s = next((x for x in students if x['id'] == sid and x.get('status') == 'active'), None)
        if not s:
            continue
        try:
            ym = a['date'][:7]
        except Exception:
            continue
        monthly.setdefault(ym, {'present': 0, 'absent': 0, 'excused': 0, 'total': 0})
        st = a.get('status', 'absent')
        monthly[ym]['total'] += 1
        if st == 'present':
            monthly[ym]['present'] += 1
        elif st == 'excused':
            monthly[ym]['excused'] += 1
        else:
            monthly[ym]['absent'] += 1

        g = s.get('group', 'Guruhsiz')
        groups[g]['total'] += 1
        if st == 'present':
            groups[g]['present'] += 1

    months_list = []
    for i in range(5, -1, -1):
        ym = (now.replace(day=1) - timedelta(days=30 * i)).strftime('%Y-%m')
        d = monthly.get(ym, {'present': 0, 'absent': 0, 'excused': 0, 'total': 0})
        rate = round(d['present'] / d['total'] * 100, 1) if d['total'] else 0
        months_list.append({'month': ym, 'present': d['present'], 'absent': d['absent'], 'excused': d['excused'], 'total': d['total'], 'rate': rate})

    group_stats = []
    for g_name, g_data in groups.items():
        rate = round(g_data['present'] / g_data['total'] * 100, 1) if g_data['total'] else 0
        group_stats.append({'group': g_name, 'students': g_data['students'], 'present': g_data['present'], 'total': g_data['total'], 'rate': rate})
    group_stats.sort(key=lambda x: x['rate'])

    overall_att = sum(m['total'] for m in months_list)
    overall_present = sum(m['present'] for m in months_list)
    overall_rate = round(overall_present / overall_att * 100, 1) if overall_att else 0

    return jsonify({
        'total_active': total_active,
        'max_capacity': max_cap,
        'capacity_pct': capacity_pct,
        'teacher_count': teacher_count,
        'student_teacher_ratio': student_teacher_ratio,
        'overall_attendance_rate': overall_rate,
        'monthly_trend': months_list,
        'group_stats': group_stats
    })

# ─── Events Calendar ───────────────────────────────────────────────────────────

@app.route('/calendar')
@login_required
def calendar_page():
    settings = load_settings()
    return render_template('calendar.html', settings=settings, admin_name=session.get('admin_name'))

@app.route('/api/events/public')
def public_events():
    events = load_json('events.json')
    today = date.today().isoformat()
    events = [e for e in events if e.get('date', '') >= today]
    events = sorted(events, key=lambda e: e['date'])[:10]
    return jsonify(events)

@app.route('/api/events', methods=['GET'])
@login_required
def get_events():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    events = load_json('events.json')
    if year and month:
        prefix = f'{year}-{month:02d}'
        events = [e for e in events if e.get('date', '').startswith(prefix)]
    return jsonify(sorted(events, key=lambda e: e.get('date', '')))

@app.route('/api/events', methods=['POST'])
@login_required
def create_event():
    data = request.get_json() or {}
    eid = str(uuid.uuid4())[:8]
    event = {
        'id': eid,
        'title': data.get('title', '').strip(),
        'date': data.get('date', ''),
        'type': data.get('type', 'other'),
        'description': data.get('description', '').strip(),
        'notify': bool(data.get('notify', False)),
        'notified': False,
        'created_at': datetime.now().isoformat()
    }
    if not event['title'] or not event['date']:
        return jsonify({'success': False, 'message': 'Sarlavha va sana majburiy'}), 400
    events = load_json('events.json')
    events.append(event)
    save_json('events.json', events)
    if event['notify']:
        try:
            _notify_event(event)
        except Exception as ex:
            print(f'Event notify error: {ex}')
    return jsonify({'success': True, 'event': event})

@app.route('/api/events/<event_id>', methods=['PUT'])
@login_required
def update_event(event_id):
    data = request.get_json() or {}
    events = load_json('events.json')
    for e in events:
        if e['id'] == event_id:
            e['title'] = data.get('title', e['title']).strip()
            e['date'] = data.get('date', e['date'])
            e['type'] = data.get('type', e['type'])
            e['description'] = data.get('description', e['description']).strip()
            was_notify = e.get('notify', False)
            e['notify'] = bool(data.get('notify', e['notify']))
            if e['notify'] and not was_notify:
                try:
                    _notify_event(e)
                except Exception as ex:
                    print(f'Event notify error: {ex}')
            save_json('events.json', events)
            return jsonify({'success': True, 'event': e})
    return jsonify({'success': False, 'message': 'Topilmadi'}), 404

@app.route('/api/events/<event_id>', methods=['DELETE'])
@login_required
def delete_event(event_id):
    events = load_json('events.json')
    events = [e for e in events if e['id'] != event_id]
    save_json('events.json', events)
    return jsonify({'success': True})

def _notify_event(event):
    settings = load_settings()
    token = settings.get('bot_token')
    if not token:
        return
    type_icons = {'holiday': '🎉', 'meeting': '📋', 'excursion': '🚌', 'other': '📌'}
    icon = type_icons.get(event['type'], '📌')
    title = event.get('title', '')
    desc = event.get('description', '')
    text = f'{icon} <b>{title}</b>\n📅 {event.get("date", "")}'
    if desc:
        text += f'\n\n{desc}'
    text += '\n\n#event'
    students = load_json('students.json')
    sent = set()
    for s in students:
        parent_phone = s.get('parent_phone', '').strip()
        chat_id = s.get('telegram_chat_id', '').strip()
        if parent_phone and chat_id and parent_phone not in sent:
            try:
                http_requests.post(f'https://api.telegram.org/bot{token}/sendMessage', json={
                    'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'
                }, timeout=8)
                sent.add(parent_phone)
            except Exception:
                pass
    admin_chat = settings.get('admin_telegram_chat_id', '').strip()
    if admin_chat:
        try:
            http_requests.post(f'https://api.telegram.org/bot{token}/sendMessage', json={
                'chat_id': admin_chat, 'text': text, 'parse_mode': 'HTML'
            }, timeout=8)
        except Exception:
            pass

# ── Google Site Verification ───────────────────────────────────────────────
@app.route('/google8e1127d0a057f256.html')
def google_verification():
    return '<html><head><meta name="google-site-verification" content="google8e1127d0a057f256"></head><body></body></html>', 200, {'Content-Type': 'text/html'}


# ── Privacy Policy ──────────────────────────────────────────────────────────

@app.route('/privacy')
def privacy_page():
    lang = session.get('lang', 'uz')
    theme = session.get('theme', 'dark')
    return render_template('privacy.html', current_lang=lang, current_theme=theme)


# ── Right to be Forgotten (parent data deletion request) ────────────────────

@app.route('/api/privacy/deletion-request', methods=['POST'])
def privacy_deletion_request():
    from privacy import create_deletion_request
    data = request.get_json() or {}
    phone = data.get('phone', '').strip()
    parent_name = data.get('parent_name', '').strip()
    reason = data.get('reason', '').strip()
    kg_id = data.get('kg_id', session.get('kindergarten_id', '')).strip()
    if not phone or not parent_name:
        return jsonify({'success': False, 'message': 'Telefon va ism majburiy'}), 400
    req = create_deletion_request(pc, kg_id, phone, parent_name, reason)
    return jsonify({'success': True, 'message': 'So\'rovingiz qabul qilindi. Admin tekshirib chiqadi.', 'id': req['id']})


# ── Admin: view & approve deletion requests ─────────────────────────────────

@app.route('/api/privacy/deletion-requests', methods=['GET'])
@login_required
def privacy_list_requests():
    reqs = pc.load_json('deletion_requests.json')
    if not isinstance(reqs, list):
        reqs = []
    kg_id = session.get('kindergarten_id')
    if kg_id:
        reqs = [r for r in reqs if r.get('kg_id') == kg_id]
    return jsonify(sorted(reqs, key=lambda r: r.get('created_at', ''), reverse=True))


@app.route('/api/privacy/deletion-requests/<req_id>/approve', methods=['POST'])
@login_required
def privacy_approve_request(req_id):
    from privacy import execute_deletion_request
    ok = execute_deletion_request(pc, req_id)
    if ok:
        return jsonify({'success': True, 'message': 'Ma\'lumotlar o\'chirildi'})
    return jsonify({'success': False, 'message': 'So\'rov topilmadi'}), 404


# ── Data Retention: purge expired records manually (admin) ──────────────────

@app.route('/api/privacy/purge-expired', methods=['POST'])
@login_required
def privacy_purge_expired():
    from privacy import purge_expired_deleted_data
    count = purge_expired_deleted_data(pc)
    return jsonify({'success': True, 'message': f'{count} ta eskirgan yozuv tozalandi', 'purged': count})


def _setup_vercel_webhooks():
    site_url = os.environ.get('SITE_URL', '').strip()
    if site_url:
        base = site_url.rstrip('/')
    else:
        vercel_url = os.environ.get('VERCEL_URL', '').strip()
        if not vercel_url:
            print("[Webhook] VERCEL_URL topilmadi, webhook o'rnatilmadi")
            return
        production_url = os.environ.get('VERCEL_PROJECT_PRODUCTION_URL', '').strip()
        if production_url and '-' in vercel_url.split('.')[0].split('//')[-1]:
            vercel_url = production_url
        base = f"https://{vercel_url}"
    print(f"[Webhook] Base URL: {base}")
    # Kg bot webhook — find all kindergartens with bot tokens
    kgs = pc.load_kindergartens()
    if isinstance(kgs, list):
        for kg in kgs:
            kg_id = kg.get('id')
            if not kg_id:
                continue
            try:
                s = load_settings(kg_id)
                token = (s or {}).get('bot_token', '')
                if token:
                    _register_webhook(token, f"{base}/api/telegram/webhook/{kg_id}", secret_token=kg_id)
            except Exception as ex:
                print(f"[Webhook] kg {kg_id} xato: {ex}")
    else:
        print("[Webhook] kindergartens.json not a list — skipping")
    # Super bot webhook
    try:
        plat = pc.load_platform()
        super_token = plat.get('super_bot_token', '').strip()
        if super_token:
            wh_secret = os.environ.get('SUPER_WEBHOOK_SECRET', '').strip()
            _register_webhook(super_token, f"{base}/api/telegram/super-webhook", secret_token=wh_secret or None)
        # Clean up old / orphaned super bot webhooks
        old_tokens = plat.get('previous_super_bot_tokens', [])
        if isinstance(old_tokens, list):
            for old_tok in old_tokens:
                t = old_tok.strip()
                if t and t != super_token:
                    try:
                        http_requests.post(f"https://api.telegram.org/bot{t}/deleteWebhook",
                            json={'drop_pending_updates': True}, timeout=5)
                        print(f"[Webhook] eski bot webhooki o'chirildi")
                    except Exception:
                        pass
    except Exception as ex:
        print(f"[Webhook] super bot xato: {ex}")

if os.environ.get('VERCEL'):
    # ── Sync platform.json from filesystem to database ──
    # (Vercel PostgreSQL may have stale data; filesystem has latest)
    fs_platform_path = os.path.join(pc.data_dir, 'platform.json')
    try:
        if os.path.exists(fs_platform_path):
            with open(fs_platform_path, 'r', encoding='utf-8') as f:
                fs_plat = json.load(f)
            existing = pc.load_platform()
            changed = False
            for k, v in fs_plat.items():
                if existing.get(k) != v:
                    existing[k] = v
                    changed = True
            if changed:
                pc.save_json('platform.json', existing)
    except Exception:
        pass
    # Force-delete old/orphaned webhooks
    for old_tok in ['8993914821:AAG16D9CVegKd5kTAy_GzGNeLnuoHhP2zjM']:
        try:
            http_requests.post(f"https://api.telegram.org/bot{old_tok}/deleteWebhook",
                json={'drop_pending_updates': True}, timeout=5)
        except Exception:
            pass
    _setup_vercel_webhooks()
else:
    _boot_telegram()

# ─── Cron endpoints (external cron-job.org) ──────────────────────────────
@app.route('/api/cron/weekly-report')
def cron_weekly_report():
    secret = request.args.get('secret', '')
    expected = os.environ.get('CRON_SECRET', '').strip()
    if not expected:
        return jsonify({'ok': False, 'message': 'CRON_SECRET not set'}), 503
    if secret != expected:
        return jsonify({'ok': False, 'message': 'bad secret'}), 403
    plat = pc.load_platform()
    st = plat.get('super_bot_token', '').strip()
    sc = plat.get('super_telegram_chat_id', '').strip()
    if not st or not sc:
        return jsonify({'ok': False, 'message': 'Telegram not configured'})
    try:
        report = audit_log.weekly_report(pc)
        from telegram_bot import send_message as tg_send
        tg_send(st, sc, report)
        return jsonify({'ok': True, 'chars': len(report)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/cron/hourly-report')
def cron_hourly_report():
    secret = request.args.get('secret', '')
    expected = os.environ.get('CRON_SECRET', '').strip()
    if not expected:
        return jsonify({'ok': False, 'message': 'CRON_SECRET not set'}), 503
    if secret != expected:
        return jsonify({'ok': False, 'message': 'bad secret'}), 403
    plat = pc.load_platform()
    st = plat.get('super_bot_token', '').strip()
    sc = plat.get('super_telegram_chat_id', '').strip()
    if not st or not sc:
        return jsonify({'ok': False, 'message': 'Telegram not configured'})
    try:
        from telegram_bot import send_message as tg_send
        now = datetime.now(timezone.utc)
        one_hour_ago = (now - timedelta(hours=1)).isoformat() + 'Z'
        logs = audit_log.get_logs(pc, limit=5000)
        hour_logs = [l for l in logs if l.get('timestamp', '') >= one_hour_ago]
        summary = {}
        ips = set()
        for l in hour_logs:
            act = l.get('action', 'unknown')
            summary[act] = summary.get(act, 0) + 1
            if l.get('ip'):
                ips.add(l.get('ip'))
        # Check visits data for this hour
        today = now.strftime('%Y-%m-%d')
        hour_str = now.strftime('%H')
        visits_data = pc.load_json(VISIT_LOG_FILE)
        hour_views = 0
        hour_unique = 0
        if isinstance(visits_data, dict) and today in visits_data:
            hr = visits_data[today].get('hours', {}).get(hour_str, {})
            hour_views = hr.get('views', 0)
            hour_unique = len(hr.get('ips', {}))
        # Find suspicious
        susp = [l for l in hour_logs if l.get('action') in
                ('payment_deleted', 'attendance_bulk_edit', 'super_admin_blocked_ip')]
        lines = [f"⏰ <b>Soatlik hisobot</b>",
                 f"🗓 {now.strftime('%Y-%m-%d %H:%M')} UTC",
                 f"📊 <a href='{BASE_URL}/super/stats'>To'liq statistika</a>",
                 ""]
        lines.append(f"👁 Ko'rishlar: {hour_views} (unikal: {hour_unique})")
        lines.append(f"👤 Kirishlar: {summary.get('login', 0)} ta")
        lines.append(f"❌ Xato kirish: {summary.get('failed_login', 0)} ta")
        lines.append(f"📊 Jami hodisa: {len(hour_logs)} ta")
        lines.append(f"🌐 Unikal IP: {len(ips)} ta")
        lines.append("")
        if summary:
            lines.append("📋 <b>Amallar:</b>")
            for act, cnt in sorted(summary.items(), key=lambda x: -x[1]):
                lines.append(f"  • {act}: {cnt}")
            lines.append("")
        if susp:
            lines.append("🚨 <b>Shubhali harakatlar:</b>")
            for l in susp:
                lines.append(f"  ⚠️ {l.get('action')} — {l.get('admin_name', '?')} | {l.get('ip', '?')} | {l.get('timestamp', '?')[:19]}")
        else:
            lines.append("✅ Shubhali harakat yo'q")
        lines.append("")
        lines.append(f"📊 <a href='{BASE_URL}/super/stats'>Statistikani ochish</a>")
        lines.append(f"👁 <a href='{BASE_URL}/super/daily-stats'>Kunlik tahlil</a>")
        report = '\n'.join(lines)
        tg_send(st, sc, report)
        return jsonify({'ok': True, 'chars': len(report), 'events': len(hour_logs)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/cron/payment-reminders')
def cron_payment_reminders():
    secret = request.args.get('secret', '')
    expected = os.environ.get('CRON_SECRET', '').strip()
    if not expected:
        return jsonify({'ok': False, 'message': 'CRON_SECRET not set'}), 503
    if secret != expected:
        return jsonify({'ok': False, 'message': 'bad secret'}), 403
    kg_id = request.args.get('kg_id', '')
    total_sent = 0
    total_failed = 0
    kindergartens = pc.load_kindergartens()
    if kg_id:
        kindergartens = [kg for kg in kindergartens if kg['id'] == kg_id]
    for kg in kindergartens:
        if kg.get('status') != 'active':
            continue
        kid = kg['id']
        settings = pc.load_settings(kid)
        token = settings.get('bot_token', '')
        if not token:
            continue
        freq = settings.get('reminder_frequency', 'off')
        if freq == 'off':
            continue
        students = pc.load_json('students.json', kid)
        payments = pc.load_json('payments.json', kid)
        today = date.today()
        this_month = today.strftime('%Y-%m')
        paid_ids = {p['student_id'] for p in payments if p.get('month') == this_month and p.get('status') == 'paid'}
        message = settings.get('reminder_message', 'Hurmatli ona/ota, bolangizning to\'lov muddati yaqinlashdi.')
        for s in students:
            if s.get('status') != 'active' or not s.get('telegram_chat_id'):
                continue
            if s['id'] in paid_ids:
                continue
            chat_id = s['telegram_chat_id']
            name = f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()
            base_url = os.environ.get('SITE_URL', 'https://sofgardercrm.vercel.app').rstrip('/')
            pay_url = f"{base_url}/pay/{s['id']}/{this_month}"
            msg = message.replace('{name}', name)
            debt = s.get('monthly_fee', 0)
            msg = msg.replace('{debt}', str(debt))
            msg = msg.replace('{currency}', settings.get('currency', 'UZS'))
            msg = msg.replace('{month}', this_month)
            msg += f"\n\n🔗 <a href='{pay_url}'>To'lov qilish</a>"
            try:
                from telegram_bot import send_message
                send_message(token, chat_id, msg)
                total_sent += 1
            except Exception:
                total_failed += 1
    return jsonify({'ok': True, 'sent': total_sent, 'failed': total_failed})

app.debug = os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true')

if __name__ == '__main__':
    print("=" * 50)
    print("EduSoft — Raufbek Turaqulov")
    print("URL: http://localhost:5000")
    print("Super admin: ro'yxatdan o'tgan login bilan")
    print("O'quv markaz admini: ro'yxatdan o'tgan login bilan")
    s = load_settings()
    if s.get('bot_token'):
        print("Telegram: polling yoqilgan (bot /start ishlashi kerak)")
    else:
        print("Telegram: token yo'q — Sozlamalardan kiriting")
    print("=" * 50)
    app.run(debug=True, port=5000, use_reloader=False)

